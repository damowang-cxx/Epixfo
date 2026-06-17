from io import BytesIO
from datetime import UTC, datetime
from decimal import Decimal
from types import SimpleNamespace

import pytest
from fastapi import HTTPException
from openpyxl import Workbook

from app.models import BoxDocument, CarrierAgent, ConsigneeContact, WarehouseReceipt
from app.schemas.box import BoxCreate
from app.services.warehouse_file_service import WarehouseFileService, assert_warehouse_upload_integrity, parse_warehouse_xlsx

_ = (CarrierAgent, ConsigneeContact)


class FakeDb:
    def __init__(self) -> None:
        self.committed = False
        self.rollback_count = 0
        self.refreshed = None
        self.added = []
        self.deleted = None
        self.deleted_objects = []

    def commit(self):
        self.committed = True

    def refresh(self, obj):
        self.refreshed = obj

    def add(self, obj):
        if getattr(obj, "id", None) is None:
            obj.id = 200 + len(self.added)
        self.added.append(obj)

    def delete(self, obj):
        self.deleted = obj
        self.deleted_objects.append(obj)

    def flush(self):
        return None

    def rollback(self):
        self.rollback_count += 1


class FakeBoxRepository:
    def __init__(self) -> None:
        self.deleted_waybill_id = None
        self.document = None
        self.added_boxes = []
        self.added_items = []
        self.deleted_item_box_ids = []
        self.receipts_by_no = {}
        self.conflict_rows = []
        self.box = SimpleNamespace(
            id=4,
            current_waybill_id=7,
            warehouse_receipt_id=88,
            box_no="BOX-OLD",
            status="bound",
            is_general_cargo=False,
            never_bound_direct_upload=False,
            unbound_reason=None,
            unbound_remark=None,
            raw_data={},
            items=[],
            document=None,
            warehouse_receipt=None,
            warehouse_waybill_no=None,
            goods_name=None,
            quantity=None,
            weight=None,
            original_volume_info=None,
            original_weight_volume_ratio=None,
            volume=None,
            weight_volume_ratio=None,
            source_row_number=None,
            document_id=None,
            created_at=datetime.now(UTC),
            updated_at=datetime.now(UTC),
        )
        self.boxes_list = [self.box]

    def add_document(self, document: BoxDocument) -> BoxDocument:
        document.id = 99
        document.uploaded_at = getattr(document, "uploaded_at", None) or datetime.now(UTC)
        self.document = document
        return document

    def delete_by_waybill(self, waybill_id: int) -> int:
        self.deleted_waybill_id = waybill_id
        return 3

    def add_boxes(self, boxes):
        self.added_boxes.extend(boxes)

    def get_by_waybill(self, waybill_id: int, box_id: int):
        for box in self.boxes_list:
            if box.current_waybill_id == waybill_id and box.id == box_id:
                return box
        return None

    def get_by_id(self, box_id: int):
        return next((box for box in self.boxes_list if box.id == box_id), None)

    def get_by_box_no(self, box_no: str):
        return next((box for box in self.boxes_list if box.box_no == box_no), None)

    def list_by_waybill(self, waybill_id: int):
        return [box for box in self.boxes_list if box.current_waybill_id == waybill_id]

    def list_conflicting_boxes(self, box_nos, target_warehouse_no):
        return self.conflict_rows

    def get_receipt_by_warehouse_no(self, warehouse_no: str):
        return self.receipts_by_no.get(warehouse_no)

    def add_receipt(self, receipt: WarehouseReceipt):
        receipt.id = 88
        self.receipts_by_no[receipt.warehouse_no] = receipt
        return receipt

    def list_by_receipt_id(self, receipt_id: int):
        return [box for box in self.boxes_list if box.warehouse_receipt_id == receipt_id]

    def list_unbound_receipt_models_ordered(self):
        return [
            receipt
            for receipt in self.receipts_by_no.values()
            if receipt.waybill_id is None and getattr(receipt, "prebooking_id", None) is None
        ]

    def get_receipt_by_id(self, receipt_id: int):
        for receipt in self.receipts_by_no.values():
            if receipt.id == receipt_id:
                return receipt
        return None

    def list_by_box_nos(self, box_nos):
        box_no_set = set(box_nos)
        return [box for box in self.boxes_list if box.box_no in box_no_set]

    def list_by_ids(self, box_ids):
        box_id_set = set(box_ids)
        return [box for box in self.boxes_list if box.id in box_id_set]

    def delete_items_for_box(self, box_id: int):
        self.deleted_item_box_ids.append(box_id)
        return 1

    def add_items(self, items):
        self.added_items.extend(items)


class FakeWaybillRepository:
    def __init__(self, waybill) -> None:
        self.waybill = waybill

    def get(self, waybill_id: int):
        return self.waybill if self.waybill.id == waybill_id else None


def _xlsx_bytes(box_no: str = "BOX-001") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["外箱条码", "提单号码", "品名", "数量", "重量", "收货体积信息", "收货重量/方"])
    sheet.append([box_no, "WH-AWB-001", "Shoes", 2, 10, "40*40*40", 0.156])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _xlsx_rows(rows: list[list]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["外箱条码", "提单号码", "品名", "数量", "重量", "收货体积信息", "收货重量/方"])
    for row in rows:
        sheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _invalid_xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["外箱条码", "提单号码", "品名", "数量", "重量", "收货体积信息"])
    sheet.append(["", "WH-AWB-001", "Shoes", 2, 10, 4])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_upload_integrity_reports_missing_excel_barcode_rows() -> None:
    parse_result = parse_warehouse_xlsx(
        "CHECK-IN.xlsx",
        _xlsx_rows(
            [
                ["DHL001", "WH-AWB-001", "Shoes", 1, 10, "40*40*40", 0.064],
                ["DHL002", "WH-AWB-002", "Bags", 1, 8, "40*40*40", 0.064],
            ]
        ),
    )

    with pytest.raises(HTTPException) as exc_info:
        assert_warehouse_upload_integrity(
            file_name="CHECK-IN.xlsx",
            warehouse_no="CHECK-IN",
            parse_result=parse_result,
            uploaded_box_nos=["DHL001"],
        )

    assert exc_info.value.status_code == 400
    assert exc_info.value.detail["error_code"] == "warehouse_upload_integrity_failed"
    assert exc_info.value.detail["expected_count"] == 2
    assert exc_info.value.detail["uploaded_count"] == 1
    assert exc_info.value.detail["issues"] == [
        {
            "row_number": 3,
            "box_no": "DHL002",
            "message": "该外箱条码未成功写入系统，请检查该行是否重复、数据格式是否错误或是否被跳过。",
        }
    ]


def _unbound_service(tmp_path):
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no=None, updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    service._store_file = lambda file_name, file_hash, content: tmp_path / file_name
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])
    return service, user


def _fake_box(box_id: int, box_no: str, weight: str, volume: str, *, items_count: int = 1, original_volume_info: str | None = None):
    return SimpleNamespace(
        id=box_id,
        current_waybill_id=7,
        warehouse_receipt_id=88,
        box_no=box_no,
        status="bound",
        is_general_cargo=False,
        never_bound_direct_upload=False,
        unbound_reason=None,
        unbound_remark=None,
        raw_data={},
        items=[
            SimpleNamespace(
                id=index + 1,
                box_id=box_id,
                document_id=None,
                warehouse_waybill_no=None,
                goods_name=None,
                quantity=None,
                weight=None,
                source_row_number=None,
                raw_data={},
                created_at=datetime.now(UTC),
                updated_at=datetime.now(UTC),
            )
            for index in range(items_count)
        ],
        document=None,
        warehouse_receipt=None,
        warehouse_waybill_no=None,
        goods_name=None,
        quantity=None,
        weight=Decimal(weight),
        original_volume_info=original_volume_info,
        original_weight_volume_ratio=None,
        volume=Decimal(volume),
        weight_volume_ratio=None,
        source_row_number=None,
        document_id=None,
        created_at=datetime.now(UTC),
        updated_at=datetime.now(UTC),
    )


def test_upload_for_waybill_replaces_boxes_and_updates_warehouse_no(tmp_path) -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no=None, updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    service._store_file = lambda file_name, file_hash, content: tmp_path / file_name
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.upload_for_waybill(7, "AMS-IN-001.xlsx", _xlsx_bytes(), user)

    assert result.warehouse_no == "AMS-IN-001"
    assert result.success_count == 1
    assert result.document_id == 99
    assert waybill.warehouse_no == "AMS-IN-001"
    assert waybill.updated_by == 5
    added_box = next(item for item in service.db.added if item.__class__.__name__ == "Box")
    assert added_box.box_no == "BOX-001"
    assert added_box.current_waybill_id == 7
    assert added_box.warehouse_receipt_id == 88
    assert added_box.never_bound_direct_upload is False
    assert added_box.original_volume_info == "40*40*40"
    assert added_box.original_weight_volume_ratio == "0.156"
    assert str(added_box.volume) == "0.064"
    assert len(service.boxes.added_items) == 1
    assert service.db.committed is True


def test_upload_for_waybill_returns_integrity_warning_without_blocking(tmp_path) -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no=None, updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    service._store_file = lambda file_name, file_hash, content: tmp_path / file_name
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.upload_for_waybill(
        7,
        "AMS-IN-001.xlsx",
        _xlsx_rows(
            [
                ["DHL001", "WH-AWB-001", "Shoes", 1, 1, "40*40*40", 0.1],
                ["DHL002", "WH-AWB-002", "Shoes", 1, 1, "40*40*40", 0.1],
            ]
        ),
        user,
        skip_conflict_box_nos=["DHL002"],
    )

    assert result.success_count == 1
    assert result.integrity_issues
    assert result.integrity_issues[0].row_number == 3
    assert result.integrity_issues[0].box_no == "DHL002"
    assert service.db.committed is True


def test_upload_for_waybill_returns_english_prohibited_goods_warning_without_blocking(tmp_path) -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no=None, updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    service._store_file = lambda file_name, file_hash, content: tmp_path / file_name
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.upload_for_waybill(
        7,
        "AMS-IN-001.xlsx",
        _xlsx_rows([["DHL001", "WH-AWB-001", "Perfume gift set", 1, 1, "40*40*40", 0.1]]),
        user,
    )

    assert result.success_count == 1
    assert result.prohibited_goods_issues
    assert result.prohibited_goods_issues[0].row_number == 2
    assert result.prohibited_goods_issues[0].box_no == "DHL001"
    assert result.prohibited_goods_issues[0].keyword == "perfume"
    assert service.db.committed is True


def test_upload_unbound_file_creates_unbound_receipt_boxes(tmp_path) -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no=None, updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    service._store_file = lambda file_name, file_hash, content: tmp_path / file_name
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.upload_unbound_file("UNBOUND-IN-001.xlsx", _xlsx_bytes("DHL001"), user)

    assert result.warehouse_no == "UNBOUND-IN-001"
    assert result.success_count == 1
    assert result.document_id == 99
    assert service.boxes.document.bound_waybill_id is None
    receipt = service.boxes.receipts_by_no["UNBOUND-IN-001"]
    assert receipt.waybill_id is None
    assert receipt.prebooking_id is None
    assert receipt.source_document_id == 99
    assert receipt.channel_tags == ["AMS"]
    assert result.channel_tags == ["AMS"]
    added_box = next(item for item in service.db.added if item.__class__.__name__ == "Box")
    assert added_box.box_no == "DHL001"
    assert added_box.current_waybill_id is None
    assert added_box.warehouse_receipt_id == receipt.id
    assert added_box.status == "unbound"
    assert added_box.never_bound_direct_upload is True
    assert added_box.unbound_reason is None
    assert added_box.raw_data["source"] == "unbound_receipt_upload"
    assert len(service.boxes.added_items) == 1
    assert service.db.committed is True


def test_upload_unbound_file_returns_chinese_prohibited_goods_warning_without_blocking(tmp_path) -> None:
    service, user = _unbound_service(tmp_path)

    result = service.upload_unbound_file(
        "UNBOUND-IN-001.xlsx",
        _xlsx_rows([["DHL001", "WH-AWB-001", "香水套装", 1, 1, "40*40*40", 0.1]]),
        user,
    )

    assert result.success_count == 1
    assert result.prohibited_goods_issues
    assert result.prohibited_goods_issues[0].row_number == 2
    assert result.prohibited_goods_issues[0].goods_name == "香水套装"
    assert result.prohibited_goods_issues[0].keyword == "香水"
    assert service.db.committed is True


def test_upload_unbound_file_reuses_receipt_without_binding_elsewhere(tmp_path) -> None:
    service, user = _unbound_service(tmp_path)
    receipt = WarehouseReceipt(
        warehouse_no="UNBOUND-IN-001",
        waybill_id=None,
        prebooking_id=None,
        total_quantity=0,
        total_weight=Decimal("0.000"),
        total_volume=Decimal("0.000"),
        weight_volume_ratio=Decimal("0.000"),
    )
    receipt.id = 88
    receipt.channel_tags = []
    service.boxes.receipts_by_no[receipt.warehouse_no] = receipt

    result = service.upload_unbound_file("UNBOUND-IN-001.xlsx", _xlsx_bytes("DHL001"), user)

    assert result.warehouse_no == "UNBOUND-IN-001"
    assert receipt.waybill_id is None
    assert receipt.prebooking_id is None
    added_box = next(item for item in service.db.added if item.__class__.__name__ == "Box" and item.box_no == "DHL001")
    assert added_box.warehouse_receipt_id == receipt.id
    assert added_box.current_waybill_id is None
    assert added_box.status == "unbound"
    assert service.db.committed is True


def test_upload_unbound_file_renames_conflicting_bound_boxes(tmp_path) -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no=None, updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.boxes.box.box_no = "DHL001"
    source_receipt = SimpleNamespace(id=77, warehouse_no="AMS-IN-001", source_document=SimpleNamespace(file_name="AMS-IN-001.xlsx"))
    service.boxes.box.warehouse_receipt = source_receipt
    service.boxes.conflict_rows = [(service.boxes.box, waybill, source_receipt)]
    service.waybills = FakeWaybillRepository(waybill)
    service._store_file = lambda file_name, file_hash, content: tmp_path / file_name
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.upload_unbound_file("UNBOUND-IN-001.xlsx", _xlsx_bytes("DHL001"), user)

    assert result.success_count == 1
    added_box = next(item for item in service.db.added if item.__class__.__name__ == "Box" and item.box_no == "DHL001-DUP1")
    assert added_box.raw_data["box_conflict"]["original_box_no"] == "DHL001"
    assert added_box.raw_data["box_conflict"]["renamed_box_no"] == "DHL001-DUP1"
    assert added_box.raw_data["box_conflict"]["waybill_no"] == "784-00000001"
    assert added_box.raw_data["box_conflict"]["warehouse_no"] == "AMS-IN-001"
    assert added_box.raw_data["box_conflict"]["source_file_name"] == "AMS-IN-001.xlsx"
    assert service.boxes.document is not None
    assert service.db.committed is True


def test_upload_unbound_file_reuses_existing_conflict_rename_for_same_receipt(tmp_path) -> None:
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no=None, updated_by=None))
    service._store_file = lambda file_name, file_hash, content: tmp_path / file_name
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])
    receipt = WarehouseReceipt(
        warehouse_no="UNBOUND-IN-001",
        waybill_id=None,
        prebooking_id=None,
        total_quantity=1,
        total_weight=Decimal("10.000"),
        total_volume=Decimal("0.064"),
        weight_volume_ratio=Decimal("156.250"),
    )
    receipt.id = 88
    receipt.channel_tags = []
    service.boxes.receipts_by_no[receipt.warehouse_no] = receipt
    source_receipt = SimpleNamespace(id=77, warehouse_no="OLD-IN", source_document=SimpleNamespace(file_name="OLD-IN.xlsx"))
    service.boxes.box.box_no = "DHL001"
    service.boxes.box.warehouse_receipt_id = 77
    service.boxes.box.current_waybill_id = None
    service.boxes.conflict_rows = [(service.boxes.box, None, source_receipt)]
    reusable_box = SimpleNamespace(
        id=9,
        box_no="DHL001-DUP1",
        warehouse_receipt_id=88,
        current_waybill_id=None,
        status="unbound",
        never_bound_direct_upload=True,
        unbound_reason=None,
        unbound_remark=None,
        raw_data={
            "box_conflict": {
                "original_box_no": "DHL001",
                "renamed_box_no": "DHL001-DUP1",
                "warehouse_receipt_id": 77,
                "warehouse_no": "OLD-IN",
                "source_file_name": "OLD-IN.xlsx",
            }
        },
        items=[],
        document=None,
        warehouse_receipt=receipt,
        warehouse_waybill_no=None,
        goods_name=None,
        quantity=None,
        weight=None,
        original_volume_info=None,
        original_weight_volume_ratio=None,
        volume=None,
        weight_volume_ratio=None,
        source_row_number=None,
        document_id=None,
        created_at=datetime.now(UTC),
        updated_at=datetime.now(UTC),
    )
    service.boxes.boxes_list.append(reusable_box)

    service.upload_unbound_file("UNBOUND-IN-001.xlsx", _xlsx_bytes("DHL001"), user)

    assert reusable_box.box_no == "DHL001-DUP1"
    assert reusable_box.raw_data["box_conflict"]["original_box_no"] == "DHL001"
    assert not [item for item in service.db.added if item.__class__.__name__ == "Box"]
    assert service.db.committed is True


def test_upload_unbound_file_rejects_prebooking_bound_receipt(tmp_path) -> None:
    service, user = _unbound_service(tmp_path)
    receipt = WarehouseReceipt(
        warehouse_no="UNBOUND-IN-001",
        waybill_id=None,
        prebooking_id=3,
        total_quantity=1,
        total_weight=Decimal("1.000"),
        total_volume=Decimal("0.064"),
        weight_volume_ratio=Decimal("15.625"),
    )
    receipt.id = 88
    receipt.channel_tags = ["AMS"]
    service.boxes.receipts_by_no[receipt.warehouse_no] = receipt

    with pytest.raises(HTTPException) as exc_info:
        service.upload_unbound_file("UNBOUND-IN-001.xlsx", _xlsx_bytes("DHL001"), user)

    assert exc_info.value.detail == "warehouse_receipt_bound_to_prebooking"
    assert service.boxes.document is None
    assert service.db.committed is False


def test_upload_unbound_file_allows_uk_channel(tmp_path) -> None:
    service, user = _unbound_service(tmp_path)

    result = service.upload_unbound_file("UK-IN-001.xlsx", _xlsx_bytes("KDP001"), user)

    assert result.channel_review is not None
    assert result.channel_review.detected_channel == "uk"
    assert result.channel_review.warnings == []
    assert result.channel_tags == ["LHR"]
    assert service.db.committed is True


def test_upload_unbound_file_allows_dpd_only_with_warning(tmp_path) -> None:
    service, user = _unbound_service(tmp_path)

    result = service.upload_unbound_file(
        "DPD-IN-001.xlsx",
        _xlsx_rows(
            [
                ["DPD001", "WH-AWB-001", "Shoes", 1, 1, "40*40*40", 0.1],
                ["DPD002", "WH-AWB-002", "Shoes", 1, 1, "40*40*40", 0.1],
            ]
        ),
        user,
    )

    assert result.channel_review is not None
    assert result.channel_review.detected_channel == "unknown"
    assert result.channel_review.warnings == ["dpd_only_channel_pending"]
    assert result.channel_tags == ["AMS"]
    assert service.db.committed is True


def test_upload_unbound_file_marks_all_ctt_receipt_with_three_tags(tmp_path) -> None:
    service, user = _unbound_service(tmp_path)

    result = service.upload_unbound_file(
        "CTT-IN-001.xlsx",
        _xlsx_rows(
            [
                ["CTT001", "WH-AWB-001", "Shoes", 1, 1, "40*40*40", 0.1],
                ["CTT002", "WH-AWB-002", "Shoes", 1, 1, "40*40*40", 0.1],
            ]
        ),
        user,
    )

    assert result.channel_tags == ["MAD", "BCN", "AMS"]
    assert service.boxes.receipts_by_no["CTT-IN-001"].channel_tags == ["MAD", "BCN", "AMS"]


def test_upload_unbound_file_reports_minority_channel_boxes_as_warning_and_writes(tmp_path) -> None:
    service, user = _unbound_service(tmp_path)

    result = service.upload_unbound_file(
        "EU-IN-001.xlsx",
        _xlsx_rows(
            [
                ["DHL001", "WH-AWB-001", "Shoes", 1, 1, "40*40*40", 0.1],
                ["UPS001", "WH-AWB-002", "Shoes", 1, 1, "40*40*40", 0.1],
                ["KDP001", "WH-AWB-003", "Shoes", 1, 1, "40*40*40", 0.1],
            ]
        ),
        user,
    )

    assert result.success_count == 3
    assert result.channel_review is not None
    assert result.channel_review.detected_channel == "europe"
    assert result.channel_review.issues[0].box_no == "KDP001"
    assert result.channel_review.issues[0].reason == "uk_box_in_europe_receipt"
    assert service.db.added
    assert service.boxes.document is not None
    assert service.db.committed is True


def test_upload_unbound_file_reports_ctt_internal_mix_as_warning(tmp_path) -> None:
    service, user = _unbound_service(tmp_path)

    result = service.upload_unbound_file(
        "CTT-IN-001.xlsx",
        _xlsx_rows(
            [
                ["CTT001", "WH-AWB-001", "Shoes", 1, 1, "40*40*40", 0.1],
                ["UPS001", "WH-AWB-002", "Shoes", 1, 1, "40*40*40", 0.1],
            ]
        ),
        user,
    )

    assert result.channel_review is not None
    assert result.channel_review.issues[0].box_no == "UPS001"
    assert "ctt_mix_not_allowed" in result.channel_review.issues[0].reason
    assert service.db.added


def test_upload_unbound_file_reports_nle_internal_mix_as_warning(tmp_path) -> None:
    service, user = _unbound_service(tmp_path)

    result = service.upload_unbound_file(
        "NLE-IN-001.xlsx",
        _xlsx_rows(
            [
                ["NLE001", "WH-AWB-001", "Shoes", 1, 1, "40*40*40", 0.1],
                ["ITE001", "WH-AWB-002", "Shoes", 1, 1, "40*40*40", 0.1],
            ]
        ),
        user,
    )

    assert result.channel_review is not None
    assert result.channel_review.issues[0].box_no == "ITE001"
    assert "nle_mix_not_allowed" in result.channel_review.issues[0].reason
    assert service.db.added


def test_upload_unbound_file_reports_ups_fed_with_too_many_dhl_boxes_as_warning(tmp_path) -> None:
    service, user = _unbound_service(tmp_path)

    result = service.upload_unbound_file(
        "UPS-IN-001.xlsx",
        _xlsx_rows(
            [
                ["UPS001", "WH-AWB-001", "Shoes", 1, 1, "40*40*40", 0.1],
                ["DHL001", "WH-AWB-002", "Shoes", 1, 1, "40*40*40", 0.1],
                ["DHL002", "WH-AWB-003", "Shoes", 1, 1, "40*40*40", 0.1],
            ]
        ),
        user,
    )

    assert result.channel_review is not None
    issues = result.channel_review.issues
    assert {item.box_no for item in issues} == {"DHL001", "DHL002"}
    assert all("dhl_ratio_too_high" in item.reason for item in issues)
    assert service.db.added


def test_update_box_no_updates_bound_box() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no=None, updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    receipt = WarehouseReceipt(
        warehouse_no="SOURCE-IN",
        waybill_id=7,
        total_quantity=1,
        total_weight=Decimal("10.000"),
        total_volume=Decimal("0.064"),
        weight_volume_ratio=Decimal("156.250"),
    )
    receipt.id = 88
    receipt.channel_tags = ["AMS"]
    service.boxes.receipts_by_no[receipt.warehouse_no] = receipt
    service.boxes.box.box_no = "DHL001"
    service.waybills = FakeWaybillRepository(waybill)
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.update_box_no(7, 4, " BOX-NEW ", user)

    assert result.box_no == "BOX-NEW"
    assert service.db.committed is True
    assert service.db.refreshed is result


def test_update_box_toggles_general_cargo() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no="AMS-IN-001", updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.update_box(7, 4, user, is_general_cargo=True)

    assert result.is_general_cargo is True
    assert service.db.committed is True


def test_update_box_updates_data_fields_and_single_item() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no="AMS-IN-001", updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    receipt = WarehouseReceipt(
        warehouse_no="AMS-IN-001",
        waybill_id=7,
        total_quantity=1,
        total_weight=Decimal("10.000"),
        total_volume=Decimal("1.000"),
        weight_volume_ratio=Decimal("10.000"),
    )
    receipt.id = 88
    service.boxes.receipts_by_no[receipt.warehouse_no] = receipt
    service.boxes.box.warehouse_receipt = receipt
    service.boxes.box.warehouse_waybill_no = "OLD-AWB"
    service.boxes.box.goods_name = "Old goods"
    service.boxes.box.quantity = 1
    service.boxes.box.weight = Decimal("10.000")
    service.boxes.box.volume = Decimal("1.000")
    service.boxes.box.items = [
        SimpleNamespace(
            warehouse_waybill_no="OLD-AWB",
            goods_name="Old goods",
            quantity=1,
            weight=Decimal("10.000"),
            raw_data={},
        )
    ]
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.update_box(
        7,
        4,
        user,
        fields_set={"warehouse_waybill_no", "goods_name", "quantity", "weight", "volume"},
        warehouse_waybill_no=" NEW-AWB ",
        goods_name=" New goods ",
        quantity=3,
        weight=Decimal("6.500"),
        volume=Decimal("2.000"),
    )

    assert result.warehouse_waybill_no == "NEW-AWB"
    assert result.goods_name == "New goods"
    assert result.quantity == 3
    assert result.weight == Decimal("6.500")
    assert result.volume == Decimal("2.000")
    assert result.weight_volume_ratio == Decimal("3.250")
    assert result.items[0].warehouse_waybill_no == "NEW-AWB"
    assert result.items[0].goods_name == "New goods"
    assert result.items[0].quantity == 3
    assert result.items[0].weight == Decimal("6.500")
    assert service.db.committed is True


def test_update_unbound_receipt_box_updates_box() -> None:
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(SimpleNamespace(id=7))
    receipt = WarehouseReceipt(
        warehouse_no="UNBOUND-IN-001",
        waybill_id=None,
        prebooking_id=None,
        total_quantity=1,
        total_weight=Decimal("1.000"),
        total_volume=Decimal("1.000"),
        weight_volume_ratio=Decimal("1.000"),
    )
    receipt.id = 88
    service.boxes.receipts_by_no[receipt.warehouse_no] = receipt
    service.boxes.box.current_waybill_id = None
    service.boxes.box.status = "unbound"
    service.boxes.box.warehouse_receipt = receipt
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.update_unbound_receipt_box(
        88,
        4,
        user,
        fields_set={"goods_name", "quantity", "volume"},
        goods_name="Bag",
        quantity=8,
        volume=Decimal("1.250"),
    )

    assert result.goods_name == "Bag"
    assert result.quantity == 8
    assert result.volume == Decimal("1.250")
    assert service.db.committed is True


def test_reorder_unbound_receipts_persists_requested_order_and_appends_missing() -> None:
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    first = WarehouseReceipt(warehouse_no="IN-A", waybill_id=None, prebooking_id=None, total_quantity=0)
    first.id = 1
    second = WarehouseReceipt(warehouse_no="IN-B", waybill_id=None, prebooking_id=None, total_quantity=0)
    second.id = 2
    third = WarehouseReceipt(warehouse_no="IN-C", waybill_id=None, prebooking_id=None, total_quantity=0)
    third.id = 3
    service.boxes.receipts_by_no = {item.warehouse_no: item for item in (first, second, third)}
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    service.reorder_unbound_receipts([3, 1], user)

    assert third.display_order == 1
    assert first.display_order == 2
    assert second.display_order == 3
    assert service.db.committed is True


def test_reorder_unbound_receipts_rejects_bound_receipts() -> None:
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    unbound = WarehouseReceipt(warehouse_no="IN-A", waybill_id=None, prebooking_id=None, total_quantity=0)
    unbound.id = 1
    bound = WarehouseReceipt(warehouse_no="IN-B", waybill_id=7, prebooking_id=None, total_quantity=0)
    bound.id = 2
    service.boxes.receipts_by_no = {item.warehouse_no: item for item in (unbound, bound)}
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    with pytest.raises(HTTPException) as exc_info:
        service.reorder_unbound_receipts([2, 1], user)

    assert exc_info.value.status_code == 400
    assert exc_info.value.detail["error_code"] == "warehouse_receipt_order_invalid_receipts"
    assert exc_info.value.detail["receipt_ids"] == [2]


def test_create_box_requires_waybill_warehouse_no() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no=None, updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    with pytest.raises(HTTPException) as exc_info:
        service.create_box(7, BoxCreate(box_no="BOX-MANUAL"), user)

    assert exc_info.value.detail == "target_warehouse_no_required"


def test_create_box_binds_to_receipt_and_adds_manual_item() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no="AMS-IN-001", updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.create_box(
        7,
        BoxCreate(
            box_no=" BOX-MANUAL ",
            warehouse_waybill_no="WH-AWB-001",
            goods_name="Shoes",
            quantity=2,
            weight=Decimal("10.5"),
            volume=Decimal("3"),
            is_general_cargo=True,
        ),
        user,
    )

    assert result.box_no == "BOX-MANUAL"
    assert result.current_waybill_id == 7
    assert result.warehouse_receipt_id == 88
    assert result.status == "bound"
    assert result.is_general_cargo is True
    assert str(result.weight_volume_ratio) == "3.500"
    assert len(service.boxes.added_items) == 1
    assert service.boxes.added_items[0].box_id == result.id
    assert service.boxes.added_items[0].warehouse_waybill_no == "WH-AWB-001"
    assert service.db.committed is True


def test_delete_box_removes_box_and_commits() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no="AMS-IN-001", updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    service.boxes.box.warehouse_receipt = WarehouseReceipt(
        warehouse_no="AMS-IN-001",
        waybill_id=7,
        total_quantity=1,
        total_weight=Decimal("10.000"),
        total_volume=Decimal("3.000"),
        weight_volume_ratio=Decimal("3.333"),
    )
    service.boxes.box.warehouse_receipt.id = 88
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    service.delete_box(7, 4, user)

    assert service.db.deleted is service.boxes.box
    assert service.db.committed is True


def test_batch_delete_unbound_receipts_deletes_valid_receipts() -> None:
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    receipt_a = WarehouseReceipt(warehouse_no="WH-A", total_quantity=1)
    receipt_a.id = 101
    receipt_b = WarehouseReceipt(warehouse_no="WH-B", total_quantity=1)
    receipt_b.id = 102
    service.boxes.receipts_by_no = {receipt_a.warehouse_no: receipt_a, receipt_b.warehouse_no: receipt_b}
    box_a = SimpleNamespace(id=1, warehouse_receipt_id=101)
    box_b = SimpleNamespace(id=2, warehouse_receipt_id=102)
    service.boxes.boxes_list = [box_a, box_b]
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.batch_delete_unbound_receipts([101, 102, 101], user)

    assert result.success_count == 2
    assert result.failed_count == 0
    assert [item.id for item in result.deleted_receipts] == [101, 102]
    assert box_a in service.db.deleted_objects
    assert box_b in service.db.deleted_objects
    assert receipt_a in service.db.deleted_objects
    assert receipt_b in service.db.deleted_objects
    assert service.db.rollback_count == 0


def test_batch_delete_unbound_receipts_reports_bound_receipt_failure() -> None:
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    receipt_a = WarehouseReceipt(warehouse_no="WH-A", total_quantity=1)
    receipt_a.id = 101
    receipt_b = WarehouseReceipt(warehouse_no="WH-B", waybill_id=7, total_quantity=1)
    receipt_b.id = 102
    service.boxes.receipts_by_no = {receipt_a.warehouse_no: receipt_a, receipt_b.warehouse_no: receipt_b}
    box_a = SimpleNamespace(id=1, warehouse_receipt_id=101)
    box_b = SimpleNamespace(id=2, warehouse_receipt_id=102)
    service.boxes.boxes_list = [box_a, box_b]
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.batch_delete_unbound_receipts([101, 102], user)

    assert result.success_count == 1
    assert result.failed_count == 1
    assert result.deleted_receipts[0].id == 101
    assert result.errors[0].id == 102
    assert result.errors[0].message == "warehouse_receipt_bound_to_waybill"
    assert box_a in service.db.deleted_objects
    assert receipt_a in service.db.deleted_objects
    assert box_b not in service.db.deleted_objects
    assert receipt_b not in service.db.deleted_objects
    assert service.db.rollback_count == 1


def test_recalculate_box_volumes_scales_single_item_boxes_to_target_volume() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no="AMS-IN-001")
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    service.boxes.boxes_list = [
        _fake_box(4, "BOX-001", "10.000", "6.000", original_volume_info="100*100*600"),
        _fake_box(5, "BOX-002", "5.000", "4.000", items_count=2, original_volume_info="100*100*400"),
    ]
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.recalculate_box_volumes(7, Decimal("9.500"), user)

    assert result.adjusted is True
    assert str(result.old_total_volume) == "10.000"
    assert str(result.original_total_volume) == "10.000"
    assert str(result.fixed_total_volume) == "4.000"
    assert str(result.adjustable_total_volume) == "6.000"
    assert str(result.new_total_volume) == "9.500"
    assert [str(box.volume) for box in service.boxes.boxes_list] == ["5.500", "4.000"]
    assert str(service.boxes.boxes_list[0].weight_volume_ratio) == "1.818"
    assert service.boxes.boxes_list[0].raw_data["volume_recalculation"]["target_volume"] == "9.500"
    assert service.boxes.boxes_list[0].raw_data["volume_recalculation"]["base_volume"] == "6.000"
    assert service.boxes.boxes_list[0].raw_data["volume_recalculation"]["calculated_volume_info"].endswith("(5.5)")
    assert "*" in service.boxes.boxes_list[0].raw_data["volume_recalculation"]["calculated_volume_info"]
    assert service.boxes.boxes_list[1].raw_data["volume_recalculation"]["adjustable"] is False
    assert service.boxes.boxes_list[1].raw_data["volume_recalculation"]["calculated_volume_info"] == "100*100*400(4)"
    assert service.db.committed is True

    service.db.committed = False
    result = service.recalculate_box_volumes(7, Decimal("10.000"), user)

    assert str(result.new_total_volume) == "10.000"
    assert [str(box.volume) for box in service.boxes.boxes_list] == ["6.000", "4.000"]
    assert service.db.committed is True


def test_recalculate_box_volumes_rejects_target_below_multi_item_fixed_volume() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no="AMS-IN-001")
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    service.boxes.boxes_list = [
        _fake_box(4, "BOX-001", "10.000", "6.000"),
        _fake_box(5, "BOX-002", "5.000", "4.000", items_count=2),
    ]
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    with pytest.raises(HTTPException) as exc_info:
        service.recalculate_box_volumes(7, Decimal("3.500"), user)

    assert exc_info.value.detail["error_code"] == "target_volume_less_than_fixed_boxes"
    assert exc_info.value.detail["fixed_total_volume"] == "4.000"
    assert service.db.committed is False


def test_recalculate_unbound_receipt_box_volumes_scales_receipt_boxes() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no="AMS-IN-001")
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    receipt = WarehouseReceipt(
        warehouse_no="UNBOUND-IN-001",
        waybill_id=None,
        prebooking_id=None,
        total_quantity=0,
        total_weight=Decimal("15.000"),
        total_volume=Decimal("8.000"),
        weight_volume_ratio=Decimal("1.875"),
    )
    receipt.id = 88
    receipt.channel_tags = []
    service.boxes.receipts_by_no[receipt.warehouse_no] = receipt
    service.boxes.boxes_list = [
        _fake_box(4, "DHL001", "10.000", "5.000"),
        _fake_box(5, "DHL002", "5.000", "3.000", items_count=2),
    ]
    for box in service.boxes.boxes_list:
        box.current_waybill_id = None
        box.status = "unbound"
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.recalculate_unbound_receipt_box_volumes(88, Decimal("7.000"), user)

    assert result.adjusted is True
    assert str(result.old_total_volume) == "8.000"
    assert str(result.fixed_total_volume) == "3.000"
    assert str(result.new_total_volume) == "7.000"
    assert [str(box.volume) for box in service.boxes.boxes_list] == ["4.000", "3.000"]
    assert str(receipt.total_volume) == "7.000"
    assert str(receipt.weight_volume_ratio) == "2.143"
    assert service.db.committed is True


def test_recalculate_unbound_receipt_box_volumes_rejects_bound_receipt() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no="AMS-IN-001")
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    receipt = WarehouseReceipt(
        warehouse_no="BOUND-IN-001",
        waybill_id=7,
        prebooking_id=None,
        total_quantity=0,
        total_weight=Decimal("10.000"),
        total_volume=Decimal("5.000"),
        weight_volume_ratio=Decimal("2.000"),
    )
    receipt.id = 88
    service.boxes.receipts_by_no[receipt.warehouse_no] = receipt
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    with pytest.raises(HTTPException) as exc_info:
        service.recalculate_unbound_receipt_box_volumes(88, Decimal("7.000"), user)

    assert exc_info.value.detail == "warehouse_receipt_not_unbound"
    assert service.db.committed is False


def test_upload_reports_failed_rows_when_no_valid_rows(tmp_path) -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no=None, updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    service._store_file = lambda file_name, file_hash, content: tmp_path / file_name
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    with pytest.raises(HTTPException) as exc_info:
        service.upload_for_waybill(7, "AMS-IN-001.xlsx", _invalid_xlsx_bytes(), user)

    assert "没有有效货物行" in exc_info.value.detail
    assert "第 2 行" in exc_info.value.detail


def test_upload_renames_box_conflicts_and_writes_conflict_metadata(tmp_path) -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no=None, updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    service._store_file = lambda file_name, file_hash, content: tmp_path / file_name
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])
    current_waybill = SimpleNamespace(id=8, waybill_no="999-00000001")
    current_receipt = SimpleNamespace(id=77, warehouse_no="OLD-IN", source_document=SimpleNamespace(file_name="OLD-IN.xlsx"))
    service.boxes.conflict_rows = [(SimpleNamespace(box_no="BOX-001"), current_waybill, current_receipt)]

    result = service.upload_for_waybill(7, "AMS-IN-001.xlsx", _xlsx_bytes(), user)

    assert result.success_count == 1
    added_box = next(item for item in service.db.added if item.__class__.__name__ == "Box" and item.box_no == "BOX-001-DUP1")
    assert added_box.current_waybill_id == 7
    assert added_box.raw_data["box_conflict"]["original_box_no"] == "BOX-001"
    assert added_box.raw_data["box_conflict"]["renamed_box_no"] == "BOX-001-DUP1"
    assert added_box.raw_data["box_conflict"]["waybill_no"] == "999-00000001"
    assert added_box.raw_data["box_conflict"]["warehouse_no"] == "OLD-IN"
    assert added_box.raw_data["box_conflict"]["source_file_name"] == "OLD-IN.xlsx"
    assert service.db.committed is True


def test_upload_for_prebooking_renames_conflicting_boxes(tmp_path) -> None:
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no=None, updated_by=None))
    service._store_file = lambda file_name, file_hash, content: tmp_path / file_name
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])
    prebooking = SimpleNamespace(id=31, status="draft", updated_by=None)
    source_receipt = SimpleNamespace(id=77, warehouse_no="OLD-IN", source_document=SimpleNamespace(file_name="OLD-IN.xlsx"))
    service.boxes.box.box_no = "BOX-001"
    service.boxes.box.warehouse_receipt_id = 77
    service.boxes.box.current_waybill_id = None
    service.boxes.conflict_rows = [(service.boxes.box, None, source_receipt)]

    result = service.upload_for_prebooking(prebooking, "PRE-IN-001.xlsx", _xlsx_bytes(), user)

    assert result.success_count == 1
    added_box = next(item for item in service.db.added if item.__class__.__name__ == "Box" and item.box_no == "BOX-001-DUP1")
    assert added_box.status == "prebooked"
    assert added_box.raw_data["box_conflict"]["original_box_no"] == "BOX-001"
    assert added_box.raw_data["box_conflict"]["warehouse_no"] == "OLD-IN"
    assert prebooking.updated_by == 5
    assert service.db.committed is True


def test_batch_bind_requires_target_waybill_warehouse_no() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no=None, updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    with pytest.raises(HTTPException) as exc_info:
        service.batch_bind_boxes([4], 7, user)

    assert exc_info.value.detail == "target_warehouse_no_required"


def test_batch_bind_moves_boxes_to_target_receipt() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no="AMS-IN-001", updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.waybills = FakeWaybillRepository(waybill)
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.batch_bind_boxes([4], 7, user)

    assert result.updated_count == 1
    assert service.boxes.box.current_waybill_id == 7
    assert service.boxes.box.warehouse_receipt_id == 88
    assert service.boxes.box.status == "bound"
    assert service.boxes.box.never_bound_direct_upload is False
    assert service.boxes.box.unbound_reason is None
    assert service.boxes.box.unbound_remark is None
    assert service.db.committed is True


def test_batch_transfer_to_unbound_records_reason_and_remark() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no="AMS-IN-001", updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    receipt = WarehouseReceipt(
        warehouse_no="SOURCE-IN",
        waybill_id=7,
        total_quantity=1,
        total_weight=Decimal("10.000"),
        total_volume=Decimal("0.064"),
        weight_volume_ratio=Decimal("156.250"),
    )
    receipt.id = 88
    receipt.channel_tags = ["AMS"]
    service.boxes.receipts_by_no[receipt.warehouse_no] = receipt
    service.boxes.box.box_no = "DHL001"
    service.waybills = FakeWaybillRepository(waybill)
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.batch_transfer_boxes(
        [4],
        "unbound",
        user,
        unbound_reason="customs_inspection",
        unbound_remark=" 海关开箱查验 ",
    )

    assert result.updated_count == 1
    assert service.boxes.box.current_waybill_id is None
    assert service.boxes.box.warehouse_receipt_id is None
    assert service.boxes.box.status == "unbound"
    assert service.boxes.box.unbound_reason == "customs_inspection"
    assert service.boxes.box.unbound_remark == "海关开箱查验"
    assert receipt.channel_tags == []
    assert service.db.committed is True


def test_batch_transfer_to_waybill_clears_unbound_reason() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no="AMS-IN-001", updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.boxes.box.current_waybill_id = None
    service.boxes.box.warehouse_receipt_id = None
    service.boxes.box.status = "unbound"
    service.boxes.box.never_bound_direct_upload = True
    service.boxes.box.unbound_reason = "other"
    service.boxes.box.unbound_remark = "待确认"
    service.waybills = FakeWaybillRepository(waybill)
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.batch_transfer_boxes([4], "waybill", user, target_waybill_id=7)

    assert result.updated_count == 1
    assert service.boxes.box.current_waybill_id == 7
    assert service.boxes.box.warehouse_receipt_id == 88
    assert service.boxes.box.status == "bound"
    assert service.boxes.box.never_bound_direct_upload is False
    assert service.boxes.box.unbound_reason is None
    assert service.boxes.box.unbound_remark is None
    assert service.db.committed is True


def test_bind_unbound_receipt_to_waybill_appends_receipt() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no="OLD-IN", updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.boxes.box.box_no = "KDP001"
    receipt = WarehouseReceipt(
        warehouse_no="UNBOUND-IN-001",
        waybill_id=None,
        total_quantity=1,
        total_weight=Decimal("10.000"),
        total_volume=Decimal("0.064"),
        weight_volume_ratio=Decimal("156.250"),
    )
    receipt.id = 88
    receipt.display_order = 9
    service.boxes.receipts_by_no[receipt.warehouse_no] = receipt
    service.boxes.box.current_waybill_id = None
    service.boxes.box.warehouse_receipt_id = 88
    service.boxes.box.status = "unbound"
    service.waybills = FakeWaybillRepository(waybill)
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.bind_receipt_to_waybill(88, 7, user)

    assert result.waybill_id == 7
    assert result.display_order is None
    assert service.boxes.box.current_waybill_id == 7
    assert service.boxes.box.status == "bound"
    assert service.boxes.box.unbound_reason is None
    assert waybill.warehouse_no == "UNBOUND-IN-001"
    assert service.db.committed is True


def test_batch_transfer_to_receipt_moves_boxes_without_waybill() -> None:
    waybill = SimpleNamespace(id=7, waybill_no="784-00000001", warehouse_no="AMS-IN-001", updated_by=None)
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.boxes.box.box_no = "KDP001"
    receipt = WarehouseReceipt(
        warehouse_no="TARGET-IN",
        waybill_id=None,
        total_quantity=0,
        total_weight=Decimal("0.000"),
        total_volume=Decimal("0.000"),
        weight_volume_ratio=Decimal("0.000"),
    )
    receipt.id = 90
    service.boxes.receipts_by_no[receipt.warehouse_no] = receipt
    service.boxes.box.current_waybill_id = None
    service.boxes.box.warehouse_receipt_id = None
    service.boxes.box.status = "unbound"
    service.boxes.box.unbound_reason = "other"
    service.waybills = FakeWaybillRepository(waybill)
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.batch_transfer_boxes([4], "receipt", user, target_receipt_id=90)

    assert result.updated_count == 1
    assert service.boxes.box.warehouse_receipt_id == 90
    assert service.boxes.box.current_waybill_id is None
    assert service.boxes.box.status == "unbound"
    assert service.boxes.box.unbound_reason is None
    assert receipt.channel_tags == ["LHR"]
    assert service.db.committed is True


def test_batch_transfer_to_prebooking_receipt_marks_boxes_prebooked() -> None:
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.boxes.box.box_no = "DHL001"
    receipt = WarehouseReceipt(
        warehouse_no="PREBOOKING-IN",
        waybill_id=None,
        prebooking_id=31,
        total_quantity=0,
        total_weight=Decimal("0.000"),
        total_volume=Decimal("0.000"),
        weight_volume_ratio=Decimal("0.000"),
    )
    receipt.id = 90
    service.boxes.receipts_by_no[receipt.warehouse_no] = receipt
    service.boxes.box.current_waybill_id = None
    service.boxes.box.warehouse_receipt_id = None
    service.boxes.box.status = "unbound"
    service.boxes.box.unbound_reason = "other"
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.batch_transfer_boxes([4], "receipt", user, target_receipt_id=90)

    assert result.updated_count == 1
    assert service.boxes.box.warehouse_receipt_id == 90
    assert service.boxes.box.current_waybill_id is None
    assert service.boxes.box.status == "prebooked"
    assert service.boxes.box.unbound_reason is None
    assert receipt.channel_tags == ["AMS"]
    assert service.db.committed is True


def test_batch_transfer_to_waybill_receipt_marks_boxes_bound() -> None:
    service = WarehouseFileService.__new__(WarehouseFileService)
    service.db = FakeDb()
    service.boxes = FakeBoxRepository()
    service.boxes.box.box_no = "DHL001"
    receipt = WarehouseReceipt(
        warehouse_no="BOUND-IN",
        waybill_id=7,
        prebooking_id=None,
        total_quantity=0,
        total_weight=Decimal("0.000"),
        total_volume=Decimal("0.000"),
        weight_volume_ratio=Decimal("0.000"),
    )
    receipt.id = 90
    service.boxes.receipts_by_no[receipt.warehouse_no] = receipt
    service.boxes.box.current_waybill_id = None
    service.boxes.box.warehouse_receipt_id = None
    service.boxes.box.status = "unbound"
    service.boxes.box.unbound_reason = "customs_inspection"
    user = SimpleNamespace(id=5, is_superuser=True, roles=[])

    result = service.batch_transfer_boxes([4], "receipt", user, target_receipt_id=90)

    assert result.updated_count == 1
    assert service.boxes.box.warehouse_receipt_id == 90
    assert service.boxes.box.current_waybill_id == 7
    assert service.boxes.box.status == "bound"
    assert service.boxes.box.unbound_reason is None
    assert receipt.channel_tags == ["AMS"]
    assert service.db.committed is True

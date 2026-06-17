from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.api.v1.waybills import general_cargo_export
from app.core.exceptions import AppHTTPException
from app.models.enums import UserRoleCode
from app.services.customs_export_service import CustomsExportService


class FakeBoxRepository:
    def __init__(self, boxes):
        self._boxes = boxes

    def list_by_waybill(self, waybill_id: int):
        return self._boxes


def _make_service(boxes):
    service = CustomsExportService.__new__(CustomsExportService)
    service.boxes = FakeBoxRepository(boxes)
    return service


def _make_monthly_service(rows):
    service = CustomsExportService.__new__(CustomsExportService)
    service._list_monthly_general_cargo_rows = lambda year, month: rows
    return service


def _make_waybill(**overrides):
    defaults = {
        "id": 7,
        "waybill_no": "176-29600664",
        "warehouse_data_remark": "入国航1号货站，喜提达打板截单时间：21号 20:00",
        "consignee": None,
        "consignee_contact": SimpleNamespace(
            name="ALLINE BV",
            address="Capronilaan 37 Schiphol-Rijk the Netherlands 1119 NG",
            email="import@alline.global",
            phone="+31 629 809 005",
            tax_info="VAT number: NL863476971B01",
            notify_party=None,
        ),
        "plan": SimpleNamespace(
            planned_flight_no="EK9871/22",
            planned_flight_date=date(2026, 5, 22),
            planned_route_text="CAN- DWC - AMS",
        ),
    }
    defaults.update(overrides)
    return SimpleNamespace(**defaults)


def _make_receipt(**overrides):
    defaults = {
        "warehouse_no": "WH-001",
        "source_document": SimpleNamespace(file_name="WH-001.xlsx"),
    }
    defaults.update(overrides)
    return SimpleNamespace(**defaults)


def _load_export(workbook_bytes: bytes):
    return load_workbook(BytesIO(workbook_bytes))


def test_customs_export_includes_two_sheets_and_single_box_row() -> None:
    box = SimpleNamespace(
        box_no="DHLAE57542",
        warehouse_waybill_no="CP147969890DE",
        goods_name="SPORTS SHOES",
        quantity=10,
        weight=Decimal("13.500"),
        original_volume_info="50*40*40",
        volume=Decimal("0.080"),
        weight_volume_ratio=Decimal("168.750"),
        items=[],
    )
    workbook = _load_export(_make_service([box]).build_waybill_export(_make_waybill()))

    assert workbook.sheetnames == ["入仓数据", "提单资料"]
    inbound = workbook["入仓数据"]
    assert [cell.value for cell in inbound[1]] == ["外箱条码", "提单号码", "品名", "数量", "重量", "收货体积信息", "收货重量/方"]
    assert [cell.value for cell in inbound[2]] == [
        "DHLAE57542",
        "CP147969890DE",
        "SPORTS SHOES",
        10,
        "13.5",
        "50*40*40",
        "0.080",
    ]
    assert [cell.value for cell in inbound[3]][3:7] == ["合计", "13.5", "168.750", "0.080"]
    for cell in inbound[3][3:7]:
        assert cell.fill.fill_type == "solid"
        assert cell.fill.fgColor.rgb.endswith("FFF2CC")

    waybill_sheet = workbook["提单资料"]
    rows = {row[0].value: row[1].value for row in waybill_sheet.iter_rows(min_row=1, max_col=2)}
    assert rows["提单号"] == "176-29600664"
    assert rows["发货人"] is None
    assert rows["品名"] is None
    assert "ALLINE BV" in rows["收货人"]
    assert "EK9871/22MAY" in rows["航班号截单时间隐含签名仓库/打板交单地址"]
    assert "EK9871/22/22MAY" not in rows["航班号截单时间隐含签名仓库/打板交单地址"]
    assert "航程：CAN- DWC - AMS" in rows["航班号截单时间隐含签名仓库/打板交单地址"]
    assert "喜提达打板截单时间" in rows["航班号截单时间隐含签名仓库/打板交单地址"]


def test_customs_export_expands_multi_item_box_with_blank_repeated_box_fields() -> None:
    box = SimpleNamespace(
        box_no="DHLAE57698",
        warehouse_waybill_no="CP147989086DE",
        goods_name="衣服",
        quantity=3,
        weight=Decimal("1.770"),
        volume=Decimal("0.150"),
        weight_volume_ratio=Decimal("11.800"),
        items=[
            SimpleNamespace(warehouse_waybill_no="CP147989086DE", goods_name="衣服", quantity=1, weight=Decimal("0.590")),
            SimpleNamespace(warehouse_waybill_no="CP147905449DE", goods_name="鞋", quantity=1, weight=Decimal("0.590")),
        ],
    )
    workbook = _load_export(_make_service([box]).build_waybill_export(_make_waybill()))
    inbound = workbook["入仓数据"]

    assert [cell.value for cell in inbound[2]] == ["DHLAE57698", "CP147989086DE", "衣服", 1, "0.59", "0.150", "0.150"]
    assert [cell.value for cell in inbound[3]] == [None, "CP147905449DE", "鞋", 1, "0.59", None, None]
    assert [cell.value for cell in inbound[4]][3:7] == ["合计", "1.18", "7.867", "0.150"]


def test_customs_export_moves_general_cargo_below_regular_rows_and_highlights() -> None:
    general_box = SimpleNamespace(
        box_no="BOX-GENERAL",
        warehouse_waybill_no="WH-GENERAL-1",
        goods_name="GENERAL GOODS",
        quantity=2,
        weight=Decimal("5.000"),
        volume=Decimal("0.200"),
        weight_volume_ratio=Decimal("25.000"),
        is_general_cargo=True,
        items=[
            SimpleNamespace(warehouse_waybill_no="WH-GENERAL-1", goods_name="GENERAL A", quantity=1, weight=Decimal("2.000")),
            SimpleNamespace(warehouse_waybill_no="WH-GENERAL-2", goods_name="GENERAL B", quantity=1, weight=Decimal("3.000")),
        ],
    )
    regular_box = SimpleNamespace(
        box_no="BOX-REGULAR",
        warehouse_waybill_no="WH-REGULAR-1",
        goods_name="REGULAR GOODS",
        quantity=1,
        weight=Decimal("10.000"),
        volume=Decimal("0.100"),
        weight_volume_ratio=Decimal("100.000"),
        is_general_cargo=False,
        items=[],
    )

    workbook = _load_export(_make_service([general_box, regular_box]).build_waybill_export(_make_waybill()))
    inbound = workbook.active

    assert inbound[2][0].value == "BOX-REGULAR"
    assert inbound[3][0].value == "BOX-GENERAL"
    assert inbound[4][0].value is None
    for row_number in (3, 4):
        for cell in inbound[row_number][:7]:
            assert cell.fill.fill_type == "solid"
            assert cell.fill.fgColor.rgb.endswith("FFF2CC")
    assert [cell.value for cell in inbound[5]][4:7] == ["15", "50.000", "0.300"]


def test_customs_export_uses_calculated_dimensions_without_cbm_suffix() -> None:
    box = SimpleNamespace(
        box_no="BOX-CALC",
        warehouse_waybill_no="WH-CALC-1",
        goods_name="CALCULATED GOODS",
        quantity=1,
        weight=Decimal("9.000"),
        original_volume_info="50*40*40",
        volume=Decimal("0.076"),
        weight_volume_ratio=Decimal("118.421"),
        raw_data={
            "volume_recalculation": {
                "calculated_volume_info": "49.15*39.32*39.32(0.076)",
            }
        },
        items=[],
    )

    workbook = _load_export(_make_service([box]).build_waybill_export(_make_waybill()))
    inbound = workbook.active

    assert inbound[2][5].value == "49.15*39.32*39.32"
    assert inbound[2][6].value == "0.076"
    assert inbound[3][5].value == "118.421"
    assert inbound[3][6].value == "0.076"


def test_monthly_general_cargo_export_includes_source_columns_and_summary() -> None:
    waybill = _make_waybill(waybill_no="176-00000001")
    receipt = _make_receipt()
    box = SimpleNamespace(
        box_no="BOX-GENERAL",
        warehouse_waybill_no="WH-GENERAL-1",
        goods_name="GENERAL GOODS",
        quantity=2,
        weight=Decimal("5.000"),
        original_volume_info="50*40*40",
        volume=Decimal("0.200"),
        weight_volume_ratio=Decimal("25.000"),
        is_general_cargo=True,
        items=[
            SimpleNamespace(warehouse_waybill_no="WH-GENERAL-1", goods_name="GENERAL A", quantity=1, weight=Decimal("2.000")),
            SimpleNamespace(warehouse_waybill_no="WH-GENERAL-2", goods_name="GENERAL B", quantity=1, weight=Decimal("3.000")),
        ],
    )

    workbook = _load_export(_make_monthly_service([(waybill, receipt, box)]).build_monthly_general_cargo_export(2026, 6))
    sheet = workbook["普货汇总"]

    assert workbook.sheetnames == ["普货汇总"]
    assert [cell.value for cell in sheet[1]] == [
        "主提单号",
        "入仓号文件",
        "外箱条码",
        "运单号",
        "品名",
        "数量",
        "重量",
        "收货体积信息",
        "收货重量/方",
    ]
    assert [cell.value for cell in sheet[2]] == [
        "176-00000001",
        "WH-001.xlsx",
        "BOX-GENERAL",
        "WH-GENERAL-1",
        "GENERAL A",
        1,
        "2",
        "50*40*40",
        "0.200",
    ]
    assert [cell.value for cell in sheet[3]] == [
        "176-00000001",
        "WH-001.xlsx",
        None,
        "WH-GENERAL-2",
        "GENERAL B",
        1,
        "3",
        None,
        None,
    ]
    assert [cell.value for cell in sheet[4]][5:9] == ["合计", "5", "25.000", "0.200"]
    for cell in sheet[4][5:9]:
        assert cell.fill.fill_type == "solid"
        assert cell.fill.fgColor.rgb.endswith("FFF2CC")


def test_monthly_general_cargo_export_empty_month_still_has_summary() -> None:
    workbook = _load_export(_make_monthly_service([]).build_monthly_general_cargo_export(2026, 6))
    sheet = workbook["普货汇总"]

    assert sheet.max_row == 2
    assert [cell.value for cell in sheet[2]][5:9] == ["合计", "0", "0.000", "0.000"]


def test_general_cargo_export_route_rejects_non_route_roles() -> None:
    user = SimpleNamespace(is_superuser=False, roles=[SimpleNamespace(code=UserRoleCode.CUSTOMER_SERVICE)])

    with pytest.raises(AppHTTPException) as exc_info:
        general_cargo_export(year=2026, month=6, current_user=user, db=SimpleNamespace())

    assert exc_info.value.status_code == 403


def test_customs_export_outputs_notify_party_only_when_different() -> None:
    same_notify = SimpleNamespace(
        name="ALLINE BV",
        address="Capronilaan 37 Schiphol-Rijk the Netherlands 1119 NG",
        email="import@alline.global",
        phone="+31 629 809 005",
        tax_info="VAT number: NL863476971B01",
        enabled=True,
    )
    contact = _make_waybill().consignee_contact
    contact.notify_party = same_notify
    workbook = _load_export(_make_service([]).build_waybill_export(_make_waybill(consignee_contact=contact)))
    labels = [row[0].value for row in workbook["提单资料"].iter_rows(min_row=1, max_col=2)]
    assert "通知人" not in labels

    different_contact = _make_waybill().consignee_contact
    different_contact.notify_party = SimpleNamespace(
        name="Notify BV",
        address="Other address",
        email="notify@example.com",
        phone="123",
        tax_info="EORI:NL123",
        enabled=True,
    )
    workbook = _load_export(_make_service([]).build_waybill_export(_make_waybill(consignee_contact=different_contact)))
    rows = {row[0].value: row[1].value for row in workbook["提单资料"].iter_rows(min_row=1, max_col=2)}
    assert "通知人" in rows
    assert "Notify BV" in rows["通知人"]

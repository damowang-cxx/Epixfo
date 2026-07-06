from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.core.exceptions import bad_request
from app.models import CarrierAgent, Consignee, ConsigneeContact, User
from app.schemas.waybill import (
    WaybillBulkImportCreated,
    WaybillBulkImportError,
    WaybillBulkImportResult,
    WaybillCreate,
)
from app.schemas.warehouse_planner import WarehousePlannerRow
from app.services.permission_service import PermissionService
from app.services.waybill_service import WaybillService
from app.utils.datetime_utils import local_now
from app.utils.planned_flight import parse_planned_flight_info


@dataclass
class ParsedWaybillImportRow:
    row_number: int
    payload: WaybillCreate | None = None
    waybill_no: str | None = None
    error: str | None = None


@dataclass
class ParsedWaybillImportFile:
    rows: list[ParsedWaybillImportRow] = field(default_factory=list)
    skipped_count: int = 0


@dataclass
class ParsedPlannerImportWarning:
    row_number: int
    field: str
    raw_value: str | None
    message: str


@dataclass
class ParsedPlannerImportError:
    row_number: int | None
    waybill_no: str | None
    message: str


@dataclass
class ParsedPlannerImportFile:
    rows: list[WarehousePlannerRow] = field(default_factory=list)
    skipped_count: int = 0
    warnings: list[ParsedPlannerImportWarning] = field(default_factory=list)
    errors: list[ParsedPlannerImportError] = field(default_factory=list)


@dataclass
class ParsedPlannerBoardGroup:
    group_id: str
    order: int
    booked_volume: Decimal | None
    booked_weight: Decimal | None


class PlannerWorksheetMergeContext:
    def __init__(self, worksheet, header_map: dict[str, int], header_row_number: int, source_id_base: int) -> None:
        self.worksheet = worksheet
        self.header_map = header_map
        self.header_row_number = header_row_number
        self.merged_parent: dict[tuple[int, int], tuple[int, int]] = {}
        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.max_row <= header_row_number:
                continue
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for column in range(merged_range.min_col, merged_range.max_col + 1):
                    self.merged_parent[(row, column)] = (merged_range.min_row, merged_range.min_col)
        self.board_groups_by_row = self._collect_board_groups(source_id_base)

    def row_values(self, row_number: int, header_count: int) -> list[Any]:
        return [self.value_by_index(row_number, index) for index in range(header_count)]

    def value_by_index(self, row_number: int, index: int) -> Any:
        column_number = index + 1
        parent_row, parent_column = self.merged_parent.get((row_number, column_number), (row_number, column_number))
        return self.worksheet.cell(parent_row, parent_column).value

    def value(self, row_number: int, header: str) -> Any:
        index = self.header_map.get(header)
        if index is None:
            return None
        return self.value_by_index(row_number, index)

    def board_group_for_row(self, row_number: int) -> ParsedPlannerBoardGroup | None:
        return self.board_groups_by_row.get(row_number)

    def _collect_board_groups(self, source_id_base: int) -> dict[int, ParsedPlannerBoardGroup]:
        volume_column = self.header_map.get("方数")
        if volume_column is None:
            return {}
        volume_column += 1
        weight_column = self.header_map.get("订舱重量")
        weight_column = weight_column + 1 if weight_column is not None else None
        groups: dict[int, ParsedPlannerBoardGroup] = {}
        for merged_range in self.worksheet.merged_cells.ranges:
            if merged_range.max_row <= self.header_row_number or merged_range.min_row == merged_range.max_row:
                continue
            if not (merged_range.min_col <= volume_column <= merged_range.max_col):
                continue
            booked_weight: Decimal | None = None
            if weight_column is not None:
                for weight_range in self.worksheet.merged_cells.ranges:
                    if (
                        weight_range.min_row == merged_range.min_row
                        and weight_range.max_row == merged_range.max_row
                        and weight_range.min_col <= weight_column <= weight_range.max_col
                    ):
                        booked_weight = _decimal_or_none(self.value(merged_range.min_row, "订舱重量"))
                        break
            group = ParsedPlannerBoardGroup(
                group_id=f"import-board-{source_id_base}-{merged_range.min_row}-{merged_range.max_row}",
                order=merged_range.min_row,
                booked_volume=_decimal_or_none(self.value(merged_range.min_row, "方数")),
                booked_weight=booked_weight,
            )
            for row_number in range(merged_range.min_row, merged_range.max_row + 1):
                groups[row_number] = group
        return groups


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def _normalize_lookup(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
    elif isinstance(value, float) and value.is_integer():
        cleaned = str(int(value))
    else:
        cleaned = str(value).strip()
    return cleaned or None


def _decimal_or_none(value: Any) -> Decimal | None:
    text = _clean_text(value)
    if not text:
        return None
    ratio_match = re.fullmatch(r"1\s*[:/]\s*([-+]?\d+(\.\d+)?)", text)
    if ratio_match:
        text = ratio_match.group(1)
    text = text.replace(",", "")
    if not re.fullmatch(r"[-+]?\d+(\.\d+)?", text):
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _date_or_none(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean_text(value)
    if not text:
        return None
    current_year = local_now().year
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d", "%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt)
        except ValueError:
            continue
        if "%Y" not in fmt:
            parsed = parsed.replace(year=current_year)
        return parsed.date()
    return None


def _datetime_or_none(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = _clean_text(value)
    if not text:
        return None
    current_year = local_now().year
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%m/%d %H:%M",
        "%m-%d %H:%M",
    ):
        try:
            parsed = datetime.strptime(text, fmt)
        except ValueError:
            continue
        if "%Y" not in fmt:
            parsed = parsed.replace(year=current_year)
        return parsed
    parsed_date = _date_or_none(text)
    return datetime.combine(parsed_date, time.min) if parsed_date else None


def _bool_or_none(value: Any) -> bool | None:
    text = _clean_text(value)
    if not text:
        return None
    lowered = re.sub(r"\s+", "", text).lower()
    if lowered in {"0", "false", "no", "n", "否", "不通知", "无", "不含tc", "不含t", "notc", "no-tc"}:
        return False
    if lowered in {"1", "true", "yes", "y", "是", "已通知", "通知", "含tc", "含t", "tc"}:
        return True
    if "不含tc" in lowered or "不含t" in lowered or "notc" in lowered or "no-tc" in lowered:
        return False
    if "含tc" in lowered or lowered == "含":
        return True
    return None


def _join_parts(parts: list[str]) -> str | None:
    cleaned = [part for part in parts if part]
    return "\n".join(cleaned) if cleaned else None


class WaybillImportTemplateParser:
    def __init__(
        self,
        *,
        agents_by_name: dict[str, int] | None = None,
        consignees_by_name: dict[str, int | None] | None = None,
        users_by_name: dict[str, int] | None = None,
    ) -> None:
        self.agents_by_name = agents_by_name or {}
        self.consignees_by_name = consignees_by_name or {}
        self.users_by_name = users_by_name or {}

    def parse(self, content: bytes) -> ParsedWaybillImportFile:
        try:
            workbook = load_workbook(BytesIO(content), data_only=True)
        except Exception as exc:
            raise bad_request("invalid_waybill_import_xlsx") from exc
        worksheet = workbook.worksheets[0]
        header_row_number = self._find_header_row(worksheet)
        if header_row_number is None:
            raise bad_request("waybill_import_header_not_found")

        headers = [_normalize_header(cell.value) for cell in worksheet[header_row_number]]
        header_map = {header: index for index, header in enumerate(headers) if header}
        if "提单号" not in header_map:
            raise bad_request("waybill_import_header_not_found")

        parsed = ParsedWaybillImportFile()
        for row_number in range(header_row_number + 1, worksheet.max_row + 1):
            values = [cell.value for cell in worksheet[row_number]]
            if not any(_clean_text(value) for value in values):
                parsed.skipped_count += 1
                continue
            parsed.rows.append(self._parse_row(row_number, values, header_map, headers))
        return parsed

    def parse_planner(self, content: bytes, *, source_id_base: int) -> ParsedPlannerImportFile:
        try:
            workbook = load_workbook(BytesIO(content), data_only=True)
        except Exception as exc:
            raise bad_request("invalid_waybill_import_xlsx") from exc
        worksheet = workbook.worksheets[0]
        header_row_number = self._find_header_row(worksheet)
        if header_row_number is None:
            raise bad_request("waybill_import_header_not_found")

        headers = [_normalize_header(cell.value) for cell in worksheet[header_row_number]]
        header_map = {header: index for index, header in enumerate(headers) if header}
        if "提单号" not in header_map:
            raise bad_request("waybill_import_header_not_found")

        parsed = ParsedPlannerImportFile()
        merge_context = PlannerWorksheetMergeContext(worksheet, header_map, header_row_number, source_id_base)
        imported_index = 0
        for row_number in range(header_row_number + 1, worksheet.max_row + 1):
            raw_values = [cell.value for cell in worksheet[row_number]]
            values = merge_context.row_values(row_number, len(headers))
            if not any(_clean_text(value) for value in raw_values):
                parsed.skipped_count += 1
                continue
            imported_index += 1
            row, warnings, errors = self._parse_planner_row(
                row_number,
                values,
                header_map,
                board_group=merge_context.board_group_for_row(row_number),
                source_id=-(source_id_base + imported_index),
            )
            parsed.warnings.extend(warnings)
            parsed.errors.extend(errors)
            if row is not None:
                parsed.rows.append(row)
        return parsed

    def _find_header_row(self, worksheet) -> int | None:
        for row_number in range(1, min(10, worksheet.max_row) + 1):
            headers = {_normalize_header(cell.value) for cell in worksheet[row_number]}
            if "提单号" in headers and "航班信息" in headers:
                return row_number
        return None

    def _parse_row(
        self,
        row_number: int,
        values: list[Any],
        header_map: dict[str, int],
        headers: list[str],
    ) -> ParsedWaybillImportRow:
        def cell(header: str) -> Any:
            index = header_map.get(header)
            if index is None or index >= len(values):
                return None
            return values[index]

        waybill_no = _clean_text(cell("提单号"))
        if not waybill_no:
            return ParsedWaybillImportRow(row_number=row_number, error="提单号不能为空")
        planned_flight_info = _clean_text(cell("航班信息"))
        if not planned_flight_info:
            return ParsedWaybillImportRow(row_number=row_number, waybill_no=waybill_no, error="航班信息不能为空")

        notes: list[str] = []
        carrier_agent_id = self._lookup_optional(self.agents_by_name, cell("航代"), "航代", notes)
        if carrier_agent_id is None and _clean_text(cell("航代")):
            return ParsedWaybillImportRow(row_number=row_number, waybill_no=waybill_no, error=notes[-1])
        consignee_contact_id = self._lookup_optional(self.consignees_by_name, cell("收件人"), "收件人", notes)
        if consignee_contact_id is None and _clean_text(cell("收件人")):
            return ParsedWaybillImportRow(row_number=row_number, waybill_no=waybill_no, error=notes[-1])

        document_operator_id = self._lookup_optional(self.users_by_name, cell("资料数据"), "资料数据", notes)
        planned_route_text = _clean_text(cell("航程"))
        cutoff_text = _clean_text(cell("截单时间"))
        if not planned_route_text and cutoff_text and "-" in cutoff_text and not _datetime_or_none(cutoff_text):
            planned_route_text = cutoff_text
            cutoff_text = None

        departure_port, destination_port = self._route_ports(planned_route_text)
        warehouse_data_remark = self._warehouse_data_remark(values, header_map, headers)
        include_tc = self._include_tc_from_warehouse_data(values, header_map)
        if include_tc is None:
            include_tc = _bool_or_none(self._blank_after(values, header_map, "报价"))
        if include_tc is None:
            include_tc = _bool_or_none(cell("报价"))
        if include_tc is None:
            include_tc = False
        notify_pickup = _bool_or_none(cell("通知提取")) or False

        delivery_time = _datetime_or_note(cell("交货时间"), "交货时间", notes)
        document_cutoff_time = _datetime_or_note(cutoff_text if cutoff_text is not None else cell("截单时间"), "截单时间", notes)
        pickup_time = _datetime_or_note(cell("提取时间"), "提取时间", notes)
        payment_date = _date_or_note(cell("付款日期"), "付款日期", notes)
        density = _decimal_or_note(cell("密度"), "密度", notes)

        flight_out_text = _clean_text(cell("飞出时间"))
        arrival_text = _clean_text(cell("到达时间"))
        carrier_text = _clean_text(cell("航司"))
        for label, value in (("航司", carrier_text), ("飞出时间", flight_out_text), ("到达时间", arrival_text)):
            if value:
                notes.append(f"{label}: {value}")

        remark = _clean_text(cell("备注"))
        if remark:
            notes.append(remark)

        try:
            payload = WaybillCreate(
                waybill_no=waybill_no,
                carrier_agent_id=carrier_agent_id,
                warehouse_no=_clean_text(cell("入仓号")),
                consignee_contact_id=consignee_contact_id,
                document_operator_id=document_operator_id,
                planned_flight_info=planned_flight_info,
                planned_route_text=planned_route_text,
                planned_destination=destination_port,
                departure_port=departure_port,
                destination_port=destination_port,
                delivery_time=delivery_time,
                document_cutoff_time=document_cutoff_time,
                booked_weight=_decimal_or_none(cell("订舱重量")),
                booked_volume=_decimal_or_none(cell("方数")),
                density=density,
                quotation=_clean_text(cell("报价")),
                include_tc=include_tc,
                warehouse_data_remark=warehouse_data_remark,
                notify_pickup=notify_pickup,
                pickup_time=pickup_time,
                internal_remark=_join_parts(notes),
                air_freight_cost=_decimal_or_none(cell("航空费")),
                payment_date=payment_date,
            )
        except Exception as exc:
            return ParsedWaybillImportRow(row_number=row_number, waybill_no=waybill_no, error=str(exc))
        return ParsedWaybillImportRow(row_number=row_number, waybill_no=waybill_no, payload=payload)

    def _parse_planner_row(
        self,
        row_number: int,
        values: list[Any],
        header_map: dict[str, int],
        *,
        board_group: ParsedPlannerBoardGroup | None,
        source_id: int,
    ) -> tuple[WarehousePlannerRow | None, list[ParsedPlannerImportWarning], list[ParsedPlannerImportError]]:
        warnings: list[ParsedPlannerImportWarning] = []
        errors: list[ParsedPlannerImportError] = []

        def cell(header: str) -> Any:
            index = header_map.get(header)
            if index is None or index >= len(values):
                return None
            return values[index]

        def warning(field: str, raw_value: Any, message: str) -> None:
            warnings.append(
                ParsedPlannerImportWarning(
                    row_number=row_number,
                    field=field,
                    raw_value=_clean_text(raw_value),
                    message=message,
                )
            )

        def lookup_optional(lookup: dict[str, int | None], value: Any, field: str) -> int | None:
            text = _clean_text(value)
            if not text:
                return None
            matched = lookup.get(_normalize_lookup(text))
            if matched is None:
                warning(field, value, "lookup_not_found")
                return None
            return matched

        def decimal_optional(value: Any, field: str) -> Decimal | None:
            parsed_value = _decimal_or_none(value)
            if _clean_text(value) and parsed_value is None:
                warning(field, value, "invalid_decimal")
            return parsed_value

        waybill_no = _clean_text(cell("提单号"))
        source_type = "import_waybill" if waybill_no else "import_prebooking"
        planned_flight_no: str | None = None
        planned_flight_date: date | None = None
        planned_flight_info = _clean_text(cell("航班信息"))
        if planned_flight_info:
            try:
                planned_flight = parse_planned_flight_info(planned_flight_info, today=local_now().date())
                planned_flight_no = planned_flight.flight_no
                planned_flight_date = planned_flight.flight_date
            except ValueError:
                warning("航班信息", planned_flight_info, "invalid_planned_flight_info")

        route_text = _clean_text(cell("航程"))
        cutoff_text = _clean_text(cell("截单时间"))
        if not route_text and cutoff_text and "-" in cutoff_text and not _datetime_or_none(cutoff_text):
            route_text = cutoff_text
        departure_port, destination_port = self._route_ports(route_text)

        include_tc: bool | None = self._include_tc_from_warehouse_data(values, header_map)
        warehouse_include_tc_text = self._warehouse_data_first_value(values, header_map)
        if _clean_text(warehouse_include_tc_text) and include_tc is None and "tc" in _normalize_lookup(warehouse_include_tc_text):
            warning("入仓数据", warehouse_include_tc_text, "invalid_boolean")
        include_tc_text = self._blank_after(values, header_map, "报价")
        if include_tc is None and _clean_text(include_tc_text):
            include_tc = _bool_or_none(include_tc_text)
            if include_tc is None:
                warning("含T", include_tc_text, "invalid_boolean")
        if include_tc is None:
            include_tc = _bool_or_none(cell("报价"))

        if board_group and board_group.booked_volume is None and _clean_text(cell("方数")):
            warning("方数", cell("方数"), "invalid_board_volume")
        if board_group and board_group.booked_weight is None and _clean_text(cell("订舱重量")):
            warning("订舱重量", cell("订舱重量"), "invalid_board_weight")

        try:
            row = WarehousePlannerRow(
                source_type=source_type,
                source_id=source_id,
                waybill_no=waybill_no,
                carrier_agent_id=lookup_optional(self.agents_by_name, cell("航代"), "航代"),
                planned_flight_no=planned_flight_no,
                planned_flight_date=planned_flight_date,
                receipt_ids=[],
                consignee_contact_id=lookup_optional(self.consignees_by_name, cell("收件人"), "收件人"),
                customs_staff_id=lookup_optional(self.users_by_name, cell("资料数据"), "资料数据"),
                board_group_id=board_group.group_id if board_group else None,
                board_group_order=board_group.order if board_group else None,
                board_booked_volume=board_group.booked_volume if board_group else None,
                board_booked_weight=board_group.booked_weight if board_group else None,
                booked_volume=None if board_group else decimal_optional(cell("方数"), "方数"),
                booked_weight=None if board_group and board_group.booked_weight is not None else decimal_optional(cell("订舱重量"), "订舱重量"),
                density=decimal_optional(cell("密度"), "密度"),
                quotation=_clean_text(cell("报价")),
                include_tc=include_tc,
                departure_port=departure_port,
                destination_port=destination_port,
                planned_route_text=route_text,
                internal_remark=_clean_text(cell("内部备注")) or None,
                source_updated_at=local_now(),
            )
        except Exception as exc:
            errors.append(ParsedPlannerImportError(row_number=row_number, waybill_no=waybill_no, message=str(exc)))
            return None, warnings, errors
        return row, warnings, errors

    def _lookup_optional(
        self,
        lookup: dict[str, int | None],
        value: Any,
        label: str,
        notes: list[str],
    ) -> int | None:
        text = _clean_text(value)
        if not text:
            return None
        matched = lookup.get(_normalize_lookup(text))
        if matched is None:
            notes.append(f"{label}未匹配: {text}")
            return None
        return matched

    def _route_ports(self, route: str | None) -> tuple[str | None, str | None]:
        if not route:
            return None, None
        parts = [part.strip().upper() for part in re.split(r"[-/>\s]+", route) if part.strip()]
        if len(parts) < 2:
            return None, None
        return parts[0][:16], parts[-1][:16]

    def _blank_after(self, values: list[Any], header_map: dict[str, int], header: str) -> Any:
        index = header_map.get(header)
        if index is None:
            return None
        next_index = index + 1
        return values[next_index] if next_index < len(values) else None

    def _warehouse_data_first_value(self, values: list[Any], header_map: dict[str, int]) -> Any:
        start = header_map.get("入仓数据")
        if start is None or start >= len(values):
            return None
        return values[start]

    def _include_tc_from_warehouse_data(self, values: list[Any], header_map: dict[str, int]) -> bool | None:
        return _bool_or_none(self._warehouse_data_first_value(values, header_map))

    def _warehouse_data_remark(self, values: list[Any], header_map: dict[str, int], headers: list[str]) -> str | None:
        start = header_map.get("入仓数据")
        if start is None:
            return None
        parts: list[str] = []
        for index in range(start, min(len(values), len(headers))):
            if index > start and headers[index]:
                break
            text = _clean_text(values[index])
            if text:
                parts.append(text)
        return " / ".join(parts) if parts else None


def _datetime_or_note(value: Any, label: str, notes: list[str]) -> datetime | None:
    parsed = _datetime_or_none(value)
    text = _clean_text(value)
    if text and parsed is None:
        notes.append(f"{label}: {text}")
    return parsed


def _date_or_note(value: Any, label: str, notes: list[str]) -> date | None:
    parsed = _date_or_none(value)
    text = _clean_text(value)
    if text and parsed is None:
        notes.append(f"{label}: {text}")
    return parsed


def _decimal_or_note(value: Any, label: str, notes: list[str]) -> Decimal | None:
    parsed = _decimal_or_none(value)
    text = _clean_text(value)
    if text and parsed is None:
        notes.append(f"{label}: {text}")
    return parsed


class WaybillBulkImportService:
    def __init__(self, db: Session):
        self.db = db
        self.waybills = WaybillService(db)

    def import_file(self, file_name: str, content: bytes, current_user: User) -> WaybillBulkImportResult:
        PermissionService.assert_waybill_write(current_user)
        if Path(file_name).suffix.lower() != ".xlsx":
            raise bad_request("waybill_import_only_xlsx_supported")

        parser = WaybillImportTemplateParser(
            agents_by_name=self._agent_lookup(),
            consignees_by_name=self._consignee_lookup(),
            users_by_name=self._user_lookup(),
        )
        parsed = parser.parse(content)

        created: list[WaybillBulkImportCreated] = []
        errors: list[WaybillBulkImportError] = []
        for row in parsed.rows:
            if row.payload is None:
                errors.append(
                    WaybillBulkImportError(row_number=row.row_number, waybill_no=row.waybill_no, message=row.error or "导入失败")
                )
                continue
            try:
                waybill = self.waybills.create(row.payload, current_user)
            except HTTPException as exc:
                self.db.rollback()
                errors.append(
                    WaybillBulkImportError(row_number=row.row_number, waybill_no=row.waybill_no, message=str(exc.detail))
                )
                continue
            except Exception as exc:
                self.db.rollback()
                errors.append(WaybillBulkImportError(row_number=row.row_number, waybill_no=row.waybill_no, message=str(exc)))
                continue
            created.append(WaybillBulkImportCreated(id=waybill.id, waybill_no=waybill.waybill_no))

        return WaybillBulkImportResult(
            file_name=file_name,
            created_count=len(created),
            skipped_count=parsed.skipped_count,
            errors=errors,
            created_waybills=created,
        )

    def _agent_lookup(self) -> dict[str, int]:
        rows = self.db.scalars(select(CarrierAgent).where(CarrierAgent.enabled.is_(True))).all()
        return _single_value_lookup((agent.agent_name, agent.id) for agent in rows)

    def _consignee_lookup(self) -> dict[str, int | None]:
        contacts = self.db.scalars(
            select(ConsigneeContact)
            .options(selectinload(ConsigneeContact.consignee))
            .join(Consignee)
            .where(Consignee.enabled.is_(True), ConsigneeContact.enabled.is_(True))
        ).all()
        pairs: list[tuple[str | None, int]] = []
        for contact in contacts:
            company = contact.consignee.name if contact.consignee else None
            pairs.append((contact.name, contact.id))
            pairs.append((company, contact.id))
            if company:
                pairs.append((f"{company}{contact.name}", contact.id))
                pairs.append((f"{company} {contact.name}", contact.id))
        return _single_value_lookup(pairs)

    def _user_lookup(self) -> dict[str, int]:
        users = self.db.scalars(select(User).where(User.is_active.is_(True))).all()
        pairs: list[tuple[str | None, int]] = []
        for user in users:
            pairs.append((user.username, user.id))
            pairs.append((user.display_name, user.id))
        return _single_value_lookup(pairs)


def _single_value_lookup(pairs) -> dict[str, int | None]:
    lookup: dict[str, int | None] = {}
    for name, value in pairs:
        key = _normalize_lookup(name)
        if not key:
            continue
        if key in lookup and lookup[key] != value:
            lookup[key] = None
        else:
            lookup[key] = value
    return lookup

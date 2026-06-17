from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from app.core.platform_patch import patch_platform_wmi

patch_platform_wmi()

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import ValidationError
from sqlalchemy import exists, select
from sqlalchemy.orm import Session, selectinload

from app.core.exceptions import bad_request
from app.models import AirWaybill, CarrierAgent, User, WarehousePlanningDraft, WarehouseReceipt, WaybillPrebooking
from app.models.enums import WaybillLifecycleStatus
from app.schemas.warehouse_planner import (
    WarehousePlannerBulkImportError,
    WarehousePlannerBulkImportResult,
    WarehousePlannerBulkImportWarning,
    WarehousePlannerCandidate,
    WarehousePlannerCandidatesOut,
    WarehousePlannerCommitRequest,
    WarehousePlannerCommitResult,
    WarehousePlannerDraftOut,
    WarehousePlannerDraftSave,
    WarehousePlannerRow,
    WarehousePlannerRowError,
    WarehousePlannerRowResult,
    WarehousePlannerRowsRequest,
    WarehousePlannerValidateResult,
)
from app.schemas.waybill import WaybillCreate, WaybillUpdate
from app.services.permission_service import PermissionService
from app.services.prebooking_service import PrebookingService
from app.services.warehouse_file_service import WarehouseFileService
from app.services.waybill_bulk_import_service import WaybillBulkImportService, WaybillImportTemplateParser
from app.services.waybill_service import WaybillService
from app.utils.datetime_utils import local_now


ACTIVE_EXCLUDED_STATUSES = {WaybillLifecycleStatus.PICKED_UP, WaybillLifecycleStatus.VOIDED}
IMPORT_SOURCE_TYPES = {"import_waybill", "import_prebooking"}
PLANNER_EXPORT_HEADERS = [
    "排仓栏位",
    "来源",
    "航代",
    "计划航班",
    "提单号",
    "出仓日期",
    "入仓号/入仓文件",
    "指定清关人员",
    "订舱方数/板总方数",
    "约定航班起飞日期",
    "订舱重量",
    "密度",
    "报价",
    "含T",
    "始发港",
    "目的港",
    "航程",
]


class WarehousePlannerService:
    def __init__(self, db: Session):
        self.db = db
        self.waybills = WaybillService(db)
        self.prebookings = PrebookingService(db)
        self.warehouse_files = WarehouseFileService(db)

    def get_draft(self, current_user: User) -> WarehousePlannerDraftOut:
        PermissionService.assert_waybill_write(current_user)
        draft = self._get_draft_model(current_user.id)
        if draft is None:
            return WarehousePlannerDraftOut(rows=[], updated_at=None)
        return WarehousePlannerDraftOut(rows=[WarehousePlannerRow.model_validate(row) for row in draft.rows], updated_at=draft.updated_at)

    def save_draft(self, payload: WarehousePlannerDraftSave, current_user: User) -> WarehousePlannerDraftOut:
        PermissionService.assert_waybill_write(current_user)
        draft = self._get_draft_model(current_user.id)
        rows = [row.model_dump(mode="json") for row in payload.rows]
        if draft is None:
            draft = WarehousePlanningDraft(user_id=current_user.id, rows=rows)
            self.db.add(draft)
        else:
            draft.rows = rows
        self.db.commit()
        return self.get_draft(current_user)

    def clear_draft(self, current_user: User) -> None:
        PermissionService.assert_waybill_write(current_user)
        draft = self._get_draft_model(current_user.id)
        if draft is not None:
            self.db.delete(draft)
            self.db.commit()

    def candidates(self, current_user: User) -> WarehousePlannerCandidatesOut:
        PermissionService.assert_waybill_write(current_user)
        waybill_items = self._candidate_waybills()
        prebooking_items = self._candidate_prebookings()
        return WarehousePlannerCandidatesOut(
            waybills=[self._waybill_candidate(item) for item in waybill_items],
            prebookings=[self._prebooking_candidate(item) for item in prebooking_items],
            unbound_receipts=self._all_unbound_receipts(),
        )

    def bulk_import(self, file_name: str, content: bytes, current_user: User) -> WarehousePlannerBulkImportResult:
        PermissionService.assert_waybill_write(current_user)
        if Path(file_name).suffix.lower() != ".xlsx":
            raise bad_request("waybill_import_only_xlsx_supported")

        lookup_service = WaybillBulkImportService(self.db)
        parser = WaybillImportTemplateParser(
            agents_by_name=lookup_service._agent_lookup(),
            consignees_by_name=lookup_service._consignee_lookup(),
            users_by_name=lookup_service._user_lookup(),
        )
        source_id_base = int(local_now().timestamp() * 1000) * 1000
        parsed = parser.parse_planner(content, source_id_base=source_id_base)
        return WarehousePlannerBulkImportResult(
            file_name=file_name,
            imported_count=len(parsed.rows),
            skipped_count=parsed.skipped_count,
            rows=parsed.rows,
            warnings=[
                WarehousePlannerBulkImportWarning(
                    row_number=item.row_number,
                    field=item.field,
                    raw_value=item.raw_value,
                    message=item.message,
                )
                for item in parsed.warnings
            ],
            errors=[
                WarehousePlannerBulkImportError(
                    row_number=item.row_number,
                    waybill_no=item.waybill_no,
                    message=item.message,
                )
                for item in parsed.errors
            ],
        )

    def validate_rows(self, payload: WarehousePlannerRowsRequest, current_user: User) -> WarehousePlannerValidateResult:
        PermissionService.assert_waybill_write(current_user)
        results = [self._validate_row(row, current_user) for row in payload.rows]
        valid_count = sum(1 for item in results if item.status == "valid")
        invalid_count = len(results) - valid_count
        return WarehousePlannerValidateResult(valid_count=valid_count, invalid_count=invalid_count, results=results)

    def commit(self, payload: WarehousePlannerCommitRequest, current_user: User) -> WarehousePlannerCommitResult:
        PermissionService.assert_waybill_write(current_user)
        validation = self.validate_rows(WarehousePlannerRowsRequest(rows=payload.rows), current_user)
        invalid_results = [item for item in validation.results if item.status == "invalid"]
        if payload.mode == "all_or_none" and invalid_results:
            return WarehousePlannerCommitResult(
                success_count=0,
                failed_count=len(invalid_results),
                results=[
                    self._failed_from_validation(item) if item.status == "invalid" else item
                    for item in validation.results
                ],
                remaining_rows=payload.rows,
                skipped_due_to_all_or_none=True,
            )

        valid_rows_by_key = {
            self._row_key(row): row
            for row, result in zip(payload.rows, validation.results, strict=False)
            if result.status == "valid"
        }
        results: list[WarehousePlannerRowResult] = []
        committed_keys: set[tuple[str, int]] = set()

        if payload.mode == "all_or_none":
            current_row: WarehousePlannerRow | None = None
            try:
                for row in payload.rows:
                    current_row = row
                    result = self._commit_row(row, current_user)
                    results.append(result)
                    committed_keys.add(self._row_key(row))
                self._save_remaining_rows(current_user, [])
                self.db.commit()
                return WarehousePlannerCommitResult(
                    success_count=len(results),
                    failed_count=0,
                    results=results,
                    remaining_rows=[],
                )
            except Exception as exc:
                self.db.rollback()
                failed_key = self._row_key(current_row) if current_row is not None else None
                failure_results = [
                    self._row_result_from_exception(row, exc)
                    if self._row_key(row) == failed_key
                    else WarehousePlannerRowResult(
                        source_type=row.source_type,
                        source_id=row.source_id,
                        status="failed",
                        waybill_no=row.waybill_no,
                        errors=[WarehousePlannerRowError(message="skipped_due_to_all_or_none")],
                    )
                    for row in payload.rows
                ]
                return WarehousePlannerCommitResult(
                    success_count=0,
                    failed_count=len(payload.rows) if payload.rows else 0,
                    results=failure_results,
                    remaining_rows=payload.rows,
                    skipped_due_to_all_or_none=True,
                )

        for row in payload.rows:
            key = self._row_key(row)
            validation_result = next((item for item in validation.results if (item.source_type, item.source_id) == key), None)
            if key not in valid_rows_by_key:
                results.append(self._failed_from_validation(validation_result) if validation_result else self._invalid_result(row, None, "invalid_row"))
                continue
            try:
                result = self._commit_row(row, current_user)
                self.db.commit()
                results.append(result)
                committed_keys.add(key)
            except Exception as exc:
                self.db.rollback()
                results.append(self._row_result_from_exception(row, exc))

        remaining_rows = [row for row in payload.rows if self._row_key(row) not in committed_keys]
        self._save_remaining_rows(current_user, remaining_rows)
        self.db.commit()
        return WarehousePlannerCommitResult(
            success_count=len(committed_keys),
            failed_count=len(payload.rows) - len(committed_keys),
            results=results,
            remaining_rows=remaining_rows,
        )

    def export_draft(self, current_user: User) -> bytes:
        draft = self.get_draft(current_user)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "排仓编辑区"
        sheet.append(PLANNER_EXPORT_HEADERS)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="EDE9FE")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in draft.rows:
            sheet.append(self._export_row_values(row))

        widths = [12, 12, 18, 16, 18, 14, 26, 18, 16, 18, 14, 12, 12, 10, 12, 12, 24]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        for item in sheet.iter_rows():
            for cell in item:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A2"

        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def _get_draft_model(self, user_id: int) -> WarehousePlanningDraft | None:
        return self.db.scalar(select(WarehousePlanningDraft).where(WarehousePlanningDraft.user_id == user_id))

    def _all_unbound_receipts(self):
        page = 1
        items = []
        while True:
            batch, total, current_page, page_size = self.warehouse_files.list_receipts(
                page=page,
                page_size=100,
                unbound_only=True,
            )
            items.extend(batch)
            if current_page * page_size >= total or not batch:
                return items
            page += 1

    def _candidate_waybills(self) -> list[AirWaybill]:
        has_receipt = select(WarehouseReceipt.id).where(WarehouseReceipt.waybill_id == AirWaybill.id).exists()
        return list(
            self.db.scalars(
                select(AirWaybill)
                .options(
                    selectinload(AirWaybill.plan),
                    selectinload(AirWaybill.carrier_agent),
                    selectinload(AirWaybill.customs_staff),
                )
                .where(
                    AirWaybill.lifecycle_status.notin_(list(ACTIVE_EXCLUDED_STATUSES)),
                    (AirWaybill.outbound_date.is_(None) | ~has_receipt),
                )
                .order_by(AirWaybill.outbound_date.is_(None).desc(), AirWaybill.id.desc())
                .limit(100)
            )
        )

    def _candidate_prebookings(self) -> list[WaybillPrebooking]:
        return list(
            self.db.scalars(
                self.prebookings.base_query()
                .where(WaybillPrebooking.status == "draft")
                .order_by(WaybillPrebooking.planned_flight_date.asc(), WaybillPrebooking.id.desc())
                .limit(100)
            )
        )

    def _waybill_candidate(self, waybill: AirWaybill) -> WarehousePlannerCandidate:
        receipts = self._receipt_summaries_for_waybill(waybill.id)
        return WarehousePlannerCandidate(
            source_type="waybill",
            source_id=waybill.id,
            label=waybill.waybill_no,
            waybill_no=waybill.waybill_no,
            carrier_agent_id=waybill.carrier_agent_id,
            carrier_agent=waybill.carrier_agent,
            planned_flight_no=getattr(waybill.plan, "planned_flight_no", None),
            planned_flight_date=getattr(waybill.plan, "planned_flight_date", None),
            outbound_date=waybill.outbound_date,
            receipts=receipts,
            customs_staff_id=waybill.customs_staff_id,
            customs_staff=waybill.customs_staff,
            booked_volume=waybill.booked_volume,
            booked_weight=waybill.booked_weight,
            density=waybill.density,
            quotation=waybill.quotation,
            include_tc=waybill.include_tc,
            departure_port=waybill.departure_port,
            destination_port=waybill.destination_port,
            planned_route_text=getattr(waybill.plan, "planned_route_text", None),
            lifecycle_status=waybill.lifecycle_status.value if hasattr(waybill.lifecycle_status, "value") else str(waybill.lifecycle_status),
            source_updated_at=waybill.updated_at,
        )

    def _prebooking_candidate(self, prebooking: WaybillPrebooking) -> WarehousePlannerCandidate:
        receipts = self._receipt_summaries_for_prebooking(prebooking.id)
        return WarehousePlannerCandidate(
            source_type="prebooking",
            source_id=prebooking.id,
            label=f"预排仓 #{prebooking.id}",
            waybill_no=prebooking.waybill_no,
            carrier_agent_id=prebooking.carrier_agent_id,
            carrier_agent=prebooking.carrier_agent,
            planned_flight_no=prebooking.planned_flight_no,
            planned_flight_date=prebooking.planned_flight_date,
            outbound_date=prebooking.outbound_date,
            receipts=receipts,
            customs_staff_id=prebooking.customs_staff_id,
            customs_staff=prebooking.customs_staff,
            booked_volume=prebooking.booked_volume,
            booked_weight=prebooking.booked_weight,
            density=prebooking.density,
            quotation=prebooking.quotation,
            include_tc=prebooking.include_tc,
            departure_port=prebooking.departure_port,
            destination_port=prebooking.destination_port,
            planned_route_text=prebooking.planned_route_text,
            lifecycle_status=prebooking.status,
            source_updated_at=prebooking.updated_at,
        )

    def _receipt_summaries_for_waybill(self, waybill_id: int):
        receipts = self.db.scalars(
            select(WarehouseReceipt)
            .where(WarehouseReceipt.waybill_id == waybill_id)
            .order_by(WarehouseReceipt.updated_at.desc(), WarehouseReceipt.id.desc())
        )
        return [self.warehouse_files.get_receipt_summary(item.id) for item in receipts]

    def _receipt_summaries_for_prebooking(self, prebooking_id: int):
        receipts = self.db.scalars(
            select(WarehouseReceipt)
            .where(WarehouseReceipt.prebooking_id == prebooking_id)
            .order_by(WarehouseReceipt.updated_at.desc(), WarehouseReceipt.id.desc())
        )
        return [self.warehouse_files.get_receipt_summary(item.id) for item in receipts]

    def _validate_row(self, row: WarehousePlannerRow, current_user: User) -> WarehousePlannerRowResult:
        errors: list[WarehousePlannerRowError] = []
        if row.source_type == "waybill":
            waybill = self.db.get(AirWaybill, row.source_id)
            if waybill is None:
                return self._invalid_result(row, None, "waybill_not_found")
            if waybill.lifecycle_status in ACTIVE_EXCLUDED_STATUSES:
                errors.append(WarehousePlannerRowError(field="source", message="waybill_not_active"))
            errors.extend(self._validate_waybill_update(row))
            errors.extend(self._validate_receipts(row, current_waybill_id=waybill.id))
            return WarehousePlannerRowResult(
                source_type=row.source_type,
                source_id=row.source_id,
                status="invalid" if errors else "valid",
                waybill_id=waybill.id,
                waybill_no=waybill.waybill_no,
                errors=errors,
            )

        if row.source_type in IMPORT_SOURCE_TYPES:
            errors.extend(self._validate_import_create(row))
            errors.extend(self._validate_receipts(row, target_new_waybill=True))
            return WarehousePlannerRowResult(
                source_type=row.source_type,
                source_id=row.source_id,
                status="invalid" if errors else "valid",
                waybill_no=row.waybill_no,
                errors=errors,
            )

        prebooking = self.db.get(WaybillPrebooking, row.source_id)
        if prebooking is None:
            return self._invalid_result(row, None, "prebooking_not_found")
        if prebooking.status != "draft":
            errors.append(WarehousePlannerRowError(field="source", message="prebooking_not_draft"))
        errors.extend(self._validate_prebooking_convert(row, prebooking))
        errors.extend(self._validate_receipts(row, current_prebooking_id=prebooking.id))
        return WarehousePlannerRowResult(
            source_type=row.source_type,
            source_id=row.source_id,
            status="invalid" if errors else "valid",
            waybill_id=prebooking.converted_waybill_id,
            waybill_no=row.waybill_no or prebooking.waybill_no,
            errors=errors,
        )

    def _validate_waybill_update(self, row: WarehousePlannerRow) -> list[WarehousePlannerRowError]:
        try:
            WaybillUpdate.model_validate(self._waybill_update_data(row))
            return []
        except ValidationError as exc:
            return [self._error_from_validation(exc)]

    def _validate_prebooking_convert(self, row: WarehousePlannerRow, prebooking: WaybillPrebooking) -> list[WarehousePlannerRowError]:
        data = self._prebooking_convert_data(row, prebooking)
        errors = self._required_formal_waybill_errors(data)
        if errors:
            return errors
        try:
            WaybillCreate.model_validate(data)
        except ValidationError as exc:
            return [self._error_from_validation(exc)]
        return []

    def _validate_import_create(self, row: WarehousePlannerRow) -> list[WarehousePlannerRowError]:
        data = self._import_create_data(row)
        errors = self._required_formal_waybill_errors(data)
        if errors:
            return errors
        try:
            WaybillCreate.model_validate(data)
        except ValidationError as exc:
            return [self._error_from_validation(exc)]
        return []

    @staticmethod
    def _required_formal_waybill_errors(data: dict[str, Any]) -> list[WarehousePlannerRowError]:
        errors: list[WarehousePlannerRowError] = []
        required = {
            "waybill_no": "waybill_no_required",
            "carrier_agent_id": "carrier_agent_required",
            "departure_port": "departure_port_required",
            "destination_port": "destination_port_required",
            "planned_route_text": "planned_route_required",
            "booked_weight": "booked_weight_required",
            "booked_volume": "booked_volume_required",
            "quotation": "quotation_required",
        }
        for field, message in required.items():
            if data.get(field) in (None, ""):
                errors.append(WarehousePlannerRowError(field=field, message=message))
        if not data.get("planned_flight_no") or not data.get("planned_flight_date"):
            errors.append(WarehousePlannerRowError(field="planned_flight_no", message="planned_flight_required"))
        return errors

    def _validate_receipts(
        self,
        row: WarehousePlannerRow,
        *,
        current_waybill_id: int | None = None,
        current_prebooking_id: int | None = None,
        target_new_waybill: bool = False,
    ) -> list[WarehousePlannerRowError]:
        errors: list[WarehousePlannerRowError] = []
        for receipt_id in row.receipt_ids:
            receipt = self.db.get(WarehouseReceipt, receipt_id)
            if receipt is None:
                errors.append(WarehousePlannerRowError(field="receipt_ids", message=f"warehouse_receipt_not_found:{receipt_id}"))
                continue
            if target_new_waybill:
                if receipt.waybill_id is not None:
                    errors.append(WarehousePlannerRowError(field="receipt_ids", message=f"receipt_bound_to_waybill:{receipt.warehouse_no}"))
                if receipt.prebooking_id is not None:
                    errors.append(WarehousePlannerRowError(field="receipt_ids", message=f"receipt_bound_to_prebooking:{receipt.warehouse_no}"))
            if current_waybill_id is not None:
                if receipt.prebooking_id is not None:
                    errors.append(WarehousePlannerRowError(field="receipt_ids", message=f"receipt_bound_to_prebooking:{receipt.warehouse_no}"))
                if receipt.waybill_id is not None and receipt.waybill_id != current_waybill_id:
                    errors.append(WarehousePlannerRowError(field="receipt_ids", message=f"receipt_bound_to_other_waybill:{receipt.warehouse_no}"))
            if current_prebooking_id is not None:
                if receipt.waybill_id is not None:
                    errors.append(WarehousePlannerRowError(field="receipt_ids", message=f"receipt_bound_to_waybill:{receipt.warehouse_no}"))
                if receipt.prebooking_id is not None and receipt.prebooking_id != current_prebooking_id:
                    errors.append(WarehousePlannerRowError(field="receipt_ids", message=f"receipt_bound_to_other_prebooking:{receipt.warehouse_no}"))
        return errors

    def _commit_row(self, row: WarehousePlannerRow, current_user: User) -> WarehousePlannerRowResult:
        if row.source_type == "waybill":
            waybill = self.waybills.update(
                row.source_id,
                WaybillUpdate.model_validate(self._waybill_update_data(row)),
                current_user,
                auto_commit=False,
            )
            for receipt_id in row.receipt_ids:
                self.warehouse_files.bind_receipt_to_waybill(receipt_id, waybill.id, current_user, auto_commit=False)
            return WarehousePlannerRowResult(
                source_type=row.source_type,
                source_id=row.source_id,
                status="committed",
                waybill_id=waybill.id,
                waybill_no=waybill.waybill_no,
            )

        if row.source_type in IMPORT_SOURCE_TYPES:
            payload = WaybillCreate.model_validate(self._import_create_data(row))
            waybill = self.waybills.create(payload, current_user, auto_commit=False)
            for receipt_id in row.receipt_ids:
                self.warehouse_files.bind_receipt_to_waybill(receipt_id, waybill.id, current_user, auto_commit=False)
            return WarehousePlannerRowResult(
                source_type=row.source_type,
                source_id=row.source_id,
                status="committed",
                waybill_id=waybill.id,
                waybill_no=waybill.waybill_no,
            )

        prebooking = self.prebookings.get(row.source_id, current_user)
        for receipt_id in row.receipt_ids:
            self.warehouse_files.bind_receipt_to_prebooking(receipt_id, prebooking, current_user, auto_commit=False)
        payload = WaybillCreate.model_validate(self._prebooking_convert_data(row, prebooking))
        waybill = self.prebookings.convert(prebooking.id, payload, current_user, auto_commit=False)
        return WarehousePlannerRowResult(
            source_type=row.source_type,
            source_id=row.source_id,
            status="committed",
            waybill_id=waybill.id,
            waybill_no=waybill.waybill_no,
        )

    def _waybill_update_data(self, row: WarehousePlannerRow) -> dict[str, Any]:
        return {
            "waybill_no": row.waybill_no,
            "carrier_agent_id": row.carrier_agent_id,
            "planned_flight_no": row.planned_flight_no,
            "planned_flight_date": row.planned_flight_date,
            "outbound_date": row.outbound_date,
            "customs_staff_id": row.customs_staff_id,
            "booked_volume": row.booked_volume,
            "booked_weight": row.booked_weight,
            "density": row.density,
            "quotation": row.quotation,
            "include_tc": row.include_tc,
            "departure_port": row.departure_port,
            "destination_port": row.destination_port,
            "planned_route_text": row.planned_route_text,
        }

    def _import_create_data(self, row: WarehousePlannerRow) -> dict[str, Any]:
        return {
            "waybill_no": row.waybill_no,
            "carrier_agent_id": row.carrier_agent_id,
            "planned_flight_no": row.planned_flight_no,
            "planned_flight_date": row.planned_flight_date,
            "planned_destination": row.destination_port,
            "outbound_date": row.outbound_date,
            "consignee_contact_id": row.consignee_contact_id,
            "customs_staff_id": row.customs_staff_id,
            "booked_volume": row.booked_volume,
            "booked_weight": row.booked_weight,
            "density": row.density,
            "quotation": row.quotation,
            "include_tc": row.include_tc if row.include_tc is not None else False,
            "departure_port": row.departure_port,
            "destination_port": row.destination_port,
            "planned_route_text": row.planned_route_text,
        }

    def _prebooking_convert_data(self, row: WarehousePlannerRow, prebooking: WaybillPrebooking) -> dict[str, Any]:
        return {
            "waybill_no": row.waybill_no or prebooking.waybill_no,
            "carrier_agent_id": row.carrier_agent_id if row.carrier_agent_id is not None else prebooking.carrier_agent_id,
            "planned_flight_no": row.planned_flight_no or prebooking.planned_flight_no,
            "planned_flight_date": row.planned_flight_date or prebooking.planned_flight_date,
            "outbound_date": row.outbound_date if row.outbound_date is not None else prebooking.outbound_date,
            "customs_staff_id": row.customs_staff_id if row.customs_staff_id is not None else prebooking.customs_staff_id,
            "booked_volume": row.booked_volume if row.booked_volume is not None else prebooking.booked_volume,
            "booked_weight": row.booked_weight if row.booked_weight is not None else prebooking.booked_weight,
            "density": row.density if row.density is not None else prebooking.density,
            "quotation": row.quotation if row.quotation is not None else prebooking.quotation,
            "include_tc": row.include_tc if row.include_tc is not None else prebooking.include_tc,
            "departure_port": row.departure_port if row.departure_port is not None else prebooking.departure_port,
            "destination_port": row.destination_port if row.destination_port is not None else prebooking.destination_port,
            "planned_route_text": row.planned_route_text if row.planned_route_text is not None else prebooking.planned_route_text,
        }

    def _save_remaining_rows(self, current_user: User, rows: list[WarehousePlannerRow]) -> None:
        draft = self._get_draft_model(current_user.id)
        serialized = [row.model_dump(mode="json") for row in rows]
        if draft is None:
            self.db.add(WarehousePlanningDraft(user_id=current_user.id, rows=serialized))
        else:
            draft.rows = serialized
        self.db.flush()

    def _export_row_values(self, row: WarehousePlannerRow) -> list[Any]:
        receipt_names = []
        for receipt_id in row.receipt_ids:
            receipt = self.db.get(WarehouseReceipt, receipt_id)
            if receipt is not None:
                receipt_names.append(receipt.warehouse_no)
        agent_name = self._agent_name(row.carrier_agent_id)
        customs_name = self._user_name(row.customs_staff_id)
        return [
            row.planning_channel,
            _source_label(row.source_type),
            agent_name,
            row.planned_flight_no,
            row.waybill_no,
            row.outbound_date.isoformat() if row.outbound_date else "",
            " / ".join(receipt_names),
            customs_name,
            _format_decimal(row.booked_volume),
            row.planned_flight_date.isoformat() if row.planned_flight_date else "",
            _format_decimal(row.booked_weight),
            _format_decimal(row.density),
            row.quotation,
            "是" if row.include_tc else "否",
            row.departure_port,
            row.destination_port,
            row.planned_route_text,
        ]

    def _agent_name(self, agent_id: int | None) -> str:
        if agent_id is None:
            return ""
        agent = self.db.get(CarrierAgent, agent_id)
        return getattr(agent, "agent_name", "") if agent else ""

    def _user_name(self, user_id: int | None) -> str:
        if user_id is None:
            return ""
        user = self.db.get(User, user_id)
        if user is None:
            return ""
        return user.display_name or user.username

    @staticmethod
    def _row_key(row: WarehousePlannerRow) -> tuple[str, int]:
        return row.source_type, row.source_id

    @staticmethod
    def _invalid_result(row: WarehousePlannerRow, field: str | None, message: str) -> WarehousePlannerRowResult:
        return WarehousePlannerRowResult(
            source_type=row.source_type,
            source_id=row.source_id,
            status="invalid",
            errors=[WarehousePlannerRowError(field=field, message=message)],
        )

    @staticmethod
    def _failed_from_validation(result: WarehousePlannerRowResult | None) -> WarehousePlannerRowResult:
        if result is None:
            return WarehousePlannerRowResult(source_type="waybill", source_id=0, status="failed", errors=[WarehousePlannerRowError(message="invalid_row")])
        return WarehousePlannerRowResult(
            source_type=result.source_type,
            source_id=result.source_id,
            status="failed",
            waybill_id=result.waybill_id,
            waybill_no=result.waybill_no,
            errors=result.errors,
        )

    @staticmethod
    def _row_result_from_exception(row: WarehousePlannerRow, exc: Exception) -> WarehousePlannerRowResult:
        message = str(exc)
        if isinstance(exc, HTTPException):
            message = str(exc.detail)
        return WarehousePlannerRowResult(
            source_type=row.source_type,
            source_id=row.source_id,
            status="failed",
            waybill_no=row.waybill_no,
            errors=[WarehousePlannerRowError(message=message or "commit_failed")],
        )

    @staticmethod
    def _error_from_validation(exc: ValidationError) -> WarehousePlannerRowError:
        first = exc.errors()[0] if exc.errors() else {}
        loc = first.get("loc") or []
        return WarehousePlannerRowError(
            field=str(loc[0]) if loc else None,
            message=str(first.get("msg") or "invalid_value"),
        )


def _format_decimal(value: Decimal | None) -> str:
    if value is None:
        return ""
    text = format(value, "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def _source_label(source_type: str) -> str:
    labels = {
        "waybill": "正式提单",
        "prebooking": "预排仓",
        "import_waybill": "导入提单",
        "import_prebooking": "导入预排仓",
    }
    return labels.get(source_type, source_type)

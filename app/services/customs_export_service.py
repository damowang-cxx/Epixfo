from __future__ import annotations

import re
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from app.core.platform_patch import patch_platform_wmi

patch_platform_wmi()

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.models import AirWaybill, Box
from app.repositories.box_repository import BoxRepository
from app.utils.planned_flight import extract_planned_flight_no


DECIMAL_001 = Decimal("0.001")
SUMMARY_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
GENERAL_CARGO_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
MONTH_CODES = ("JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC")
DIMENSION_VOLUME_PATTERN = re.compile(
    r"(\d+(?:\.\d+)?)\s*(?:\*|x|X|×)\s*(\d+(?:\.\d+)?)\s*(?:\*|x|X|×)\s*(\d+(?:\.\d+)?)"
)


class CustomsExportService:
    def __init__(self, db: Session):
        self.db = db
        self.boxes = BoxRepository(db)

    def build_waybill_export(self, waybill: AirWaybill) -> bytes:
        workbook = Workbook()
        inbound_sheet = workbook.active
        inbound_sheet.title = "入仓数据"
        self._write_inbound_sheet(inbound_sheet, self.boxes.list_by_waybill(waybill.id))
        self._write_waybill_sheet(workbook.create_sheet("提单资料"), waybill)

        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def _write_inbound_sheet(self, sheet: Worksheet, boxes: list[Box]) -> None:
        headers = ["外箱条码", "提单号码", "品名", "数量", "重量", "收货体积信息", "收货重量/方"]
        sheet.append(headers)
        self._style_header(sheet, len(headers))
        total_weight = Decimal("0.000")
        total_volume = Decimal("0.000")

        ordered_boxes = [box for box in boxes if not getattr(box, "is_general_cargo", False)]
        ordered_boxes.extend(box for box in boxes if getattr(box, "is_general_cargo", False))

        for box in ordered_boxes:
            is_general_cargo = bool(getattr(box, "is_general_cargo", False))
            items = list(box.items or [])
            if not items:
                total_weight += _to_decimal(box.weight) or Decimal("0.000")
                total_volume += _to_decimal(box.volume) or Decimal("0.000")
                sheet.append(
                    [
                        box.box_no,
                        box.warehouse_waybill_no,
                        box.goods_name,
                        box.quantity,
                        _format_decimal_trim(box.weight),
                        _format_box_volume_info(box),
                        _format_decimal_3(box.weight_volume_ratio),
                    ]
                )
                if is_general_cargo:
                    self._highlight_general_cargo_row(sheet)
                continue

            total_volume += _to_decimal(box.volume) or Decimal("0.000")
            for index, item in enumerate(items):
                is_first = index == 0
                total_weight += _to_decimal(item.weight) or Decimal("0.000")
                sheet.append(
                    [
                        box.box_no if is_first else "",
                        item.warehouse_waybill_no,
                        item.goods_name,
                        item.quantity,
                        _format_decimal_trim(item.weight),
                        _format_box_volume_info(box) if is_first else "",
                        _format_decimal_3(box.weight_volume_ratio) if is_first else "",
                    ]
                )
                if is_general_cargo:
                    self._highlight_general_cargo_row(sheet)

        self._append_inbound_summary_row(sheet, total_weight, total_volume)
        self._finish_table(sheet, widths=[18, 20, 28, 10, 12, 16, 16])

    @staticmethod
    def _highlight_general_cargo_row(sheet: Worksheet) -> None:
        for cell in sheet[sheet.max_row][:7]:
            cell.fill = GENERAL_CARGO_FILL

    @staticmethod
    def _append_inbound_summary_row(sheet: Worksheet, total_weight: Decimal, total_volume: Decimal) -> None:
        total_weight = total_weight.quantize(DECIMAL_001)
        total_volume = total_volume.quantize(DECIMAL_001)
        total_density = (
            (total_weight / total_volume).quantize(DECIMAL_001)
            if total_volume > 0
            else Decimal("0.000")
        )
        sheet.append(
            [
                "",
                "",
                "",
                "合计",
                _format_decimal_trim(total_weight),
                _format_decimal_3(total_volume),
                _format_decimal_3(total_density),
            ]
        )
        for cell in sheet[sheet.max_row][3:7]:
            cell.fill = SUMMARY_FILL
            cell.font = Font(bold=True)

    def _write_waybill_sheet(self, sheet: Worksheet, waybill: AirWaybill) -> None:
        rows: list[tuple[str, str]] = [
            ("提单号", waybill.waybill_no),
            ("发货人", ""),
            ("品名", ""),
            ("收货人", self._format_consignee(waybill)),
        ]
        notify_text = self._format_notify_party_if_different(waybill)
        if notify_text:
            rows.append(("通知人", notify_text))
        rows.append(("航班号截单时间隐含签名仓库/打板交单地址", self._format_flight_and_warehouse_text(waybill)))

        for row in rows:
            sheet.append(row)

        self._finish_table(sheet, widths=[34, 92])
        for row in sheet.iter_rows(min_row=1, max_col=2):
            row[0].font = Font(bold=True)
            row[1].alignment = Alignment(wrap_text=True, vertical="top")

    def _format_consignee(self, waybill: AirWaybill) -> str:
        contact = getattr(waybill, "consignee_contact", None)
        if contact is None:
            return waybill.consignee or ""
        return _format_party(
            name=getattr(contact, "name", None),
            address=getattr(contact, "address", None),
            email=getattr(contact, "email", None),
            phone=getattr(contact, "phone", None),
            tax_info=getattr(contact, "tax_info", None),
        )

    def _format_notify_party_if_different(self, waybill: AirWaybill) -> str:
        contact = getattr(waybill, "consignee_contact", None)
        notify = getattr(contact, "notify_party", None) if contact is not None else None
        if notify is None or getattr(notify, "enabled", True) is False:
            return ""

        contact_signature = _party_signature(contact)
        notify_signature = _party_signature(notify)
        if contact_signature == notify_signature:
            return ""
        return _format_party(
            name=getattr(notify, "name", None),
            address=getattr(notify, "address", None),
            email=getattr(notify, "email", None),
            phone=getattr(notify, "phone", None),
            tax_info=getattr(notify, "tax_info", None),
        )

    def _format_flight_and_warehouse_text(self, waybill: AirWaybill) -> str:
        plan = getattr(waybill, "plan", None)
        flight_no = _clean(extract_planned_flight_no(getattr(plan, "planned_flight_no", None)))
        flight_date = _format_flight_date(getattr(plan, "planned_flight_date", None))
        flight_text = "/".join(part for part in [flight_no, flight_date] if part)
        route_text = _clean(getattr(plan, "planned_route_text", None))
        warehouse_text = _clean(getattr(waybill, "warehouse_data_remark", None))
        return f"{waybill.waybill_no}航班：{flight_text}航程：{route_text}{warehouse_text}"

    @staticmethod
    def _style_header(sheet: Worksheet, columns: int) -> None:
        for cell in sheet[1][:columns]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    @staticmethod
    def _finish_table(sheet: Worksheet, widths: list[int]) -> None:
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A2"


def _party_signature(party: Any) -> tuple[str, str, str, str, str]:
    return (
        _normalize(getattr(party, "name", None)),
        _normalize(getattr(party, "address", None)),
        _normalize(getattr(party, "email", None)),
        _normalize(getattr(party, "phone", None)),
        _normalize(getattr(party, "tax_info", None)),
    )


def _format_party(*, name: str | None, address: str | None, email: str | None, phone: str | None, tax_info: str | None) -> str:
    parts = [_clean(name), _clean(address), _clean(tax_info)]
    if _clean(phone):
        parts.append(f"PHONE:{_clean(phone)}")
    if _clean(email):
        parts.append(f"Email:{_clean(email)}")
    return "\n".join(part for part in parts if part)


def _format_flight_date(value: date | None) -> str:
    if value is None:
        return ""
    return f"{value.day:02d}{MONTH_CODES[value.month - 1]}"


def _format_decimal_3(value: Decimal | int | float | str | None) -> str:
    decimal = _to_decimal(value)
    if decimal is None:
        return ""
    return str(decimal.quantize(DECIMAL_001))


def _format_decimal_trim(value: Decimal | int | float | str | None) -> str:
    decimal = _to_decimal(value)
    if decimal is None:
        return ""
    text = str(decimal.normalize())
    return text if "E" not in text else format(decimal, "f").rstrip("0").rstrip(".")


def _format_box_volume_info(box: Box) -> str:
    calculated = _extract_calculated_volume_dimensions(box)
    if calculated:
        return calculated

    original = _extract_dimensions_text(getattr(box, "original_volume_info", None))
    if original:
        return original

    return _format_decimal_3(getattr(box, "volume", None))


def _extract_calculated_volume_dimensions(box: Box) -> str:
    raw_data = getattr(box, "raw_data", None)
    if not isinstance(raw_data, dict):
        return ""
    recalculation = raw_data.get("volume_recalculation")
    if not isinstance(recalculation, dict):
        return ""
    calculated_info = _clean(recalculation.get("calculated_volume_info"))
    if not calculated_info:
        return ""
    before_parenthesis = calculated_info.split("(", 1)[0].strip()
    return _extract_dimensions_text(before_parenthesis)


def _extract_dimensions_text(value: Any) -> str:
    text = _clean(value).replace(",", "")
    if not text:
        return ""
    match = DIMENSION_VOLUME_PATTERN.search(text)
    if not match:
        return ""
    return "*".join(_format_dimension_part(part) for part in match.groups())


def _format_dimension_part(value: str) -> str:
    decimal = Decimal(value)
    text = format(decimal, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def _to_decimal(value: Decimal | int | float | str | None) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _clean(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _normalize(value: Any) -> str:
    return " ".join(_clean(value).lower().split())

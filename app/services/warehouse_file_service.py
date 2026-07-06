from __future__ import annotations

import hashlib
import re
from collections import Counter
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation, ROUND_DOWN, ROUND_HALF_UP
from itertools import product
from io import BytesIO
from pathlib import Path
from typing import Any, Callable

from fastapi import HTTPException, status

from app.core.platform_patch import patch_platform_wmi

patch_platform_wmi()

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.exceptions import bad_request
from app.models import AirWaybill, Box, BoxDocument, BoxItem, User, WarehouseReceipt, WaybillPrebooking
from app.repositories.box_repository import BoxRepository
from app.repositories.waybill_repository import WaybillRepository
from app.schemas.box import (
    BoxBatchOperationResult,
    BoxCreate,
    BoxVolumeRecalculationResult,
    WarehouseChannelReviewIssue,
    WarehouseChannelReviewOut,
    WarehouseBoxConflict,
    WarehouseFileImportError,
    WarehouseFileUploadResult,
    WarehouseProhibitedGoodsIssue,
    WarehouseReceiptBatchDeleteError,
    WarehouseReceiptBatchDeleteItem,
    WarehouseReceiptBatchDeleteResult,
    WarehouseReceiptListOut,
    WarehouseUploadIntegrityIssue,
)
from app.services.permission_service import PermissionService


REQUIRED_COLUMNS = {
    "outer_barcode": {"外箱条码", "外箱條碼", "箱号", "箱號", "外箱号", "外箱號", "box_no", "box no", "barcode"},
    "warehouse_waybill_no": {"提单号码", "提單號碼", "提单号", "提單號", "仓库提单号", "倉庫提單號", "warehouse waybill no"},
    "goods_name": {"品名", "货物品名", "貨物品名", "goods_name", "goods name", "cargo name"},
    "quantity": {"数量", "數量", "件数", "件數", "qty", "quantity"},
    "weight": {"重量", "收货重量", "收貨重量", "weight", "weight kg"},
    "volume": {"收货体积信息", "收貨體積信息", "体积", "體積", "方数", "方數", "volume", "volume cbm"},
}
OPTIONAL_COLUMNS = {
    "original_weight_volume_ratio": {"收货重量/方", "收貨重量/方", "重量/方", "weight/volume", "weight volume ratio"},
}

DECIMAL_001 = Decimal("0.001")
TARGET_VOLUME_TOLERANCE = Decimal("0.500")
MAX_DIMENSION_CANDIDATES_PER_BOX = 48
MAX_DIMENSION_FIT_STATES = 60000
BOX_NO_MAX_LENGTH = 128
BOX_CONFLICT_RAW_KEY = "box_conflict"
UNBOUND_REASONS = {"customs_inspection", "other"}
EUROPE_CHANNEL_PREFIXES = {"UPS", "DHL", "DPD", "FED", "CTT", "FRE", "ITE", "NLE"}
UK_CHANNEL_PREFIXES = {"KDP", "KTK", "DPD"}
DUAL_CHANNEL_PREFIXES = {"DPD"}
CTT_ALLOWED_PREFIXES = {"CTT", "FRE", "ITE", "DHL"}
NLE_ALLOWED_PREFIXES = {"NLE", "DHL"}
EUROPE_CHANNEL_TAGS = ["AMS"]
UK_CHANNEL_TAGS = ["LHR"]
ALL_CTT_CHANNEL_TAGS = ["MAD", "BCN", "AMS"]
PROHIBITED_GOODS_KEYWORDS = ("香水", "perfume")
GENERAL_CARGO_MARKERS = ("\u666e\u8d27", "\u666e\u8ca8")
DIMENSION_VOLUME_PATTERN = re.compile(
    r"(\d+(?:\.\d+)?)\s*(?:\*|x|X|×)\s*(\d+(?:\.\d+)?)\s*(?:\*|x|X|×)\s*(\d+(?:\.\d+)?)"
)


@dataclass(frozen=True)
class _IntegerDimensionCandidate:
    dimensions: tuple[int, int, int]
    volume: Decimal
    units: int
    score: float


@dataclass(frozen=True)
class _IntegerDimensionFit:
    volumes: dict[int, Decimal]
    dimensions: dict[int, tuple[int, int, int]]
    total_volume: Decimal


@dataclass
class ParsedWarehouseBoxItem:
    warehouse_waybill_no: str | None
    goods_name: str | None
    quantity: int | None
    weight: Decimal
    source_row_number: int
    raw_data: dict[str, Any]


@dataclass
class ParsedWarehouseBox:
    box_no: str
    warehouse_waybill_no: str | None
    goods_name: str | None
    quantity: int | None
    weight: Decimal
    original_volume_info: str | None
    original_weight_volume_ratio: str | None
    volume: Decimal
    weight_volume_ratio: Decimal
    source_row_number: int
    is_general_cargo: bool
    raw_data: dict[str, Any]
    items: list[ParsedWarehouseBoxItem]


@dataclass
class WarehouseFileParseResult:
    boxes: list[ParsedWarehouseBox]
    skipped_count: int
    errors: list[WarehouseFileImportError]
    barcode_cells: list[WarehouseUploadIntegrityIssue]


@dataclass
class WarehouseChannelReviewResult:
    review: WarehouseChannelReviewOut
    issues: list[WarehouseChannelReviewIssue]


def _http_exception_message(exc: HTTPException) -> str:
    detail = exc.detail
    if isinstance(detail, dict):
        message = detail.get("message") or detail.get("error_code")
        if isinstance(message, str):
            return message
    if isinstance(detail, str):
        return detail
    return "warehouse_receipt_delete_failed"


class WarehouseFileService:
    def __init__(self, db: Session):
        self.db = db
        self.boxes = BoxRepository(db)
        self.waybills = WaybillRepository(db)

    def list_boxes(self, waybill_id: int) -> list[Box]:
        return self.boxes.list_by_waybill(waybill_id)

    def list_prebooking_boxes(self, prebooking_id: int) -> list[Box]:
        return self.boxes.list_by_prebooking(prebooking_id)

    def list_unbound_boxes(self, *, page: int, page_size: int) -> tuple[list[Box], int, int, int]:
        page = max(page, 1)
        page_size = min(max(page_size, 1), 100)
        items, total = self.boxes.list_unbound(page=page, page_size=page_size)
        return items, total, page, page_size

    def delete_unbound_box(self, box_id: int, current_user: User) -> None:
        PermissionService.assert_waybill_write(current_user)
        box = self.boxes.get_by_id(box_id)
        if box is None:
            raise bad_request("box_not_found")
        if box.warehouse_receipt_id is not None or box.current_waybill_id is not None:
            raise bad_request("box_is_not_unbound")
        self.db.delete(box)
        self.db.commit()

    def list_receipts(
        self,
        *,
        page: int,
        page_size: int,
        unbound_only: bool = False,
    ) -> tuple[list[WarehouseReceiptListOut], int, int, int]:
        page = max(page, 1)
        page_size = min(max(page_size, 1), 100)
        rows, total = self.boxes.list_receipts(page=page, page_size=page_size, unbound_only=unbound_only)
        return [self._receipt_list_out(*row) for row in rows], total, page, page_size

    def list_receipt_boxes(self, receipt_id: int) -> list[Box]:
        if self.boxes.get_receipt_by_id(receipt_id) is None:
            raise bad_request("warehouse_receipt_not_found")
        return self.boxes.list_by_receipt_id(receipt_id)

    def reorder_unbound_receipts(self, receipt_ids: list[int], current_user: User) -> None:
        PermissionService.assert_waybill_write(current_user)
        ordered_receipts = self.boxes.list_unbound_receipt_models_ordered()
        receipts_by_id = {item.id: item for item in ordered_receipts}
        seen: set[int] = set()
        requested_ids: list[int] = []
        for receipt_id in receipt_ids:
            if receipt_id in seen:
                continue
            seen.add(receipt_id)
            requested_ids.append(receipt_id)

        invalid_ids = [receipt_id for receipt_id in requested_ids if receipt_id not in receipts_by_id]
        if invalid_ids:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={
                    "error_code": "warehouse_receipt_order_invalid_receipts",
                    "message": "排序中包含不存在或已绑定的入仓号。",
                    "receipt_ids": invalid_ids,
                },
            )

        ordered_ids = [*requested_ids, *[item.id for item in ordered_receipts if item.id not in seen]]
        for index, receipt_id in enumerate(ordered_ids, start=1):
            receipts_by_id[receipt_id].display_order = index
        self.db.commit()

    def get_receipt_summary(self, receipt_id: int) -> WarehouseReceiptListOut:
        receipt = self.boxes.get_receipt_by_id(receipt_id)
        if receipt is None:
            raise bad_request("warehouse_receipt_not_found")
        waybill_no = None
        if receipt.waybill_id is not None:
            waybill = self.waybills.get(receipt.waybill_id)
            waybill_no = waybill.waybill_no if waybill else None
        document = self.db.get(BoxDocument, receipt.source_document_id) if receipt.source_document_id else None
        prebooking = self.db.get(WaybillPrebooking, receipt.prebooking_id) if receipt.prebooking_id else None
        boxes = self.boxes.list_by_receipt_id(receipt.id)
        box_count = len(boxes)
        general_cargo_count = sum(1 for box in boxes if box.is_general_cargo)
        return self._receipt_list_out(
            receipt,
            waybill_no,
            prebooking.id if prebooking else None,
            prebooking.status if prebooking else None,
            prebooking.planned_flight_date if prebooking else None,
            document.file_name if document else None,
            document.uploaded_at if document else None,
            box_count,
            general_cargo_count,
        )

    def update_box_no(self, waybill_id: int, box_id: int, box_no: str, current_user: User) -> Box:
        """向后兼容：旧调用方仍可只改 box_no。"""
        return self.update_box(waybill_id, box_id, current_user, box_no=box_no)

    def _clean_optional_text(self, value: str | None) -> str | None:
        if value is None:
            return None
        cleaned = value.strip()
        return cleaned or None

    def _quantize_decimal(self, value: Decimal | None) -> Decimal | None:
        if value is None:
            return None
        return value.quantize(DECIMAL_001, rounding=ROUND_HALF_UP)

    def _recompute_weight_volume_ratio(self, box: Box) -> Decimal | None:
        if box.weight is None and box.volume is None:
            return None
        if box.weight is not None and box.volume is not None and box.volume > 0:
            return (box.weight / box.volume).quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
        return Decimal("0.000")

    def _sync_single_item_after_box_update(self, box: Box, fields: set[str]) -> None:
        item_fields = {"warehouse_waybill_no", "goods_name", "quantity", "weight"}
        if not fields.intersection(item_fields):
            return

        if len(box.items) == 1:
            item = box.items[0]
            if "warehouse_waybill_no" in fields:
                item.warehouse_waybill_no = box.warehouse_waybill_no
            if "goods_name" in fields:
                item.goods_name = box.goods_name
            if "quantity" in fields:
                item.quantity = box.quantity
            if "weight" in fields:
                item.weight = box.weight
            item.raw_data = {**(item.raw_data or {}), "manual_edit": True}
            return

        if box.items:
            return

        has_item_data = any(
            value not in (None, "")
            for value in (box.warehouse_waybill_no, box.goods_name, box.quantity, box.weight)
        )
        if not has_item_data:
            return

        box.items.append(
            BoxItem(
                box_id=box.id,
                document_id=box.document_id,
                warehouse_waybill_no=box.warehouse_waybill_no,
                goods_name=box.goods_name,
                quantity=box.quantity,
                weight=box.weight,
                raw_data={"source": "manual_edit"},
            )
        )

    def _refresh_receipt_channel_tags(self, receipt: WarehouseReceipt | None) -> None:
        if receipt is None:
            return
        receipt.channel_tags = compute_warehouse_receipt_channel_tags(
            [box.box_no for box in self.boxes.list_by_receipt_id(receipt.id)]
        )

    def _apply_box_update(
        self,
        box: Box,
        current_user: User,
        *,
        fields_set: set[str] | None = None,
        box_no: str | None = None,
        warehouse_waybill_no: str | None = None,
        goods_name: str | None = None,
        quantity: int | None = None,
        weight: Decimal | None = None,
        volume: Decimal | None = None,
        weight_volume_ratio: Decimal | None = None,
        is_general_cargo: bool | None = None,
    ) -> None:
        supplied_values = {
            "box_no": box_no,
            "warehouse_waybill_no": warehouse_waybill_no,
            "goods_name": goods_name,
            "quantity": quantity,
            "weight": weight,
            "volume": volume,
            "weight_volume_ratio": weight_volume_ratio,
            "is_general_cargo": is_general_cargo,
        }
        fields = set(fields_set or [])
        if not fields:
            fields = {field for field, value in supplied_values.items() if value is not None}

        if "box_no" in fields:
            cleaned_box_no = (box_no or "").strip()
            if not cleaned_box_no:
                raise bad_request("box_no_required")
            existing = self.boxes.get_by_box_no(cleaned_box_no)
            if existing is not None and existing.id != box.id:
                raise bad_request("box_no_exists")
            box.box_no = cleaned_box_no

        if "warehouse_waybill_no" in fields:
            box.warehouse_waybill_no = self._clean_optional_text(warehouse_waybill_no)
        if "goods_name" in fields:
            box.goods_name = self._clean_optional_text(goods_name)
        if "quantity" in fields:
            box.quantity = quantity
        if "weight" in fields:
            box.weight = self._quantize_decimal(weight)
        if "volume" in fields:
            box.volume = self._quantize_decimal(volume)
        if "weight_volume_ratio" in fields:
            box.weight_volume_ratio = self._quantize_decimal(weight_volume_ratio)
        elif "weight" in fields or "volume" in fields:
            box.weight_volume_ratio = self._recompute_weight_volume_ratio(box)
        if "is_general_cargo" in fields:
            box.is_general_cargo = bool(is_general_cargo)

        self._sync_single_item_after_box_update(box, fields)
        box.raw_data = {
            **(box.raw_data or {}),
            "manual_edit": True,
            "manual_edit_updated_by": current_user.id,
            "manual_edit_updated_at": datetime.now(UTC).isoformat(),
        }

    def update_box(
        self,
        waybill_id: int,
        box_id: int,
        current_user: User,
        *,
        fields_set: set[str] | None = None,
        box_no: str | None = None,
        warehouse_waybill_no: str | None = None,
        goods_name: str | None = None,
        quantity: int | None = None,
        weight: Decimal | None = None,
        volume: Decimal | None = None,
        weight_volume_ratio: Decimal | None = None,
        is_general_cargo: bool | None = None,
    ) -> Box:
        """部分更新：传入字段可被清空，未传字段保持不变。"""
        PermissionService.assert_waybill_write(current_user)
        if self.waybills.get(waybill_id) is None:
            raise bad_request("waybill_not_found")
        box = self.boxes.get_by_waybill(waybill_id, box_id)
        if box is None:
            raise bad_request("box_not_found")

        self._apply_box_update(
            box,
            current_user,
            fields_set=fields_set,
            box_no=box_no,
            warehouse_waybill_no=warehouse_waybill_no,
            goods_name=goods_name,
            quantity=quantity,
            weight=weight,
            volume=volume,
            weight_volume_ratio=weight_volume_ratio,
            is_general_cargo=is_general_cargo,
        )
        self._refresh_receipt_totals(box.warehouse_receipt)
        self._refresh_receipt_channel_tags(box.warehouse_receipt)
        self.db.commit()
        self.db.refresh(box)
        return box

    def update_prebooking_box(
        self,
        prebooking: WaybillPrebooking,
        box_id: int,
        current_user: User,
        *,
        fields_set: set[str] | None = None,
        box_no: str | None = None,
        warehouse_waybill_no: str | None = None,
        goods_name: str | None = None,
        quantity: int | None = None,
        weight: Decimal | None = None,
        volume: Decimal | None = None,
        weight_volume_ratio: Decimal | None = None,
        is_general_cargo: bool | None = None,
    ) -> Box:
        PermissionService.assert_waybill_write(current_user)
        if prebooking.status != "draft":
            raise bad_request("prebooking_not_editable")
        box = self.boxes.get_by_prebooking(prebooking.id, box_id)
        if box is None:
            raise bad_request("box_not_found")

        self._apply_box_update(
            box,
            current_user,
            fields_set=fields_set,
            box_no=box_no,
            warehouse_waybill_no=warehouse_waybill_no,
            goods_name=goods_name,
            quantity=quantity,
            weight=weight,
            volume=volume,
            weight_volume_ratio=weight_volume_ratio,
            is_general_cargo=is_general_cargo,
        )
        self._refresh_receipt_totals(box.warehouse_receipt)
        self._refresh_receipt_channel_tags(box.warehouse_receipt)
        self.db.commit()
        self.db.refresh(box)
        return box

    def update_unbound_receipt_box(
        self,
        receipt_id: int,
        box_id: int,
        current_user: User,
        *,
        fields_set: set[str] | None = None,
        box_no: str | None = None,
        warehouse_waybill_no: str | None = None,
        goods_name: str | None = None,
        quantity: int | None = None,
        weight: Decimal | None = None,
        volume: Decimal | None = None,
        weight_volume_ratio: Decimal | None = None,
        is_general_cargo: bool | None = None,
    ) -> Box:
        PermissionService.assert_waybill_write(current_user)
        receipt = self.boxes.get_receipt_by_id(receipt_id)
        if receipt is None:
            raise bad_request("warehouse_receipt_not_found")
        if receipt.waybill_id is not None or receipt.prebooking_id is not None:
            raise bad_request("warehouse_receipt_not_unbound")
        box = self.boxes.get_by_id(box_id)
        if box is None or box.warehouse_receipt_id != receipt.id:
            raise bad_request("box_not_found")

        self._apply_box_update(
            box,
            current_user,
            fields_set=fields_set,
            box_no=box_no,
            warehouse_waybill_no=warehouse_waybill_no,
            goods_name=goods_name,
            quantity=quantity,
            weight=weight,
            volume=volume,
            weight_volume_ratio=weight_volume_ratio,
            is_general_cargo=is_general_cargo,
        )
        self._refresh_receipt_totals(receipt)
        self._refresh_receipt_channel_tags(receipt)
        self.db.commit()
        self.db.refresh(box)
        return box

    def create_box(self, waybill_id: int, payload: BoxCreate, current_user: User) -> Box:
        PermissionService.assert_waybill_write(current_user)
        waybill = self.waybills.get(waybill_id)
        if waybill is None:
            raise bad_request("waybill_not_found")
        cleaned_box_no = (payload.box_no or "").strip()
        if not cleaned_box_no:
            raise bad_request("box_no_required")
        if self.boxes.get_by_box_no(cleaned_box_no) is not None:
            raise bad_request("box_no_exists")

        receipt = self._resolve_waybill_receipt_for_manual_box(waybill, payload.warehouse_receipt_id, current_user)
        weight = payload.weight.quantize(DECIMAL_001, rounding=ROUND_HALF_UP) if payload.weight is not None else None
        volume = payload.volume.quantize(DECIMAL_001, rounding=ROUND_HALF_UP) if payload.volume is not None else None
        weight_volume_ratio = (
            (weight / volume).quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
            if weight is not None and volume is not None and volume > 0
            else (Decimal("0.000") if weight is not None or volume is not None else None)
        )

        box = Box(
            box_no=cleaned_box_no,
            warehouse_receipt_id=receipt.id,
            current_waybill_id=waybill.id,
            warehouse_waybill_no=payload.warehouse_waybill_no,
            goods_name=payload.goods_name,
            quantity=payload.quantity,
            weight=weight,
            volume=volume,
            weight_volume_ratio=weight_volume_ratio,
            is_general_cargo=payload.is_general_cargo,
            never_bound_direct_upload=False,
            status="bound",
            raw_data={"source": "manual"},
        )
        self.db.add(box)
        self.db.flush()

        has_item_data = any(
            value not in (None, "")
            for value in (payload.warehouse_waybill_no, payload.goods_name, payload.quantity, weight)
        )
        if has_item_data:
            self.boxes.add_items(
                [
                    BoxItem(
                        box_id=box.id,
                        warehouse_waybill_no=payload.warehouse_waybill_no,
                        goods_name=payload.goods_name,
                        quantity=payload.quantity,
                        weight=weight,
                        raw_data={"source": "manual"},
                    )
                ]
            )

        self._refresh_receipt_totals(receipt)
        self.db.commit()
        self.db.refresh(box)
        return box

    def create_prebooking_box(self, prebooking: WaybillPrebooking, payload: BoxCreate, current_user: User) -> Box:
        PermissionService.assert_waybill_write(current_user)
        if prebooking.status != "draft":
            raise bad_request("prebooking_not_editable")
        receipt = self._resolve_prebooking_receipt_for_manual_box(prebooking, payload.warehouse_receipt_id)

        cleaned_box_no = (payload.box_no or "").strip()
        if not cleaned_box_no:
            raise bad_request("box_no_required")
        if self.boxes.get_by_box_no(cleaned_box_no) is not None:
            raise bad_request("box_no_exists")

        weight = payload.weight.quantize(DECIMAL_001, rounding=ROUND_HALF_UP) if payload.weight is not None else None
        volume = payload.volume.quantize(DECIMAL_001, rounding=ROUND_HALF_UP) if payload.volume is not None else None
        weight_volume_ratio = (
            (weight / volume).quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
            if weight is not None and volume is not None and volume > 0
            else (Decimal("0.000") if weight is not None or volume is not None else None)
        )

        box = Box(
            box_no=cleaned_box_no,
            warehouse_receipt_id=receipt.id,
            current_waybill_id=None,
            warehouse_waybill_no=payload.warehouse_waybill_no,
            goods_name=payload.goods_name,
            quantity=payload.quantity,
            weight=weight,
            volume=volume,
            weight_volume_ratio=weight_volume_ratio,
            is_general_cargo=payload.is_general_cargo,
            never_bound_direct_upload=False,
            status="prebooked",
            raw_data={"source": "manual_prebooking"},
        )
        self.db.add(box)
        self.db.flush()

        has_item_data = any(
            value not in (None, "")
            for value in (payload.warehouse_waybill_no, payload.goods_name, payload.quantity, weight)
        )
        if has_item_data:
            self.boxes.add_items(
                [
                    BoxItem(
                        box_id=box.id,
                        warehouse_waybill_no=payload.warehouse_waybill_no,
                        goods_name=payload.goods_name,
                        quantity=payload.quantity,
                        weight=weight,
                        raw_data={"source": "manual_prebooking"},
                    )
                ]
            )

        self._refresh_receipt_totals(receipt)
        self.db.commit()
        self.db.refresh(box)
        return box

    def delete_box(self, waybill_id: int, box_id: int, current_user: User) -> None:
        PermissionService.assert_waybill_write(current_user)
        if self.waybills.get(waybill_id) is None:
            raise bad_request("waybill_not_found")
        box = self.boxes.get_by_waybill(waybill_id, box_id)
        if box is None:
            raise bad_request("box_not_found")
        receipt = box.warehouse_receipt
        self.db.delete(box)
        self.db.flush()
        self._refresh_receipt_totals(receipt)
        self.db.commit()

    def delete_prebooking_box(self, prebooking: WaybillPrebooking, box_id: int, current_user: User) -> None:
        PermissionService.assert_waybill_write(current_user)
        if prebooking.status != "draft":
            raise bad_request("prebooking_not_editable")
        box = self.boxes.get_by_prebooking(prebooking.id, box_id)
        if box is None:
            raise bad_request("box_not_found")
        receipt = box.warehouse_receipt
        self.db.delete(box)
        self.db.flush()
        self._refresh_receipt_totals(receipt)
        self.db.commit()

    def recalculate_box_volumes(
        self,
        waybill_id: int,
        target_volume: Decimal,
        current_user: User,
        warehouse_receipt_id: int | None = None,
    ) -> BoxVolumeRecalculationResult:
        PermissionService.assert_waybill_write(current_user)
        waybill = self.waybills.get(waybill_id)
        if waybill is None:
            raise bad_request("waybill_not_found")

        if warehouse_receipt_id is not None:
            receipt = self.boxes.get_receipt_by_id(warehouse_receipt_id)
            if receipt is None or receipt.waybill_id != waybill_id:
                raise bad_request("warehouse_receipt_not_found")
            boxes = self.boxes.list_by_receipt_id(warehouse_receipt_id)
        else:
            boxes = self.boxes.list_by_waybill(waybill_id)
        return self._recalculate_box_volume_set(
            boxes=boxes,
            target_volume=target_volume,
            result_boxes_loader=lambda: self.boxes.list_by_waybill(waybill_id),
        )

    def recalculate_prebooking_box_volumes(
        self,
        prebooking: WaybillPrebooking,
        target_volume: Decimal,
        current_user: User,
        warehouse_receipt_id: int | None = None,
    ) -> BoxVolumeRecalculationResult:
        PermissionService.assert_waybill_write(current_user)
        if prebooking.status != "draft":
            raise bad_request("prebooking_not_editable")

        if warehouse_receipt_id is not None:
            receipt = self.boxes.get_receipt_by_id(warehouse_receipt_id)
            if receipt is None or receipt.prebooking_id != prebooking.id:
                raise bad_request("warehouse_receipt_not_found")
            boxes = self.boxes.list_by_receipt_id(warehouse_receipt_id)
        else:
            boxes = self.boxes.list_by_prebooking(prebooking.id)
        return self._recalculate_box_volume_set(
            boxes=boxes,
            target_volume=target_volume,
            result_boxes_loader=lambda: self.boxes.list_by_prebooking(prebooking.id),
        )

    def recalculate_unbound_receipt_box_volumes(
        self,
        receipt_id: int,
        target_volume: Decimal,
        current_user: User,
    ) -> BoxVolumeRecalculationResult:
        PermissionService.assert_waybill_write(current_user)
        receipt = self.boxes.get_receipt_by_id(receipt_id)
        if receipt is None:
            raise bad_request("warehouse_receipt_not_found")
        if receipt.waybill_id is not None or receipt.prebooking_id is not None:
            raise bad_request("warehouse_receipt_not_unbound")
        return self._recalculate_box_volume_set(
            boxes=self.boxes.list_by_receipt_id(receipt_id),
            target_volume=target_volume,
            result_boxes_loader=lambda: self.boxes.list_by_receipt_id(receipt_id),
        )

    def _recalculate_box_volume_set(
        self,
        *,
        boxes: list[Box],
        target_volume: Decimal,
        result_boxes_loader: Callable[[], list[Box]],
    ) -> BoxVolumeRecalculationResult:
        target_volume = target_volume.quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
        if target_volume <= 0:
            raise bad_request("target_volume_required")
        if not boxes:
            raise bad_request("warehouse_boxes_required")

        old_total_volume = sum((item.volume or Decimal("0.000") for item in boxes), Decimal("0.000")).quantize(
            DECIMAL_001,
            rounding=ROUND_HALF_UP,
        )
        total_weight = sum((item.weight or Decimal("0.000") for item in boxes), Decimal("0.000")).quantize(
            DECIMAL_001,
            rounding=ROUND_HALF_UP,
        )
        base_volumes = {box.id: _box_original_volume(box) for box in boxes}
        original_total_volume = sum(base_volumes.values(), Decimal("0.000")).quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
        if original_total_volume <= 0:
            raise bad_request("warehouse_volume_required")

        target_volume_upper = (target_volume + TARGET_VOLUME_TOLERANCE).quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
        dimensions_by_box_id: dict[int, tuple[Decimal, Decimal, Decimal]] = {}
        fixed_boxes: list[Box] = []
        adjustable_boxes: list[Box] = []
        for box in boxes:
            dimensions = _parse_dimensions_text(box.original_volume_info)
            if len(box.items or []) > 1 or dimensions is None or base_volumes[box.id] <= 0:
                fixed_boxes.append(box)
                continue
            dimensions_by_box_id[box.id] = dimensions
            adjustable_boxes.append(box)

        fixed_total_volume = sum((base_volumes[box.id] for box in fixed_boxes), Decimal("0.000")).quantize(
            DECIMAL_001,
            rounding=ROUND_HALF_UP,
        )
        adjustable_base_total = sum((base_volumes[box.id] for box in adjustable_boxes), Decimal("0.000")).quantize(
            DECIMAL_001,
            rounding=ROUND_HALF_UP,
        )
        if fixed_total_volume > target_volume_upper:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={
                    "error_code": "target_volume_less_than_fixed_boxes",
                    "message": "固定箱号方数已超过目标方数允许上限，无法只调整一箱一件且有长宽高的箱号。",
                    "target_volume": str(target_volume),
                    "target_volume_upper": str(target_volume_upper),
                    "fixed_total_volume": str(fixed_total_volume),
                    "original_total_volume": str(original_total_volume),
                    "total_volume": str(old_total_volume),
                },
            )

        adjustable_target_min = max(
            Decimal("0.000"),
            (target_volume - fixed_total_volume).quantize(DECIMAL_001, rounding=ROUND_HALF_UP),
        )
        adjustable_target_max = (target_volume_upper - fixed_total_volume).quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
        integer_fit: _IntegerDimensionFit | None = None
        if adjustable_boxes:
            integer_fit = _fit_integer_dimension_volumes_to_target(
                boxes=adjustable_boxes,
                dimensions_by_box_id=dimensions_by_box_id,
                base_volumes=base_volumes,
                target_min=adjustable_target_min,
                target_max=adjustable_target_max,
            )
        elif target_volume <= fixed_total_volume <= target_volume_upper:
            integer_fit = _IntegerDimensionFit(volumes={}, dimensions={}, total_volume=Decimal("0.000"))

        if integer_fit is None and adjustable_base_total <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={
                    "error_code": "warehouse_adjustable_volume_required",
                    "message": "没有可用于整数长宽高调整的一箱一件箱号。",
                    "target_volume": str(target_volume),
                    "target_volume_upper": str(target_volume_upper),
                    "fixed_total_volume": str(fixed_total_volume),
                    "adjustable_total_volume": str(adjustable_base_total),
                    "original_total_volume": str(original_total_volume),
                    "total_volume": str(old_total_volume),
                },
            )
        if integer_fit is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={
                    "error_code": "warehouse_integer_volume_fit_unavailable",
                    "message": "无法找到满足整数长宽高且落入目标方数区间的调整方案。",
                    "target_volume": str(target_volume),
                    "target_volume_upper": str(target_volume_upper),
                    "fixed_total_volume": str(fixed_total_volume),
                    "adjustable_total_volume": str(adjustable_base_total),
                    "original_total_volume": str(original_total_volume),
                    "total_volume": str(old_total_volume),
                },
            )

        fitted_volumes = integer_fit.volumes
        fitted_dimensions = integer_fit.dimensions
        ratio = integer_fit.total_volume / adjustable_base_total if adjustable_base_total > 0 else Decimal("0.000")
        adjustable_box_ids = {box.id for box in adjustable_boxes}
        adjusted_box_count = sum(
            1
            for box in adjustable_boxes
            if fitted_volumes.get(box.id, base_volumes[box.id]) != base_volumes[box.id]
        )
        recalculated_at = datetime.now(UTC).isoformat()
        touched_receipt_ids = {box.warehouse_receipt_id for box in boxes if box.warehouse_receipt_id is not None}
        for box in boxes:
            old_volume = box.volume or Decimal("0.000")
            new_volume = fitted_volumes.get(box.id, base_volumes[box.id])
            box.volume = new_volume
            box.weight_volume_ratio = (
                ((box.weight or Decimal("0.000")) / new_volume).quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
                if new_volume > 0
                else Decimal("0.000")
            )
            calculated_volume_info = (
                _calculated_volume_info_from_dimensions(fitted_dimensions[box.id], new_volume)
                if box.id in fitted_dimensions
                else _calculated_volume_info(box, new_volume)
            )
            raw_data = dict(box.raw_data or {})
            recalculation = {
                "source": "target_volume_integer_dimensions",
                "base_volume": str(base_volumes[box.id]),
                "old_volume": str(old_volume.quantize(DECIMAL_001, rounding=ROUND_HALF_UP)),
                "new_volume": str(new_volume),
                "calculated_volume_info": calculated_volume_info,
                "old_total_volume": str(old_total_volume),
                "original_total_volume": str(original_total_volume),
                "target_volume": str(target_volume),
                "target_volume_upper": str(target_volume_upper),
                "fixed_total_volume": str(fixed_total_volume),
                "adjustable_target_volume": str(adjustable_target_min),
                "adjustable_target_volume_upper": str(adjustable_target_max),
                "fit_adjustable_volume": str(integer_fit.total_volume),
                "ratio": str(ratio.quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP)),
                "adjustable": box.id in adjustable_box_ids,
                "recalculated_at": recalculated_at,
            }
            if box.id in fitted_dimensions:
                recalculation["integer_dimensions"] = list(fitted_dimensions[box.id])
            raw_data["volume_recalculation"] = recalculation
            box.raw_data = raw_data

        for receipt_id in touched_receipt_ids:
            self._refresh_receipt_totals(self.boxes.get_receipt_by_id(receipt_id))
        self.db.commit()
        updated_boxes = result_boxes_loader()
        new_total_volume = sum((item.volume or Decimal("0.000") for item in updated_boxes), Decimal("0.000")).quantize(
            DECIMAL_001,
            rounding=ROUND_HALF_UP,
        )
        return BoxVolumeRecalculationResult(
            target_volume=target_volume,
            total_weight=total_weight,
            original_total_volume=original_total_volume,
            old_total_volume=old_total_volume,
            fixed_total_volume=fixed_total_volume,
            adjustable_total_volume=adjustable_base_total,
            new_total_volume=new_total_volume,
            adjusted=new_total_volume != old_total_volume,
            adjusted_box_count=adjusted_box_count,
            fixed_box_count=len(fixed_boxes),
            boxes=updated_boxes,
        )

    def upload_for_waybill(
        self,
        waybill_id: int,
        file_name: str,
        content: bytes,
        current_user: User,
        force_move_box_nos: list[str] | None = None,
        skip_conflict_box_nos: list[str] | None = None,
    ) -> WarehouseFileUploadResult:
        PermissionService.assert_waybill_write(current_user)
        waybill = self.waybills.get(waybill_id)
        if waybill is None:
            raise bad_request("waybill_not_found")

        parse_result = parse_warehouse_xlsx(file_name, content)
        if not parse_result.boxes:
            raise bad_request(_format_no_valid_rows_error(parse_result.errors))

        warehouse_no = Path(file_name).stem[:128]
        if not warehouse_no:
            raise bad_request("warehouse_no_required")

        skipped_conflict_box_nos = {item.strip() for item in skip_conflict_box_nos or [] if item and item.strip()}
        bindable_boxes = [item for item in parse_result.boxes if item.box_no not in skipped_conflict_box_nos]
        if not bindable_boxes:
            raise bad_request("warehouse_file_no_bindable_boxes")
        integrity_issues = warehouse_upload_integrity_issues(
            parse_result=parse_result,
            uploaded_box_nos=[item.box_no for item in bindable_boxes],
        )
        prohibited_goods_issues = warehouse_prohibited_goods_issues(bindable_boxes)
        target_receipt = self.boxes.get_receipt_by_warehouse_no(warehouse_no)
        if target_receipt is not None and target_receipt.waybill_id not in (None, waybill.id):
            raise bad_request("warehouse_receipt_bound_to_other_waybill")
        if target_receipt is not None and target_receipt.prebooking_id is not None:
            raise bad_request("warehouse_receipt_bound_to_prebooking")
        self._apply_conflict_box_renames(
            bindable_boxes,
            target_receipt=target_receipt,
            target_warehouse_no=warehouse_no,
        )

        file_hash = hashlib.sha256(content).hexdigest()
        stored_path = self._store_file(file_name, file_hash, content)

        document = self.boxes.add_document(
            BoxDocument(
                file_name=file_name,
                file_path=str(stored_path),
                file_hash=file_hash,
                bound_waybill_id=waybill_id,
                uploaded_by=current_user.id,
                uploaded_at=datetime.now(UTC),
            )
        )

        receipt = self._ensure_receipt(warehouse_no, waybill, document=document, current_user=current_user)
        touched_receipt_ids = self._sync_receipt_boxes(receipt, waybill, document, bindable_boxes)
        waybill.warehouse_no = warehouse_no
        waybill.updated_by = current_user.id
        self._refresh_receipt_totals(receipt)
        for receipt_id in touched_receipt_ids:
            if receipt_id != receipt.id:
                self._refresh_receipt_totals(self.boxes.get_receipt_by_id(receipt_id))
        receipt.channel_tags = compute_warehouse_receipt_channel_tags([item.box_no for item in bindable_boxes])
        self.db.commit()

        return WarehouseFileUploadResult(
            file_name=file_name,
            warehouse_no=warehouse_no,
            document_id=document.id,
            uploaded_at=document.uploaded_at,
            success_count=len(bindable_boxes),
            skipped_count=parse_result.skipped_count,
            errors=parse_result.errors,
            channel_tags=list(receipt.channel_tags or []),
            integrity_issues=integrity_issues,
            prohibited_goods_issues=prohibited_goods_issues,
        )

    def upload_for_prebooking(
        self,
        prebooking: WaybillPrebooking,
        file_name: str,
        content: bytes,
        current_user: User,
    ) -> WarehouseFileUploadResult:
        PermissionService.assert_waybill_write(current_user)
        if prebooking.status != "draft":
            raise bad_request("prebooking_not_editable")

        parse_result = parse_warehouse_xlsx(file_name, content)
        if not parse_result.boxes:
            raise bad_request(_format_no_valid_rows_error(parse_result.errors))

        warehouse_no = Path(file_name).stem[:128]
        if not warehouse_no:
            raise bad_request("warehouse_no_required")

        receipt = self.boxes.get_receipt_by_warehouse_no(warehouse_no)
        if receipt is not None and receipt.waybill_id is not None:
            raise bad_request("warehouse_receipt_bound_to_waybill")
        if receipt is not None and receipt.prebooking_id is not None:
            raise bad_request("warehouse_receipt_bound_to_prebooking")
        if receipt is not None and receipt.prebooking_id not in (None, prebooking.id):
            raise bad_request("warehouse_receipt_bound_to_other_prebooking")

        integrity_issues = warehouse_upload_integrity_issues(
            parse_result=parse_result,
            uploaded_box_nos=[item.box_no for item in parse_result.boxes],
        )
        prohibited_goods_issues = warehouse_prohibited_goods_issues(parse_result.boxes)
        self._apply_conflict_box_renames(
            parse_result.boxes,
            target_receipt=receipt,
            target_warehouse_no=warehouse_no,
        )

        file_hash = hashlib.sha256(content).hexdigest()
        stored_path = self._store_file(file_name, file_hash, content)
        document = self.boxes.add_document(
            BoxDocument(
                file_name=file_name,
                file_path=str(stored_path),
                file_hash=file_hash,
                bound_waybill_id=None,
                uploaded_by=current_user.id,
                uploaded_at=datetime.now(UTC),
            )
        )

        if receipt is None:
            receipt = self.boxes.add_receipt(
                WarehouseReceipt(
                    warehouse_no=warehouse_no,
                    waybill_id=None,
                    prebooking_id=prebooking.id,
                    source_document_id=document.id,
                    uploaded_by=current_user.id,
                    total_quantity=0,
                    total_weight=Decimal("0.000"),
                    total_volume=Decimal("0.000"),
                    weight_volume_ratio=Decimal("0.000"),
                )
            )
        else:
            receipt.prebooking_id = prebooking.id
            receipt.source_document_id = document.id
            receipt.uploaded_by = current_user.id
            receipt.display_order = None

        self._sync_prebooking_receipt_boxes(receipt, document, parse_result.boxes)
        self._refresh_receipt_totals(receipt)
        prebooking.updated_by = current_user.id
        self.db.commit()
        return WarehouseFileUploadResult(
            file_name=file_name,
            warehouse_no=warehouse_no,
            document_id=document.id,
            uploaded_at=document.uploaded_at,
            success_count=len(parse_result.boxes),
            skipped_count=parse_result.skipped_count,
            errors=parse_result.errors,
            channel_tags=list(receipt.channel_tags or []),
            integrity_issues=integrity_issues,
            prohibited_goods_issues=prohibited_goods_issues,
        )

    def upload_unbound_file(self, file_name: str, content: bytes, current_user: User) -> WarehouseFileUploadResult:
        PermissionService.assert_waybill_write(current_user)
        parse_result = parse_warehouse_xlsx(file_name, content)
        if not parse_result.boxes:
            raise bad_request(_format_no_valid_rows_error(parse_result.errors))

        warehouse_no = Path(file_name).stem[:128]
        if not warehouse_no:
            raise bad_request("warehouse_no_required")

        channel_review = review_warehouse_file_channels(warehouse_no, file_name, parse_result.boxes)
        channel_review.review.issues = channel_review.issues

        receipt = self.boxes.get_receipt_by_warehouse_no(warehouse_no)
        if receipt is not None and receipt.waybill_id is not None:
            raise bad_request("warehouse_receipt_bound_to_waybill")
        if receipt is not None and receipt.prebooking_id is not None:
            raise bad_request("warehouse_receipt_bound_to_prebooking")

        integrity_issues = warehouse_upload_integrity_issues(
            parse_result=parse_result,
            uploaded_box_nos=[item.box_no for item in parse_result.boxes],
        )
        prohibited_goods_issues = warehouse_prohibited_goods_issues(parse_result.boxes)
        self._apply_conflict_box_renames(
            parse_result.boxes,
            target_receipt=receipt,
            target_warehouse_no=warehouse_no,
        )

        file_hash = hashlib.sha256(content).hexdigest()
        stored_path = self._store_file(file_name, file_hash, content)
        document = self.boxes.add_document(
            BoxDocument(
                file_name=file_name,
                file_path=str(stored_path),
                file_hash=file_hash,
                bound_waybill_id=None,
                uploaded_by=current_user.id,
                uploaded_at=datetime.now(UTC),
            )
        )

        if receipt is None:
            receipt = self.boxes.add_receipt(
                WarehouseReceipt(
                    warehouse_no=warehouse_no,
                    waybill_id=None,
                    prebooking_id=None,
                    source_document_id=document.id,
                    uploaded_by=current_user.id,
                    total_quantity=0,
                    total_weight=Decimal("0.000"),
                    total_volume=Decimal("0.000"),
                    weight_volume_ratio=Decimal("0.000"),
                )
            )
        else:
            receipt.waybill_id = None
            receipt.prebooking_id = None
            receipt.source_document_id = document.id
            receipt.uploaded_by = current_user.id

        self._sync_unbound_receipt_boxes(receipt, document, parse_result.boxes)
        receipt.waybill_id = None
        receipt.prebooking_id = None
        self._refresh_receipt_totals(receipt)
        receipt.channel_tags = compute_warehouse_receipt_channel_tags([item.box_no for item in parse_result.boxes])

        self.db.commit()
        return WarehouseFileUploadResult(
            file_name=file_name,
            warehouse_no=warehouse_no,
            document_id=document.id,
            uploaded_at=document.uploaded_at,
            success_count=len(parse_result.boxes),
            skipped_count=parse_result.skipped_count,
            errors=parse_result.errors,
            channel_review=channel_review.review,
            channel_tags=list(receipt.channel_tags or []),
            integrity_issues=integrity_issues,
            prohibited_goods_issues=prohibited_goods_issues,
        )

    def bind_receipt_to_waybill(
        self,
        receipt_id: int,
        target_waybill_id: int,
        current_user: User,
        *,
        auto_commit: bool = True,
    ) -> WarehouseReceipt:
        PermissionService.assert_waybill_write(current_user)
        receipt = self.boxes.get_receipt_by_id(receipt_id)
        if receipt is None:
            raise bad_request("warehouse_receipt_not_found")
        if receipt.waybill_id is not None and receipt.waybill_id != target_waybill_id:
            raise bad_request("warehouse_receipt_bound_to_other_waybill")
        if receipt.prebooking_id is not None:
            raise bad_request("warehouse_receipt_bound_to_prebooking")
        waybill = self.waybills.get(target_waybill_id)
        if waybill is None:
            raise bad_request("target_waybill_not_found")

        receipt.waybill_id = waybill.id
        receipt.prebooking_id = None
        receipt.display_order = None
        receipt.uploaded_by = receipt.uploaded_by or current_user.id
        for box in self.boxes.list_by_receipt_id(receipt.id):
            box.current_waybill_id = waybill.id
            box.status = "bound"
            box.never_bound_direct_upload = False
            box.unbound_reason = None
            box.unbound_remark = None
        waybill.warehouse_no = receipt.warehouse_no
        waybill.updated_by = current_user.id
        self._refresh_receipt_totals(receipt)
        if auto_commit:
            self.db.commit()
            self.db.refresh(receipt)
        else:
            self.db.flush()
        return receipt

    def unbind_receipt_from_waybill(
        self,
        waybill_id: int,
        receipt_id: int,
        current_user: User,
        *,
        auto_commit: bool = True,
    ) -> WarehouseReceipt:
        PermissionService.assert_waybill_write(current_user)
        waybill = self.waybills.get(waybill_id)
        if waybill is None:
            raise bad_request("waybill_not_found")
        receipt = self.boxes.get_receipt_by_id(receipt_id)
        if receipt is None:
            raise bad_request("warehouse_receipt_not_found")
        if receipt.waybill_id != waybill.id:
            raise bad_request("warehouse_receipt_not_bound_to_waybill")

        receipt.waybill_id = None
        receipt.prebooking_id = None
        receipt.display_order = None
        if receipt.source_document_id is not None:
            document = self.db.get(BoxDocument, receipt.source_document_id)
            if document is not None and document.bound_waybill_id == waybill.id:
                document.bound_waybill_id = None

        for box in self.boxes.list_by_receipt_id(receipt.id):
            box.current_waybill_id = None
            box.status = "unbound"
            box.never_bound_direct_upload = False
            box.unbound_reason = None
            box.unbound_remark = None

        waybill.warehouse_no = self._latest_bound_receipt_warehouse_no(waybill.id)
        waybill.updated_by = current_user.id
        self._refresh_receipt_totals(receipt)
        if auto_commit:
            self.db.commit()
            self.db.refresh(receipt)
        else:
            self.db.flush()
        return receipt

    def bind_receipt_to_prebooking(
        self,
        receipt_id: int,
        prebooking: WaybillPrebooking,
        current_user: User,
        *,
        auto_commit: bool = True,
    ) -> WarehouseReceipt:
        PermissionService.assert_waybill_write(current_user)
        if prebooking.status != "draft":
            raise bad_request("prebooking_not_editable")
        receipt = self.boxes.get_receipt_by_id(receipt_id)
        if receipt is None:
            raise bad_request("warehouse_receipt_not_found")
        if receipt.waybill_id is not None:
            raise bad_request("warehouse_receipt_bound_to_waybill")
        if receipt.prebooking_id is not None and receipt.prebooking_id != prebooking.id:
            raise bad_request("warehouse_receipt_bound_to_other_prebooking")

        receipt.prebooking_id = prebooking.id
        receipt.display_order = None
        receipt.uploaded_by = receipt.uploaded_by or current_user.id
        for box in self.boxes.list_by_receipt_id(receipt.id):
            box.current_waybill_id = None
            box.status = "prebooked"
            box.never_bound_direct_upload = False
            box.unbound_reason = None
            box.unbound_remark = None
        prebooking.updated_by = current_user.id
        self._refresh_receipt_totals(receipt)
        if auto_commit:
            self.db.commit()
            self.db.refresh(receipt)
        else:
            self.db.flush()
        return receipt

    def delete_unbound_receipt(self, receipt_id: int, current_user: User) -> None:
        PermissionService.assert_waybill_write(current_user)
        self._delete_unbound_receipt(receipt_id)
        self.db.commit()

    def batch_delete_unbound_receipts(
        self,
        receipt_ids: list[int],
        current_user: User,
    ) -> WarehouseReceiptBatchDeleteResult:
        PermissionService.assert_waybill_write(current_user)
        seen: set[int] = set()
        unique_receipt_ids: list[int] = []
        for receipt_id in receipt_ids:
            if receipt_id in seen:
                continue
            seen.add(receipt_id)
            unique_receipt_ids.append(receipt_id)
        deleted: list[WarehouseReceiptBatchDeleteItem] = []
        errors: list[WarehouseReceiptBatchDeleteError] = []
        for receipt_id in unique_receipt_ids:
            receipt = self.boxes.get_receipt_by_id(receipt_id)
            warehouse_no = receipt.warehouse_no if receipt is not None else None
            try:
                self._delete_unbound_receipt(receipt_id, receipt=receipt)
                self.db.commit()
                deleted.append(WarehouseReceiptBatchDeleteItem(id=receipt_id, warehouse_no=warehouse_no))
            except HTTPException as exc:
                self.db.rollback()
                errors.append(
                    WarehouseReceiptBatchDeleteError(
                        id=receipt_id,
                        warehouse_no=warehouse_no,
                        message=_http_exception_message(exc),
                    )
                )
            except Exception as exc:
                self.db.rollback()
                errors.append(
                    WarehouseReceiptBatchDeleteError(
                        id=receipt_id,
                        warehouse_no=warehouse_no,
                        message=str(exc) or "warehouse_receipt_delete_failed",
                    )
                )

        return WarehouseReceiptBatchDeleteResult(
            success_count=len(deleted),
            failed_count=len(errors),
            deleted_receipts=deleted,
            errors=errors,
        )

    def _delete_unbound_receipt(self, receipt_id: int, *, receipt: WarehouseReceipt | None = None) -> WarehouseReceipt:
        receipt = receipt or self.boxes.get_receipt_by_id(receipt_id)
        if receipt is None:
            raise bad_request("warehouse_receipt_not_found")
        if receipt.waybill_id is not None:
            raise bad_request("warehouse_receipt_bound_to_waybill")
        if receipt.prebooking_id is not None:
            raise bad_request("warehouse_receipt_bound_to_prebooking")
        for box in self.boxes.list_by_receipt_id(receipt.id):
            self.db.delete(box)
        self.db.delete(receipt)
        return receipt

    def batch_bind_boxes(self, box_ids: list[int], target_waybill_id: int, current_user: User) -> BoxBatchOperationResult:
        return self.batch_transfer_boxes(
            box_ids,
            "waybill",
            current_user,
            target_waybill_id=target_waybill_id,
        )

    def batch_unbind_boxes(self, box_ids: list[int], current_user: User) -> BoxBatchOperationResult:
        PermissionService.assert_waybill_write(current_user)
        boxes = self._load_boxes_for_batch(box_ids)
        return self._move_boxes_to_unbound(boxes, unbound_reason=None, unbound_remark=None)

    def batch_transfer_boxes(
        self,
        box_ids: list[int],
        target_type: str,
        current_user: User,
        *,
        target_waybill_id: int | None = None,
        target_receipt_id: int | None = None,
        unbound_reason: str | None = None,
        unbound_remark: str | None = None,
    ) -> BoxBatchOperationResult:
        PermissionService.assert_waybill_write(current_user)
        boxes = self._load_boxes_for_batch(box_ids)
        if target_type == "waybill":
            if target_waybill_id is None:
                raise bad_request("target_waybill_required")
            return self._move_boxes_to_waybill(boxes, target_waybill_id, current_user)
        if target_type == "receipt":
            if target_receipt_id is None:
                raise bad_request("target_receipt_required")
            return self._move_boxes_to_receipt(boxes, target_receipt_id)
        if target_type == "unbound":
            reason = unbound_reason or "other"
            if reason not in UNBOUND_REASONS:
                raise bad_request("invalid_unbound_reason")
            return self._move_boxes_to_unbound(boxes, unbound_reason=reason, unbound_remark=unbound_remark)
        raise bad_request("invalid_transfer_target_type")

    def _move_boxes_to_waybill(
        self,
        boxes: list[Box],
        target_waybill_id: int,
        current_user: User,
    ) -> BoxBatchOperationResult:
        waybill = self.waybills.get(target_waybill_id)
        if waybill is None:
            raise bad_request("target_waybill_not_found")
        if not waybill.warehouse_no:
            raise bad_request("target_warehouse_no_required")
        receipt = self._ensure_receipt(waybill.warehouse_no, waybill, document=None, current_user=current_user)
        touched_receipt_ids = {box.warehouse_receipt_id for box in boxes if box.warehouse_receipt_id is not None}
        for box in boxes:
            box.warehouse_receipt_id = receipt.id
            box.current_waybill_id = waybill.id
            box.status = "bound"
            box.never_bound_direct_upload = False
            box.unbound_reason = None
            box.unbound_remark = None
        self._refresh_receipt_totals(receipt)
        for receipt_id in touched_receipt_ids:
            if receipt_id != receipt.id:
                self._refresh_receipt_totals(self.boxes.get_receipt_by_id(receipt_id))
        self.db.commit()
        return BoxBatchOperationResult(updated_count=len(boxes), boxes=self.boxes.list_by_ids([box.id for box in boxes]))

    def _move_boxes_to_receipt(
        self,
        boxes: list[Box],
        target_receipt_id: int,
    ) -> BoxBatchOperationResult:
        receipt = self.boxes.get_receipt_by_id(target_receipt_id)
        if receipt is None:
            raise bad_request("target_receipt_not_found")
        touched_receipt_ids = {box.warehouse_receipt_id for box in boxes if box.warehouse_receipt_id is not None}
        for box in boxes:
            box.warehouse_receipt_id = receipt.id
            box.current_waybill_id = receipt.waybill_id
            box.status = "bound" if receipt.waybill_id is not None else ("prebooked" if receipt.prebooking_id is not None else "unbound")
            box.never_bound_direct_upload = False
            box.unbound_reason = None
            box.unbound_remark = None
        self._refresh_receipt_totals(receipt)
        for receipt_id in touched_receipt_ids:
            if receipt_id != receipt.id:
                self._refresh_receipt_totals(self.boxes.get_receipt_by_id(receipt_id))
        self.db.commit()
        return BoxBatchOperationResult(updated_count=len(boxes), boxes=self.boxes.list_by_ids([box.id for box in boxes]))

    def _move_boxes_to_unbound(
        self,
        boxes: list[Box],
        *,
        unbound_reason: str | None,
        unbound_remark: str | None,
    ) -> BoxBatchOperationResult:
        touched_receipt_ids = {box.warehouse_receipt_id for box in boxes if box.warehouse_receipt_id is not None}
        remark = unbound_remark.strip() if unbound_remark else None
        for box in boxes:
            box.warehouse_receipt_id = None
            box.current_waybill_id = None
            box.status = "unbound"
            box.never_bound_direct_upload = False
            box.unbound_reason = unbound_reason
            box.unbound_remark = remark
        for receipt_id in touched_receipt_ids:
            self._refresh_receipt_totals(self.boxes.get_receipt_by_id(receipt_id))
        self.db.commit()
        return BoxBatchOperationResult(updated_count=len(boxes), boxes=self.boxes.list_by_ids([box.id for box in boxes]))

    def _load_boxes_for_batch(self, box_ids: list[int]) -> list[Box]:
        unique_ids = list(dict.fromkeys(box_ids))
        boxes = self.boxes.list_by_ids(unique_ids)
        if len(boxes) != len(unique_ids):
            raise bad_request("box_not_found")
        return boxes

    def _upload_conflicts(
        self,
        parsed_boxes: list[ParsedWarehouseBox],
        target_waybill: AirWaybill,
        target_warehouse_no: str,
        forced_box_nos: set[str],
    ) -> list[WarehouseBoxConflict]:
        parsed_box_nos = [item.box_no for item in parsed_boxes]
        conflicts: list[WarehouseBoxConflict] = []
        for box, current_waybill, current_receipt in self.boxes.list_conflicting_boxes(parsed_box_nos, target_warehouse_no):
            if box.box_no in forced_box_nos:
                continue
            conflicts.append(
                WarehouseBoxConflict(
                    box_no=box.box_no,
                    current_waybill_id=current_waybill.id if current_waybill else None,
                    current_waybill_no=current_waybill.waybill_no if current_waybill else None,
                    current_warehouse_no=current_receipt.warehouse_no if current_receipt else None,
                    target_waybill_id=target_waybill.id,
                    target_waybill_no=target_waybill.waybill_no,
                    target_warehouse_no=target_warehouse_no,
                )
            )
        return conflicts

    def _apply_conflict_box_renames(
        self,
        parsed_boxes: list[ParsedWarehouseBox],
        *,
        target_receipt: WarehouseReceipt | None,
        target_warehouse_no: str,
    ) -> None:
        if not parsed_boxes:
            return

        source_rows = self.boxes.list_conflicting_boxes([item.box_no for item in parsed_boxes], target_warehouse_no)
        if not source_rows:
            return

        source_by_box_no: dict[str, tuple[Box, AirWaybill | None, WarehouseReceipt | None]] = {}
        target_receipt_id = target_receipt.id if target_receipt is not None else None
        for box, current_waybill, current_receipt in source_rows:
            box_receipt_id = getattr(box, "warehouse_receipt_id", None)
            box_waybill_id = getattr(box, "current_waybill_id", None)
            if box_receipt_id is None and current_receipt is not None:
                box_receipt_id = current_receipt.id
            if box_waybill_id is None and current_waybill is not None:
                box_waybill_id = current_waybill.id
            if target_receipt_id is not None and box_receipt_id == target_receipt_id:
                continue
            if box_receipt_id is None and box_waybill_id is None:
                continue
            source_by_box_no.setdefault(box.box_no, (box, current_waybill, current_receipt))
        if not source_by_box_no:
            return

        reusable_by_original: dict[str, Box] = {}
        if target_receipt_id is not None:
            for box in self.boxes.list_by_receipt_id(target_receipt_id):
                conflict = _box_conflict_from_raw(box.raw_data)
                original_box_no = conflict.get("original_box_no") if conflict else None
                if isinstance(original_box_no, str) and original_box_no:
                    reusable_by_original.setdefault(original_box_no, box)

        reserved_names = {item.box_no for item in parsed_boxes}
        for parsed in parsed_boxes:
            source_row = source_by_box_no.get(parsed.box_no)
            if source_row is None:
                continue
            original_box_no = parsed.box_no
            reusable = reusable_by_original.get(original_box_no)
            if reusable is not None:
                renamed_box_no = reusable.box_no
                conflict_info = _box_conflict_from_raw(reusable.raw_data) or self._build_box_conflict_info(
                    original_box_no,
                    renamed_box_no,
                    *source_row,
                )
            else:
                renamed_box_no = self._next_conflict_box_no(original_box_no, reserved_names)
                conflict_info = self._build_box_conflict_info(original_box_no, renamed_box_no, *source_row)

            parsed.box_no = renamed_box_no
            parsed.raw_data = {**(parsed.raw_data or {}), BOX_CONFLICT_RAW_KEY: conflict_info}
            for item in parsed.items:
                item.raw_data = {**(item.raw_data or {}), BOX_CONFLICT_RAW_KEY: conflict_info}
            reserved_names.add(renamed_box_no)

    def _next_conflict_box_no(self, original_box_no: str, reserved_names: set[str]) -> str:
        index = 1
        while True:
            suffix = f"-DUP{index}"
            candidate = f"{original_box_no[: BOX_NO_MAX_LENGTH - len(suffix)]}{suffix}"
            if candidate not in reserved_names and self.boxes.get_by_box_no(candidate) is None:
                return candidate
            index += 1

    def _build_box_conflict_info(
        self,
        original_box_no: str,
        renamed_box_no: str,
        source_box: Box,
        source_waybill: AirWaybill | None,
        source_receipt: WarehouseReceipt | None,
    ) -> dict[str, Any]:
        source_document = getattr(source_receipt, "source_document", None) if source_receipt is not None else None
        return {
            "original_box_no": original_box_no,
            "renamed_box_no": renamed_box_no,
            "waybill_id": source_waybill.id if source_waybill else getattr(source_box, "current_waybill_id", None),
            "waybill_no": source_waybill.waybill_no if source_waybill else None,
            "warehouse_receipt_id": source_receipt.id if source_receipt else getattr(source_box, "warehouse_receipt_id", None),
            "warehouse_no": source_receipt.warehouse_no if source_receipt else None,
            "source_file_name": source_document.file_name if source_document else None,
        }

    def _ensure_receipt(
        self,
        warehouse_no: str,
        waybill: AirWaybill,
        *,
        document: BoxDocument | None,
        current_user: User,
    ) -> WarehouseReceipt:
        receipt = self.boxes.get_receipt_by_warehouse_no(warehouse_no)
        if receipt is not None and receipt.waybill_id not in (None, waybill.id):
            raise bad_request("warehouse_receipt_bound_to_other_waybill")
        if receipt is not None and receipt.prebooking_id is not None:
            raise bad_request("warehouse_receipt_bound_to_prebooking")
        if receipt is None:
            receipt = self.boxes.add_receipt(
                WarehouseReceipt(
                    warehouse_no=warehouse_no,
                    waybill_id=waybill.id,
                    source_document_id=document.id if document else None,
                    uploaded_by=current_user.id,
                    total_quantity=0,
                    total_weight=Decimal("0.000"),
                    total_volume=Decimal("0.000"),
                    weight_volume_ratio=Decimal("0.000"),
                )
            )
        else:
            receipt.waybill_id = waybill.id
            receipt.prebooking_id = None
            receipt.display_order = None
            if document is not None:
                receipt.source_document_id = document.id
                receipt.uploaded_by = current_user.id
        return receipt

    def _resolve_waybill_receipt_for_manual_box(
        self,
        waybill: AirWaybill,
        warehouse_receipt_id: int | None,
        current_user: User,
    ) -> WarehouseReceipt:
        if warehouse_receipt_id is not None:
            receipt = self.boxes.get_receipt_by_id(warehouse_receipt_id)
            if receipt is None or receipt.waybill_id != waybill.id:
                raise bad_request("warehouse_receipt_not_found")
            return receipt
        warehouse_no = (waybill.warehouse_no or "").strip()
        if not warehouse_no:
            raise bad_request("target_warehouse_no_required")
        return self._ensure_receipt(warehouse_no, waybill, document=None, current_user=current_user)

    def _resolve_prebooking_receipt_for_manual_box(
        self,
        prebooking: WaybillPrebooking,
        warehouse_receipt_id: int | None,
    ) -> WarehouseReceipt:
        if warehouse_receipt_id is None:
            raise bad_request("target_warehouse_receipt_required")
        receipt = self.boxes.get_receipt_by_id(warehouse_receipt_id)
        if receipt is None or receipt.prebooking_id != prebooking.id:
            raise bad_request("warehouse_receipt_not_found")
        return receipt

    def _unbind_previous_waybill_receipt_if_needed(self, waybill: AirWaybill, next_warehouse_no: str) -> None:
        if not waybill.warehouse_no or waybill.warehouse_no == next_warehouse_no:
            return
        previous = self.boxes.get_receipt_by_warehouse_no(waybill.warehouse_no)
        if previous is None or previous.waybill_id != waybill.id:
            return
        previous_boxes = self.boxes.list_by_receipt_id(previous.id)
        for box in previous_boxes:
            box.warehouse_receipt_id = None
            box.current_waybill_id = None
            box.status = "unbound"
            box.never_bound_direct_upload = False
            box.unbound_reason = None
            box.unbound_remark = None
        previous.waybill_id = None
        self._refresh_receipt_totals(previous)

    def _latest_bound_receipt_warehouse_no(self, waybill_id: int) -> str | None:
        receipt = self.db.scalar(
            select(WarehouseReceipt)
            .where(WarehouseReceipt.waybill_id == waybill_id)
            .order_by(WarehouseReceipt.updated_at.desc(), WarehouseReceipt.id.desc())
        )
        return receipt.warehouse_no if receipt is not None else None

    def _sync_receipt_boxes(
        self,
        receipt: WarehouseReceipt,
        waybill: AirWaybill,
        document: BoxDocument,
        parsed_boxes: list[ParsedWarehouseBox],
    ) -> set[int]:
        touched_receipt_ids: set[int] = set()
        parsed_by_no = {item.box_no: item for item in parsed_boxes}
        existing_in_receipt = self.boxes.list_by_receipt_id(receipt.id)
        for existing in existing_in_receipt:
            if existing.box_no not in parsed_by_no:
                if existing.warehouse_receipt_id is not None:
                    touched_receipt_ids.add(existing.warehouse_receipt_id)
                existing.warehouse_receipt_id = None
                existing.current_waybill_id = None
                existing.status = "unbound"
                existing.never_bound_direct_upload = False
                existing.unbound_reason = None
                existing.unbound_remark = None

        existing_by_no = {box.box_no: box for box in self.boxes.list_by_box_nos(list(parsed_by_no))}
        for parsed in parsed_boxes:
            box = existing_by_no.get(parsed.box_no)
            if box is None:
                box = Box(box_no=parsed.box_no)
                self.db.add(box)
                self.db.flush()
            elif box.warehouse_receipt_id is not None:
                touched_receipt_ids.add(box.warehouse_receipt_id)
            box.document_id = document.id
            box.warehouse_receipt_id = receipt.id
            box.current_waybill_id = waybill.id
            box.warehouse_waybill_no = parsed.warehouse_waybill_no
            box.goods_name = parsed.goods_name
            box.quantity = parsed.quantity
            box.weight = parsed.weight
            box.original_volume_info = parsed.original_volume_info
            box.original_weight_volume_ratio = parsed.original_weight_volume_ratio
            box.volume = parsed.volume
            box.weight_volume_ratio = parsed.weight_volume_ratio
            box.is_general_cargo = parsed.is_general_cargo
            box.source_row_number = parsed.source_row_number
            box.status = "bound"
            box.never_bound_direct_upload = False
            box.unbound_reason = None
            box.unbound_remark = None
            box.raw_data = parsed.raw_data
            self.db.flush()

            self.boxes.delete_items_for_box(box.id)
            self.boxes.add_items(
                [
                    BoxItem(
                        box_id=box.id,
                        document_id=document.id,
                        warehouse_waybill_no=item.warehouse_waybill_no,
                        goods_name=item.goods_name,
                        quantity=item.quantity,
                        weight=item.weight,
                        source_row_number=item.source_row_number,
                        raw_data=item.raw_data,
                    )
                    for item in parsed.items
                ]
            )
        return touched_receipt_ids

    def _sync_prebooking_receipt_boxes(
        self,
        receipt: WarehouseReceipt,
        document: BoxDocument,
        parsed_boxes: list[ParsedWarehouseBox],
    ) -> set[int]:
        touched_receipt_ids: set[int] = set()
        parsed_by_no = {item.box_no: item for item in parsed_boxes}
        existing_in_receipt = self.boxes.list_by_receipt_id(receipt.id)
        for existing in existing_in_receipt:
            if existing.box_no not in parsed_by_no:
                existing.warehouse_receipt_id = None
                existing.current_waybill_id = None
                existing.status = "unbound"
                existing.never_bound_direct_upload = False
                existing.unbound_reason = None
                existing.unbound_remark = None

        existing_by_no = {box.box_no: box for box in self.boxes.list_by_box_nos(list(parsed_by_no))}
        for parsed in parsed_boxes:
            box = existing_by_no.get(parsed.box_no)
            if box is None:
                box = Box(box_no=parsed.box_no)
                self.db.add(box)
                self.db.flush()
            elif box.warehouse_receipt_id is not None:
                touched_receipt_ids.add(box.warehouse_receipt_id)
            box.document_id = document.id
            box.warehouse_receipt_id = receipt.id
            box.current_waybill_id = None
            box.warehouse_waybill_no = parsed.warehouse_waybill_no
            box.goods_name = parsed.goods_name
            box.quantity = parsed.quantity
            box.weight = parsed.weight
            box.original_volume_info = parsed.original_volume_info
            box.original_weight_volume_ratio = parsed.original_weight_volume_ratio
            box.volume = parsed.volume
            box.weight_volume_ratio = parsed.weight_volume_ratio
            box.is_general_cargo = parsed.is_general_cargo
            box.source_row_number = parsed.source_row_number
            box.status = "prebooked"
            box.never_bound_direct_upload = False
            box.unbound_reason = None
            box.unbound_remark = None
            box.raw_data = {**(parsed.raw_data or {}), "source": "prebooking_receipt_upload"}
            self.db.flush()

            self.boxes.delete_items_for_box(box.id)
            self.boxes.add_items(
                [
                    BoxItem(
                        box_id=box.id,
                        document_id=document.id,
                        warehouse_waybill_no=item.warehouse_waybill_no,
                        goods_name=item.goods_name,
                        quantity=item.quantity,
                        weight=item.weight,
                        source_row_number=item.source_row_number,
                        raw_data=item.raw_data,
                    )
                    for item in parsed.items
                ]
            )
        return touched_receipt_ids

    def _sync_unbound_receipt_boxes(
        self,
        receipt: WarehouseReceipt,
        document: BoxDocument,
        parsed_boxes: list[ParsedWarehouseBox],
    ) -> set[int]:
        touched_receipt_ids: set[int] = set()
        parsed_by_no = {item.box_no: item for item in parsed_boxes}
        existing_in_receipt = self.boxes.list_by_receipt_id(receipt.id)
        for existing in existing_in_receipt:
            if existing.box_no not in parsed_by_no:
                existing.warehouse_receipt_id = None
                existing.current_waybill_id = None
                existing.status = "unbound"
                existing.never_bound_direct_upload = False
                existing.unbound_reason = None
                existing.unbound_remark = None

        existing_by_no = {box.box_no: box for box in self.boxes.list_by_box_nos(list(parsed_by_no))}
        for parsed in parsed_boxes:
            box = existing_by_no.get(parsed.box_no)
            if box is None:
                box = Box(box_no=parsed.box_no)
                self.db.add(box)
                self.db.flush()
            elif box.warehouse_receipt_id is not None:
                touched_receipt_ids.add(box.warehouse_receipt_id)
            box.document_id = document.id
            box.warehouse_receipt_id = receipt.id
            box.current_waybill_id = None
            box.warehouse_waybill_no = parsed.warehouse_waybill_no
            box.goods_name = parsed.goods_name
            box.quantity = parsed.quantity
            box.weight = parsed.weight
            box.original_volume_info = parsed.original_volume_info
            box.original_weight_volume_ratio = parsed.original_weight_volume_ratio
            box.volume = parsed.volume
            box.weight_volume_ratio = parsed.weight_volume_ratio
            box.is_general_cargo = parsed.is_general_cargo
            box.source_row_number = parsed.source_row_number
            box.status = "unbound"
            box.never_bound_direct_upload = True
            box.unbound_reason = None
            box.unbound_remark = None
            box.raw_data = {**(parsed.raw_data or {}), "source": "unbound_receipt_upload"}
            self.db.flush()

            self.boxes.delete_items_for_box(box.id)
            self.boxes.add_items(
                [
                    BoxItem(
                        box_id=box.id,
                        document_id=document.id,
                        warehouse_waybill_no=item.warehouse_waybill_no,
                        goods_name=item.goods_name,
                        quantity=item.quantity,
                        weight=item.weight,
                        source_row_number=item.source_row_number,
                        raw_data=item.raw_data,
                    )
                    for item in parsed.items
                ]
            )
        return touched_receipt_ids

    def _receipt_list_out(
        self,
        receipt: WarehouseReceipt,
        waybill_no: str | None,
        prebooking_id: int | None,
        prebooking_status: str | None,
        prebooking_planned_flight_date: date | None,
        source_file_name: str | None,
        source_uploaded_at: datetime | None,
        box_count: int,
        general_cargo_count: int,
    ) -> WarehouseReceiptListOut:
        prebooking_label = None
        if prebooking_id is not None:
            prebooking_label = f"预排仓 #{prebooking_id} / {prebooking_planned_flight_date}"
        return WarehouseReceiptListOut(
            id=receipt.id,
            warehouse_no=receipt.warehouse_no,
            waybill_id=receipt.waybill_id,
            waybill_no=waybill_no,
            prebooking_id=prebooking_id,
            prebooking_status=prebooking_status,
            prebooking_label=prebooking_label,
            source_document_id=receipt.source_document_id,
            source_file_name=source_file_name,
            uploaded_by=receipt.uploaded_by,
            total_quantity=receipt.total_quantity,
            total_weight=receipt.total_weight,
            total_volume=receipt.total_volume,
            weight_volume_ratio=receipt.weight_volume_ratio,
            channel_tags=list(receipt.channel_tags or []),
            box_count=box_count,
            general_cargo_count=general_cargo_count,
            display_order=receipt.display_order,
            uploaded_at=source_uploaded_at or receipt.created_at,
            created_at=receipt.created_at,
            updated_at=receipt.updated_at,
        )

    def _refresh_receipt_totals(self, receipt: WarehouseReceipt | None) -> None:
        if receipt is None:
            return
        boxes = self.boxes.list_by_receipt_id(receipt.id)
        total_quantity = sum(item.quantity or 0 for item in boxes)
        total_weight = sum((item.weight or Decimal("0.000") for item in boxes), Decimal("0.000"))
        total_volume = sum((item.volume or Decimal("0.000") for item in boxes), Decimal("0.000"))
        receipt.total_quantity = total_quantity
        receipt.total_weight = total_weight.quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
        receipt.total_volume = total_volume.quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
        receipt.weight_volume_ratio = (
            (receipt.total_weight / receipt.total_volume).quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
            if receipt.total_volume and receipt.total_volume > 0
            else Decimal("0.000")
        )
        receipt.channel_tags = compute_warehouse_receipt_channel_tags([item.box_no for item in boxes])

    def _store_file(self, file_name: str, file_hash: str, content: bytes) -> Path:
        storage_dir = Path(settings.warehouse_file_storage_dir)
        storage_dir.mkdir(parents=True, exist_ok=True)
        suffix = Path(file_name).suffix.lower()
        safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(file_name).stem).strip("._") or "warehouse-file"
        path = storage_dir / f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{file_hash[:12]}_{safe_name}{suffix}"
        path.write_bytes(content)
        return path


def parse_warehouse_xlsx(file_name: str, content: bytes) -> WarehouseFileParseResult:
    if Path(file_name).suffix.lower() != ".xlsx":
        raise bad_request("warehouse_file_only_xlsx_supported")
    if not content:
        raise bad_request("warehouse_file_empty")

    try:
        workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
    except Exception as exc:
        raise bad_request("warehouse_file_invalid_xlsx") from exc

    worksheet = workbook.active
    rows = worksheet.iter_rows()
    header_row_number = 0
    header_values: list[Any] = []
    for row_number, row in enumerate(rows, start=1):
        row_values = [cell.value for cell in row]
        if any(_clean_text(value) for value in row_values):
            header_row_number = row_number
            header_values = row_values
            break
    if not header_values:
        raise bad_request("warehouse_file_missing_header")

    column_map = _build_column_map(header_values)
    missing = [field for field in REQUIRED_COLUMNS if field != "quantity" and field not in column_map]
    if missing:
        raise bad_request(f"warehouse_file_missing_columns:{','.join(missing)}")

    boxes_by_no: dict[str, ParsedWarehouseBox] = {}
    box_order: list[str] = []
    errors: list[WarehouseFileImportError] = []
    barcode_cells: list[WarehouseUploadIntegrityIssue] = []
    skipped_count = 0
    normalized_headers = [_clean_text(value) or f"column_{idx + 1}" for idx, value in enumerate(header_values)]
    last_valid_box_no: str | None = None

    for row_number, row in enumerate(rows, start=header_row_number + 1):
        row_cells = list(row)
        values = [cell.value for cell in row_cells]
        if not any(_clean_text(value) for value in values):
            skipped_count += 1
            continue

        raw_data = {
            normalized_headers[idx]: _raw_json_value(values[idx] if idx < len(values) else None)
            for idx in range(len(normalized_headers))
        }
        try:
            raw_box_no = _optional_text(values, column_map["outer_barcode"])
            if raw_box_no:
                barcode_cells.append(
                    WarehouseUploadIntegrityIssue(
                        row_number=row_number,
                        box_no=raw_box_no,
                        message="",
                    )
                )
            box_no = raw_box_no or last_valid_box_no
            if not box_no:
                raise ValueError("外箱条码不能为空")
            is_continuation_row = raw_box_no is None
            warehouse_waybill_no = _optional_text(values, column_map["warehouse_waybill_no"])
            goods_name = _optional_text(values, column_map["goods_name"])
            quantity = _parse_quantity(values, column_map.get("quantity"))
            weight = _parse_decimal_cell(values, column_map["weight"], "重量")
            original_volume_info = _optional_text(values, column_map["volume"])
            original_weight_volume_ratio = _optional_column_text(values, column_map, "original_weight_volume_ratio")
            volume = _parse_volume_cell(values, column_map["volume"], allow_empty=is_continuation_row)
            if not is_continuation_row and volume <= 0:
                raise ValueError("收货体积信息必须大于 0")
        except ValueError as exc:
            errors.append(WarehouseFileImportError(row_number=row_number, message=str(exc)))
            continue

        last_valid_box_no = box_no
        is_general_cargo = _row_marks_general_cargo(row_cells, values, column_map)
        item = ParsedWarehouseBoxItem(
            warehouse_waybill_no=warehouse_waybill_no,
            goods_name=goods_name,
            quantity=quantity,
            weight=weight,
            source_row_number=row_number,
            raw_data=raw_data,
        )
        parsed_box = boxes_by_no.get(box_no)
        if parsed_box is None:
            parsed_box = ParsedWarehouseBox(
                box_no=box_no,
                warehouse_waybill_no=warehouse_waybill_no,
                goods_name=goods_name,
                quantity=None,
                weight=Decimal("0.000"),
                original_volume_info=original_volume_info,
                original_weight_volume_ratio=original_weight_volume_ratio,
                volume=volume,
                weight_volume_ratio=Decimal("0.000"),
                source_row_number=row_number,
                is_general_cargo=is_general_cargo,
                raw_data=raw_data,
                items=[],
            )
            boxes_by_no[box_no] = parsed_box
            box_order.append(box_no)

        parsed_box.items.append(item)
        parsed_box.is_general_cargo = parsed_box.is_general_cargo or is_general_cargo
        if quantity is not None:
            parsed_box.quantity = (parsed_box.quantity or 0) + quantity
        parsed_box.weight = (parsed_box.weight + weight).quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
        if parsed_box.original_volume_info is None and original_volume_info:
            parsed_box.original_volume_info = original_volume_info
        if parsed_box.original_weight_volume_ratio is None and original_weight_volume_ratio:
            parsed_box.original_weight_volume_ratio = original_weight_volume_ratio
        if parsed_box.volume <= 0 and volume > 0:
            parsed_box.volume = volume
        if parsed_box.warehouse_waybill_no is None and warehouse_waybill_no:
            parsed_box.warehouse_waybill_no = warehouse_waybill_no
        if parsed_box.goods_name is None and goods_name:
            parsed_box.goods_name = goods_name
        parsed_box.weight_volume_ratio = (
            (parsed_box.weight / parsed_box.volume).quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
            if parsed_box.volume > 0
            else Decimal("0.000")
        )

    return WarehouseFileParseResult(
        boxes=[boxes_by_no[box_no] for box_no in box_order],
        skipped_count=skipped_count,
        errors=errors,
        barcode_cells=barcode_cells,
    )


def warehouse_upload_integrity_issues(
    parse_result: WarehouseFileParseResult,
    uploaded_box_nos: list[str],
) -> list[WarehouseUploadIntegrityIssue]:
    uploaded_counter = Counter(uploaded_box_nos)
    issues: list[WarehouseUploadIntegrityIssue] = []
    for cell in parse_result.barcode_cells:
        if uploaded_counter[cell.box_no] > 0:
            uploaded_counter[cell.box_no] -= 1
            continue
        issues.append(
            WarehouseUploadIntegrityIssue(
                row_number=cell.row_number,
                box_no=cell.box_no,
                message="该外箱条码未成功写入系统，请检查该行是否重复、数据格式是否错误或是否被跳过。",
            )
        )
    return issues


def warehouse_prohibited_goods_issues(
    boxes: list[ParsedWarehouseBox],
) -> list[WarehouseProhibitedGoodsIssue]:
    issues: list[WarehouseProhibitedGoodsIssue] = []
    for box in boxes:
        for item in box.items:
            goods_name = (item.goods_name or "").strip()
            if not goods_name:
                continue
            lower_goods_name = goods_name.lower()
            keyword = next(
                (
                    candidate
                    for candidate in PROHIBITED_GOODS_KEYWORDS
                    if (candidate.isascii() and candidate in lower_goods_name) or (not candidate.isascii() and candidate in goods_name)
                ),
                None,
            )
            if keyword is None:
                continue
            issues.append(
                WarehouseProhibitedGoodsIssue(
                    row_number=item.source_row_number,
                    box_no=box.box_no,
                    warehouse_waybill_no=item.warehouse_waybill_no,
                    goods_name=goods_name,
                    keyword=keyword,
                    message=f"品名包含违禁词“{keyword}”，请人工复核。",
                )
            )
    return issues


def assert_warehouse_upload_integrity(
    *,
    file_name: str,
    warehouse_no: str,
    parse_result: WarehouseFileParseResult,
    uploaded_box_nos: list[str],
) -> None:
    issues = warehouse_upload_integrity_issues(parse_result, uploaded_box_nos)
    expected_count = len(parse_result.barcode_cells)
    uploaded_count = len(uploaded_box_nos)
    if expected_count == uploaded_count and not issues:
        return
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail={
            "error_code": "warehouse_upload_integrity_failed",
            "message": "入仓文件外箱数量校验失败，请检查未成功上传的外箱行。",
            "file_name": file_name,
            "warehouse_no": warehouse_no,
            "expected_count": expected_count,
            "uploaded_count": uploaded_count,
            "issues": [item.model_dump() for item in issues],
        },
    )


def review_warehouse_file_channels(
    warehouse_no: str,
    file_name: str,
    boxes: list[ParsedWarehouseBox],
) -> WarehouseChannelReviewResult:
    issue_by_box_no: dict[str, WarehouseChannelReviewIssue] = {}
    issue_order: list[str] = []
    valid_rows: list[tuple[ParsedWarehouseBox, str]] = []

    def add_issue(box: ParsedWarehouseBox, prefix: str, reason: str, message: str) -> None:
        existing = issue_by_box_no.get(box.box_no)
        if existing is None:
            issue_by_box_no[box.box_no] = WarehouseChannelReviewIssue(
                box_no=box.box_no,
                prefix=prefix,
                reason=reason,
                message=message,
            )
            issue_order.append(box.box_no)
            return
        reasons = {item.strip() for item in existing.reason.split(",") if item.strip()}
        if reason not in reasons:
            existing.reason = f"{existing.reason},{reason}"
            existing.message = f"{existing.message}；{message}"

    for box in boxes:
        prefix = _channel_prefix(box.box_no)
        if prefix is None:
            display_prefix = (box.box_no or "").strip().upper()[:3]
            add_issue(box, display_prefix, "unknown_channel_prefix", "外箱条码前三位必须是可识别的渠道字母前缀。")
            continue
        if prefix not in EUROPE_CHANNEL_PREFIXES and prefix not in UK_CHANNEL_PREFIXES:
            add_issue(box, prefix, "unknown_channel_prefix", "外箱条码前三位不是已配置的欧洲或英国渠道前缀。")
            continue
        valid_rows.append((box, prefix))

    europe_rows = [(box, prefix) for box, prefix in valid_rows if prefix in EUROPE_CHANNEL_PREFIXES - DUAL_CHANNEL_PREFIXES]
    uk_rows = [(box, prefix) for box, prefix in valid_rows if prefix in UK_CHANNEL_PREFIXES - DUAL_CHANNEL_PREFIXES]
    warnings: list[str] = []

    if len(europe_rows) > len(uk_rows):
        detected_channel = "europe"
        for box, prefix in uk_rows:
            add_issue(box, prefix, "uk_box_in_europe_receipt", "当前文件欧洲渠道箱号较多，该英国渠道箱号不应混入同一入仓号。")
    elif len(uk_rows) > len(europe_rows):
        detected_channel = "uk"
        for box, prefix in europe_rows:
            add_issue(box, prefix, "europe_box_in_uk_receipt", "当前文件英国渠道箱号较多，该欧洲渠道箱号不应混入同一入仓号。")
    elif europe_rows and uk_rows:
        detected_channel = "mixed"
        for box, prefix in europe_rows + uk_rows:
            add_issue(box, prefix, "channel_tie_unresolved", "当前文件欧洲和英国渠道数量相同，无法判定该入仓号所属渠道。")
    else:
        detected_channel = "unknown"
        if valid_rows and all(prefix == "DPD" for _, prefix in valid_rows):
            warnings.append("dpd_only_channel_pending")

    valid_prefixes = {prefix for _, prefix in valid_rows}
    if "CTT" in valid_prefixes:
        for box, prefix in valid_rows:
            if prefix not in CTT_ALLOWED_PREFIXES:
                add_issue(box, prefix, "ctt_mix_not_allowed", "CTT 入仓号只允许搭配 CTT/FRE/ITE/DHL 渠道箱号。")

    if "NLE" in valid_prefixes:
        for box, prefix in valid_rows:
            if prefix not in NLE_ALLOWED_PREFIXES:
                add_issue(box, prefix, "nle_mix_not_allowed", "NLE 入仓号只允许搭配 NLE/DHL 渠道箱号。")

    if {"UPS", "FED"}.intersection(valid_prefixes):
        dhl_rows = [(box, prefix) for box, prefix in valid_rows if prefix == "DHL"]
        if len(dhl_rows) > len(valid_rows) / 2:
            for box, prefix in dhl_rows:
                add_issue(box, prefix, "dhl_ratio_too_high", "UPS/FED 入仓号中 DHL 箱号数量不能超过当前文件箱号总数的一半。")

    return WarehouseChannelReviewResult(
        review=WarehouseChannelReviewOut(
            detected_channel=detected_channel,
            warnings=warnings,
        ),
        issues=[issue_by_box_no[box_no] for box_no in issue_order],
    )


def compute_warehouse_receipt_channel_tags(box_nos: list[str | None]) -> list[str]:
    prefixes = [_channel_prefix(box_no) for box_no in box_nos if box_no]
    known_prefixes = [prefix for prefix in prefixes if prefix in EUROPE_CHANNEL_PREFIXES or prefix in UK_CHANNEL_PREFIXES]
    if not known_prefixes:
        return []

    if len(known_prefixes) == len(prefixes) and all(prefix == "CTT" for prefix in known_prefixes):
        return list(ALL_CTT_CHANNEL_TAGS)

    europe_count = sum(1 for prefix in known_prefixes if prefix in EUROPE_CHANNEL_PREFIXES - DUAL_CHANNEL_PREFIXES)
    uk_count = sum(1 for prefix in known_prefixes if prefix in UK_CHANNEL_PREFIXES - DUAL_CHANNEL_PREFIXES)
    if europe_count > uk_count:
        return list(EUROPE_CHANNEL_TAGS)
    if uk_count > europe_count:
        return list(UK_CHANNEL_TAGS)
    if europe_count == 0 and uk_count == 0 and all(prefix == "DPD" for prefix in known_prefixes):
        return list(EUROPE_CHANNEL_TAGS)
    return []


def _fit_integer_dimension_volumes_to_target(
    *,
    boxes: list[Box],
    dimensions_by_box_id: dict[int, tuple[Decimal, Decimal, Decimal]],
    base_volumes: dict[int, Decimal],
    target_min: Decimal,
    target_max: Decimal,
) -> _IntegerDimensionFit | None:
    target_min = max(Decimal("0.000"), target_min.quantize(DECIMAL_001, rounding=ROUND_HALF_UP))
    target_max = target_max.quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
    if target_max < target_min:
        return None

    base_total = sum((base_volumes[box.id] for box in boxes), Decimal("0.000")).quantize(
        DECIMAL_001,
        rounding=ROUND_HALF_UP,
    )
    if base_total <= 0:
        return None

    ideal_total = min(max(base_total, target_min), target_max)
    ratio = ideal_total / base_total
    candidate_lists: list[list[_IntegerDimensionCandidate]] = []
    for box in boxes:
        dimensions = dimensions_by_box_id.get(box.id)
        if dimensions is None:
            return None
        ideal_volume = (base_volumes[box.id] * ratio).quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
        candidates = _integer_dimension_candidates_for_box(
            dimensions=dimensions,
            base_volume=base_volumes[box.id],
            ideal_volume=ideal_volume,
            max_volume=target_max,
        )
        if not candidates:
            return None
        candidate_lists.append(candidates)

    min_units = _volume_units(target_min)
    max_units = _volume_units(target_max)
    suffix_min_units = [0] * (len(candidate_lists) + 1)
    suffix_max_units = [0] * (len(candidate_lists) + 1)
    for index in range(len(candidate_lists) - 1, -1, -1):
        suffix_min_units[index] = suffix_min_units[index + 1] + min(candidate.units for candidate in candidate_lists[index])
        suffix_max_units[index] = suffix_max_units[index + 1] + max(candidate.units for candidate in candidate_lists[index])
    if suffix_min_units[0] > max_units or suffix_max_units[0] < min_units:
        return None

    states: dict[int, tuple[float, tuple[int, ...]]] = {0: (0.0, ())}
    ideal_units = _volume_units(ideal_total)
    for index, candidates in enumerate(candidate_lists):
        next_states: dict[int, tuple[float, tuple[int, ...]]] = {}
        remaining_min_units = suffix_min_units[index + 1]
        remaining_max_units = suffix_max_units[index + 1]
        for current_units, (current_score, selected_indexes) in states.items():
            for candidate_index, candidate in enumerate(candidates):
                next_units = current_units + candidate.units
                if next_units + remaining_min_units > max_units:
                    continue
                if next_units + remaining_max_units < min_units:
                    continue
                next_score = current_score + candidate.score
                previous = next_states.get(next_units)
                if previous is None or next_score < previous[0]:
                    next_states[next_units] = (next_score, selected_indexes + (candidate_index,))
        if not next_states:
            return None
        if len(next_states) > MAX_DIMENSION_FIT_STATES:
            expected_units = round(ideal_units * ((index + 1) / len(candidate_lists)))
            next_states = _prune_dimension_fit_states(next_states, expected_units)
        states = next_states

    valid_states = [
        (units, state)
        for units, state in states.items()
        if min_units <= units <= max_units
    ]
    if not valid_states:
        return None

    best_units, (_best_score, selected_indexes) = min(
        valid_states,
        key=lambda item: (item[1][0] + (item[0] - min_units) * 0.00000001, item[0]),
    )
    selected_volumes: dict[int, Decimal] = {}
    selected_dimensions: dict[int, tuple[int, int, int]] = {}
    for box, candidates, candidate_index in zip(boxes, candidate_lists, selected_indexes, strict=True):
        candidate = candidates[candidate_index]
        selected_volumes[box.id] = candidate.volume
        selected_dimensions[box.id] = candidate.dimensions

    total_volume = sum(selected_volumes.values(), Decimal("0.000")).quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
    if _volume_units(total_volume) != best_units:
        total_volume = (Decimal(best_units) * DECIMAL_001).quantize(DECIMAL_001, rounding=ROUND_HALF_UP)
    return _IntegerDimensionFit(volumes=selected_volumes, dimensions=selected_dimensions, total_volume=total_volume)


def _integer_dimension_candidates_for_box(
    *,
    dimensions: tuple[Decimal, Decimal, Decimal],
    base_volume: Decimal,
    ideal_volume: Decimal,
    max_volume: Decimal,
) -> list[_IntegerDimensionCandidate]:
    dimension_volume = _volume_from_dimensions(dimensions)
    if dimension_volume <= 0 or ideal_volume <= 0:
        return []

    scale = Decimal(str(float(ideal_volume / dimension_volume) ** (1 / 3)))
    centers = tuple(max(Decimal("1"), dimension * scale) for dimension in dimensions)
    center_ints = tuple(_positive_int_dimension(center) for center in centers)
    original_ints = tuple(_positive_int_dimension(dimension) for dimension in dimensions)
    dimension_options: set[tuple[int, int, int]] = {original_ints, center_ints}

    rounded_ranges: list[range] = []
    for center in centers:
        floored = max(1, int(center.to_integral_value(rounding=ROUND_DOWN)))
        rounded_ranges.append(range(floored, floored + 2))
    dimension_options.update(tuple(items) for items in product(*rounded_ranges))

    for radius in (1, 2, 4, 6):
        ranges = [
            range(max(1, center - radius), center + radius + 1)
            for center in center_ints
        ]
        dimension_options.update(tuple(items) for items in product(*ranges))

    candidates_by_units: dict[int, _IntegerDimensionCandidate] = {}
    for candidate_dimensions in dimension_options:
        volume = _volume_from_integer_dimensions(candidate_dimensions)
        if volume <= 0 or volume > max_volume:
            continue
        units = _volume_units(volume)
        score = _integer_dimension_candidate_score(candidate_dimensions, dimensions, volume, base_volume, ideal_volume)
        candidate = _IntegerDimensionCandidate(
            dimensions=candidate_dimensions,
            volume=volume,
            units=units,
            score=score,
        )
        previous = candidates_by_units.get(units)
        if previous is None or candidate.score < previous.score:
            candidates_by_units[units] = candidate

    return sorted(
        candidates_by_units.values(),
        key=lambda item: (item.score, abs(item.volume - ideal_volume), abs(item.volume - base_volume)),
    )[:MAX_DIMENSION_CANDIDATES_PER_BOX]


def _integer_dimension_candidate_score(
    candidate_dimensions: tuple[int, int, int],
    original_dimensions: tuple[Decimal, Decimal, Decimal],
    volume: Decimal,
    base_volume: Decimal,
    ideal_volume: Decimal,
) -> float:
    dimension_score = 0.0
    changed_dimensions = 0
    for candidate, original in zip(candidate_dimensions, original_dimensions, strict=True):
        original_float = float(original)
        if original_float <= 0:
            continue
        relative_delta = (candidate - original_float) / original_float
        dimension_score += relative_delta * relative_delta
        if Decimal(candidate) != original:
            changed_dimensions += 1

    ideal_float = float(ideal_volume) if ideal_volume > 0 else 0.0
    ideal_score = 0.0
    if ideal_float > 0:
        ideal_delta = (float(volume) - ideal_float) / ideal_float
        ideal_score = ideal_delta * ideal_delta

    base_float = float(base_volume) if base_volume > 0 else 0.0
    base_score = 0.0
    if base_float > 0:
        base_delta = (float(volume) - base_float) / base_float
        base_score = base_delta * base_delta

    return dimension_score + ideal_score * 0.75 + base_score * 0.001 + changed_dimensions * 0.000001


def _prune_dimension_fit_states(
    states: dict[int, tuple[float, tuple[int, ...]]],
    expected_units: int,
) -> dict[int, tuple[float, tuple[int, ...]]]:
    ranked = sorted(
        states.items(),
        key=lambda item: (item[1][0] + abs(item[0] - expected_units) * 0.000001, item[1][0], item[0]),
    )
    return dict(ranked[:MAX_DIMENSION_FIT_STATES])


def _volume_units(value: Decimal) -> int:
    return int((value.quantize(DECIMAL_001, rounding=ROUND_HALF_UP) / DECIMAL_001).to_integral_value(rounding=ROUND_HALF_UP))


def _volume_from_integer_dimensions(dimensions: tuple[int, int, int]) -> Decimal:
    length, width, height = dimensions
    return (Decimal(length) * Decimal(width) * Decimal(height) / Decimal("1000000")).quantize(
        DECIMAL_001,
        rounding=ROUND_HALF_UP,
    )


def _positive_int_dimension(value: Decimal) -> int:
    return max(1, int(value.to_integral_value(rounding=ROUND_HALF_UP)))


def _channel_prefix(box_no: str | None) -> str | None:
    text = (box_no or "").strip().upper()
    if len(text) < 3:
        return None
    prefix = text[:3]
    return prefix if re.fullmatch(r"[A-Z]{3}", prefix) else None


def _box_original_volume(box: Box) -> Decimal:
    raw_data = box.raw_data or {}
    if isinstance(raw_data, dict):
        recalculation = raw_data.get("volume_recalculation")
        if isinstance(recalculation, dict):
            stored_base = _decimal_from_optional_value(recalculation.get("base_volume"))
            if stored_base is not None:
                return stored_base

    original_volume = _parse_original_volume_text(box.original_volume_info)
    if original_volume is not None:
        return original_volume
    return (box.volume or Decimal("0.000")).quantize(DECIMAL_001, rounding=ROUND_HALF_UP)


def _calculated_volume_info(box: Box, volume: Decimal) -> str | None:
    dimensions = _parse_dimensions_text(box.original_volume_info)
    if dimensions is None:
        return None
    original_volume = _volume_from_dimensions(dimensions)
    if original_volume <= 0 or volume <= 0:
        return None
    scale = Decimal(str(float(volume / original_volume) ** (1 / 3)))
    scaled_dimensions = tuple((item * scale).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) for item in dimensions)
    dimension_text = "*".join(_format_dimension_value(item) for item in scaled_dimensions)
    return f"{dimension_text}({_format_decimal_display(volume)})"


def _calculated_volume_info_from_dimensions(dimensions: tuple[int, int, int], volume: Decimal) -> str:
    dimension_text = "*".join(str(item) for item in dimensions)
    return f"{dimension_text}({_format_decimal_display(volume)})"


def _parse_dimensions_text(value: str | None) -> tuple[Decimal, Decimal, Decimal] | None:
    if not value:
        return None
    text = _clean_text(value).replace(",", "")
    dimension_match = DIMENSION_VOLUME_PATTERN.search(text)
    if not dimension_match:
        return None
    try:
        dimensions = tuple(Decimal(part) for part in dimension_match.groups())
    except (InvalidOperation, ValueError):
        return None
    if any(item <= 0 for item in dimensions):
        return None
    return dimensions


def _volume_from_dimensions(dimensions: tuple[Decimal, Decimal, Decimal]) -> Decimal:
    length, width, height = dimensions
    return (length * width * height / Decimal("1000000")).quantize(DECIMAL_001, rounding=ROUND_HALF_UP)


def _format_dimension_value(value: Decimal) -> str:
    text = format(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), "f").rstrip("0").rstrip(".")
    return text or "0"


def _format_decimal_display(value: Decimal) -> str:
    text = format(value.quantize(DECIMAL_001, rounding=ROUND_HALF_UP), "f").rstrip("0").rstrip(".")
    return text or "0"


def _parse_original_volume_text(value: str | None) -> Decimal | None:
    if not value:
        return None
    text = _clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        dimensions = _parse_dimensions_text(text)
        if dimensions is not None:
            decimal = _volume_from_dimensions(dimensions)
        else:
            decimal = _to_decimal(text, "original_volume_info")
    except (InvalidOperation, ValueError):
        return None
    if decimal < 0:
        return None
    return decimal.quantize(DECIMAL_001, rounding=ROUND_HALF_UP)


def _decimal_from_optional_value(value: Any) -> Decimal | None:
    if value is None:
        return None
    try:
        decimal = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
    if decimal < 0:
        return None
    return decimal.quantize(DECIMAL_001, rounding=ROUND_HALF_UP)


def _format_no_valid_rows_error(errors: list[WarehouseFileImportError]) -> str:
    if not errors:
        return "warehouse_file_no_valid_rows"
    details = "；".join(f"第 {item.row_number} 行（{item.message}）" for item in errors[:5])
    suffix = f"；另 {len(errors) - 5} 行" if len(errors) > 5 else ""
    return f"没有有效货物行，失败行：{details}{suffix}"


def _row_marks_general_cargo(row_cells: list[Any], values: list[Any], column_map: dict[str, int]) -> bool:
    return _row_has_general_cargo_fill(row_cells, values, column_map) or _row_has_general_cargo_marker(values, column_map)


def _row_has_general_cargo_marker(values: list[Any], column_map: dict[str, int]) -> bool:
    if not column_map:
        return False
    weight_volume_index = column_map.get("original_weight_volume_ratio")
    marker_index = (weight_volume_index if weight_volume_index is not None else max(column_map.values())) + 1
    marker_text = _clean_text(_value_at(values, marker_index))
    return any(marker in marker_text for marker in GENERAL_CARGO_MARKERS)


def _row_has_general_cargo_fill(row_cells: list[Any], values: list[Any], column_map: dict[str, int]) -> bool:
    effective_indices = [
        index
        for index in sorted(set(column_map.values()))
        if _clean_text(_value_at(values, index))
    ]
    if not effective_indices:
        return False
    return all(index < len(row_cells) and _cell_has_non_default_fill(row_cells[index]) for index in effective_indices)


def _cell_has_non_default_fill(cell: Any) -> bool:
    fill = getattr(cell, "fill", None)
    if fill is None:
        return False
    fill_type = getattr(fill, "fill_type", None) or getattr(fill, "patternType", None)
    return bool(fill_type and fill_type != "none")


def _build_column_map(header_values: list[Any]) -> dict[str, int]:
    normalized = [_normalize_header(value) for value in header_values]
    column_map: dict[str, int] = {}
    for field, aliases in {**REQUIRED_COLUMNS, **OPTIONAL_COLUMNS}.items():
        normalized_aliases = {_normalize_header(alias) for alias in aliases}
        for index, header in enumerate(normalized):
            if header in normalized_aliases:
                column_map[field] = index
                break
    return column_map


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", _clean_text(value).lower())


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _value_at(values: list[Any], index: int) -> Any:
    return values[index] if index < len(values) else None


def _required_text(values: list[Any], index: int, label: str) -> str:
    value = _optional_text(values, index)
    if not value:
        raise ValueError(f"{label}不能为空")
    return value


def _optional_text(values: list[Any], index: int) -> str | None:
    value = _clean_text(_value_at(values, index))
    return value or None


def _optional_column_text(values: list[Any], column_map: dict[str, int], field: str) -> str | None:
    index = column_map.get(field)
    return _optional_text(values, index) if index is not None else None


def _parse_quantity(values: list[Any], index: int | None) -> int | None:
    if index is None:
        return None
    raw = _value_at(values, index)
    if raw is None or _clean_text(raw) == "":
        return None
    decimal = _to_decimal(raw, "数量")
    if decimal <= 0:
        raise ValueError("数量必须大于 0")
    if decimal != decimal.to_integral_value():
        raise ValueError("数量必须是整数")
    return int(decimal)


def _parse_decimal_cell(values: list[Any], index: int, label: str) -> Decimal:
    decimal = _to_decimal(_value_at(values, index), label)
    if decimal < 0:
        raise ValueError(f"{label}不能小于 0")
    return decimal.quantize(DECIMAL_001, rounding=ROUND_HALF_UP)


def _parse_volume_cell(values: list[Any], index: int, allow_empty: bool = False) -> Decimal:
    raw = _value_at(values, index)
    if raw is None or _clean_text(raw) == "":
        if allow_empty:
            return Decimal("0.000")
        raise ValueError("收货体积信息不能为空")

    text = _clean_text(raw).replace(",", "")
    dimension_match = DIMENSION_VOLUME_PATTERN.search(text)
    if dimension_match:
        length, width, height = (Decimal(part) for part in dimension_match.groups())
        decimal = length * width * height / Decimal("1000000")
    else:
        decimal = _to_decimal(raw, "收货体积信息")

    if decimal < 0:
        raise ValueError("收货体积信息不能小于 0")
    return decimal.quantize(DECIMAL_001, rounding=ROUND_HALF_UP)


def _to_decimal(value: Any, label: str) -> Decimal:
    if value is None or _clean_text(value) == "":
        raise ValueError(f"{label}不能为空")
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        raw = str(value)
    else:
        text = _clean_text(value).replace(",", "")
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if not match:
            raise ValueError(f"{label}必须是有效数字")
        raw = match.group(0)
    try:
        return Decimal(raw)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{label}必须是有效数字") from exc


def _raw_json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _box_conflict_from_raw(raw_data: Any) -> dict[str, Any] | None:
    if not isinstance(raw_data, dict):
        return None
    conflict = raw_data.get(BOX_CONFLICT_RAW_KEY)
    return conflict if isinstance(conflict, dict) else None
